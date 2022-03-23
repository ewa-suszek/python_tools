import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
import requests
from requests_kerberos import HTTPKerberosAuth
import urllib3

# this is to ignore the ssl insecure warning as we are passing in 'verify=false'
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

def main():
    config_wb = openpyxl.load_workbook('Config.cfg.xlsx')
    output_wb = openpyxl.Workbook()
    create_metric(config_wb, output_wb)
    save_output_wb(output_wb)

def save_output_wb(wb):
    ts = get_filename_timestamp()
    filename = f'{ts}-Metrics.xlsx'
    wb.save(filename)
    print(f'COMPLETED: Saved the output file:{filename}')

def get_filename_timestamp():
    ts = str(datetime.now())
    ts = ts.replace(':', '_')
    return ts

def create_metric(config_wb, output_wb):
    projects = get_hsd_es_data(config_wb)
    update_config_file(config_wb,projects)
    generate_project_ws(output_wb, projects)
    generate_detail_ws(output_wb, projects)
    generate_summary_ws(output_wb, projects)

def get_hsd_es_data(config_wb):
    ws = config_wb['hsdes']
    projects = []
    column_disable = get_column(ws,'Disable')
    column_query_name = get_column(ws,'short name')
    column_query_id = get_column(ws,'Query id')
    for row in range (2,ws.max_row+1):
        disable = ws.cell(row=row, column=column_disable).value
        query_name = ws.cell(row=row, column=column_query_name).value
        query_id = str(ws.cell(row=row, column=column_query_id).value)
        if (disable == None and query_name != None and  query_id != None):
            project = get_query_data(query_id,query_name)
            projects.append(project)    
    return projects

def get_column(ws,column_header):
    for col in range (1, ws.max_column):
        value = ws.cell(row=1,column=col).value
        if column_header.lower() == value.lower():
            return col
    raise Exception(f'unable to find column:{column_header}')

def get_query_data(query_id,query_name):
    headers = { 'Content-type': 'application/json' }
    query_subject = get_query_subject(query_id)
    data_fields = set_data_fields_by_subject(query_subject)
    url = 'https://hsdes-api.intel.com/rest/query/'+query_id+'?include_text_fields=Y&start_at=1&fields='+data_fields+'&expand=metadata'
    response = requests.get(url, verify=False, auth=HTTPKerberosAuth(), headers = headers)
    results = response.json()
    column_names = results['data'][0].keys() # to do - if no data - no data[0].keys()!!!
    
    # use only user defined columns - system columns are starting from  hierarchy_id 
    # todo: when selecting fields - no need to check hierarchy_id - include all selected
    #print(column_names)
    #for key in results['data'][0].keys():
    #    if key == 'hierarchy_id':
    #        break    
    #    column_names.append(key)

    query_title = results['metadata']['title']
    query_link = 'https://hsdes.intel.com/appstore/community/#/query?queryId='+query_id
    project = {'title': query_name, 'data': results['data'], 'column_names': column_names, 'query_title': query_title,'query_link': query_link, 'query_subject': query_subject}
    return project

def get_query_subject(query_id):
    headers = { 'Content-type': 'application/json' }
    url = 'https://hsdes-api.intel.com/rest/query/'+query_id+'?include_text_fields=N&start_at=1&max_results=1&fields=id%2Ctenant%2Csubject&expand=metadata'
    response = requests.get(url, verify=False, auth=HTTPKerberosAuth(), headers = headers)
    results = response.json()
    query_subject = results['metadata']['query.parent_subject_list']
    return query_subject

def set_data_fields_by_subject(query_subject):
    bug_data_fields = 'id%2Ctitle%2Crelease%2Ccomponent%2Cstatus%2Creason%2Cpriority%2Cexposure%2Ctag%2Csubmitted_date%2Csubmitted_by%2Cupdated_date%2Cupdated_by%2Crev%2Ctenant%2Csubject'
    # todo: add more fields by subjects
    # now limited to bugs
    data_fields = bug_data_fields
    return data_fields

def update_config_file(config_wb,projects):
    ws = config_wb['hsdes']
    ws.cell(row=1, column=4, value='Query title')
    ws.cell(row=1, column=5, value='Query subject')
    for i in range(len(projects)):
        ws.cell(row=i+2, column=4, value=projects[i]['query_title'])
        ws.cell(row=i+2, column=4).style = "Hyperlink"
        ws.cell(row=i+2, column=4).hyperlink = projects[i]['query_link']
        ws.cell(row=i+2, column=5, value=projects[i]['query_subject'])
    config_wb.save('Config.cfg.xlsx')

def generate_project_ws(output_wb, projects):
    for project in projects:
        project_name = project['title']
        ws = output_wb.create_sheet(f'hsd_es-{project_name}')
        row=1
        col=0
        for column_name in project['column_names']:
            col+=1
            ws.cell(row=row, column=col, value=column_name)
        for result in project['data']:
            row+=1
            col=0
            for column_name in project['column_names']:
                col+=1
                ws.cell(row=row, column=col, value=result[column_name])
        ws.auto_filter.ref = ws.dimensions
        set_width(ws)

def set_width(ws):
    column_widths = []
    max_num_lines = []

    for row_cnt, row in enumerate(ws.iter_rows()):
        max_num_lines.append(1)

        for cnt, cell in enumerate(row):
            # Ignore cells that don't wrap
            if cell.alignment.wrap_text == False:
                continue
            if isinstance(cell.value, str):
                if 'HYPERLINK' in cell.value:
                    pieces = cell.value.split(',')
                    val = pieces[-1].split('"')[1]
                    length = len(val)
                else:
                    if '\n' in cell.value:
                        max_num_lines[row_cnt] = max(max_num_lines[row_cnt], len(cell.value.split('\n')))
                        length = max([len(s) for s in cell.value.split('\n')])
                    else:
                        length = len(cell.value)
            elif isinstance(cell.value, bool):
                length = len('False')
            elif isinstance(cell.value, int):
                if cell.value == 0:
                    length = 1
                else:
                    if cell.value < 0:
                        val = -1 * cell.value
                    else:
                        val = cell.value
                    length = math.ceil(math.log10(abs(cell.value)))
            else:
                length = 0
            length += 2
            if len(column_widths) > cnt:
                if length > column_widths[cnt]:
                    column_widths[cnt] = length
            else:
                column_widths += [length]

    max_column_width_allowed = 60
    for cnt, column_width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(cnt + 1)].width = min(column_width, max_column_width_allowed)

    row_height = 14
    for cnt, num_lines in enumerate(max_num_lines, 1):
        if num_lines > 1:
            ws.row_dimensions[1].height = row_height * num_lines

def generate_detail_ws(output_wb, projects):
    ws = output_wb.create_sheet('hsd_es-detail')
    ws.cell(row=1, column=1, value='detail')

def generate_summary_ws(output_wb, projects):
    ws = output_wb.create_sheet('hsd_es-summary')
    row=1
    ws.cell(row=row, column=2, value='priority')
    ws.cell(row=row, column=7, value='priority %')
    row+=1
    ws.cell(row=row, column=1, value='Query title')
    ws.cell(row=row, column=2, value='P1')
    ws.cell(row=row, column=3, value='P2')
    ws.cell(row=row, column=4, value='P3')
    ws.cell(row=row, column=5, value='P4')
    ws.cell(row=row, column=7, value='P1 %')
    ws.cell(row=row, column=8, value='P2 %')
    ws.cell(row=row, column=9, value='P3 %')
    ws.cell(row=row, column=10, value='P4 %')
        
    for i in range(len(projects)):
        row+=1
        ws.cell(row=row, column=1, value=projects[i]['query_title'])
        priority_stat = list()
        for bug in range(len(projects[i]['data'])):
             priority_stat.append(projects[i]['data'][bug]['priority'])
        ws.cell(row=row, column=2, value=priority_stat.count('p1-showstopper'))
        ws.cell(row=row, column=7, value=round(priority_stat.count('p1-showstopper')/len(priority_stat),4))
        ws.cell(row=row, column=7).style = 'Percent'
        ws.cell(row=row, column=3, value=priority_stat.count('p2-high'))
        ws.cell(row=row, column=8, value=round(priority_stat.count('p2-high')/len(priority_stat),4))
        ws.cell(row=row, column=8).style = 'Percent'
        ws.cell(row=row, column=4, value=priority_stat.count('p3-medium'))
        ws.cell(row=row, column=9, value=round(priority_stat.count('p3-medium')/len(priority_stat),4))
        ws.cell(row=row, column=9).style = 'Percent'
        ws.cell(row=row, column=5, value=priority_stat.count('p4-low'))
        ws.cell(row=row, column=10, value=round(priority_stat.count('p4-low')/len(priority_stat),4))
        ws.cell(row=row, column=10).style = 'Percent'

if __name__ == '__main__':
    main()
