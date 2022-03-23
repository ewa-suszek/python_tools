import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from datetime import datetime
import requests
from requests_kerberos import HTTPKerberosAuth
import urllib3
import json
import math

# this is to ignore the ssl insecure warning as we are passing in 'verify=false'
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

def main():
    config_wb = openpyxl.load_workbook('Config.cfg.xlsx')
    output_wb = openpyxl.Workbook()
    create_metric(config_wb, output_wb)
    save_output_wb(output_wb)

def save_output_wb(wb):
    #ts = get_filename_timestamp()
    #filename = f'{ts}-Metrics.xlsx' 
    filename = 'latest-Metrics.xlsx' # remove for final
    wb.save(filename)
    print(f'COMPLETED: Saved the output file:{filename}')

def get_filename_timestamp():
    ts = str(datetime.now())
    ts = ts.replace(':', '_')
    return ts

def create_metric(config_wb, output_wb):
    projects = get_hsd_es_data(config_wb)
    update_config_file(config_wb,projects)
    generate_project_ws(output_wb, projects)
    generate_project_metrics_ws(output_wb, projects)
    generate_detail_ws(output_wb, projects)
    generate_summary_ws(output_wb, projects)

def get_hsd_es_data(config_wb):
    ws = config_wb['hsdes']
    projects = []
    column_disable = get_column(ws,'Disable')
    column_query_name = get_column(ws,'short name')
    column_query_id = get_column(ws,'Query id')
    for row in range (2,ws.max_row+1):
        disable = ws.cell(row=row, column=column_disable).value
        query_name = ws.cell(row=row, column=column_query_name).value
        query_id = str(ws.cell(row=row, column=column_query_id).value)
        if (disable == None and query_name != None and  query_id != None):
            project = get_query_data(query_id,query_name)
            projects.append(project)    
    return projects

def get_column(ws,column_header):
    for col in range (1, ws.max_column):
        value = ws.cell(row=1,column=col).value
        if column_header.lower() == value.lower():
            return col
    raise Exception(f'unable to find column:{column_header}')

def get_query_data(query_id,query_name):
    query_info = get_query_info(query_id)
    query_subject = query_info['query_subject']
    query_fields = query_info['query_fields']

    query_xml = query_info['query_xml']
    query_title = query_info['query_title']

    query_data_fields = '%2C'.join(query_fields)
    query_data = get_xml_query_data(query_xml,query_data_fields) #todo: query_data_fields should not be needed to query - should be included in the query xml

    # add time calculations
    add_time_to_close(query_data,query_fields)
    add_time_since_last_update(query_data,query_fields)
    add_time_since_submitted_date(query_data,query_fields)

    
    query_link = 'https://hsdes.intel.com/appstore/community/#/query?queryId='+query_id
    project = {'title': query_name, 'data': query_data, 'column_names': query_fields, 'query_title': query_title,'query_link': query_link, 'query_subject': query_subject}
    return project

def get_query_info(query_id):
    headers = { 'Content-type': 'application/json' }
    url = 'https://hsdes-api.intel.com/rest/query/'+query_id+'?verbose=true&include_text_fields=Y&start_at=0&max_results=0&expand=metadata'
    response = requests.get(url, verify=False, auth=HTTPKerberosAuth(), headers = headers)
    results = response.json()
    query_subject = results['metadata']['query.parent_subject_list']
    query_title = results['metadata']['title']
    query_xml_user = results['metadata']['query.query_xml']
    column_names = set_data_fields_by_subject(query_subject)
    #todo parse query xml for existing fields and add missing
    # add user fields to colum names
    query_fields = column_names
    query_xml_fields = '<Display Displayas=\"shortname\"><DisplayField Visible=\"false\" Shortname=\"id\" Fullname=\"id\"/><DisplayField Visible=\"true\" Shortname=\"title\" Fullname=\"title\"/><DisplayField Visible=\"true\" Shortname=\"release\" Fullname=\"release\"/><DisplayField Visible=\"true\" Shortname=\"component\" Fullname=\"component\"/><DisplayField Visible=\"true\" Shortname=\"status\" Fullname=\"status\"/><DisplayField Visible=\"true\" Shortname=\"reason\" Fullname=\"reason\"/><DisplayField Visible=\"true\" Shortname=\"priority\" Fullname=\"priority\"/><DisplayField Visible=\"true\" Shortname=\"tag\" Fullname=\"tag\"/><DisplayField Visible=\"true\" Shortname=\"flags\" Fullname=\"server_rackscale.bug.flags\"/><DisplayField Visible=\"true\" Shortname=\"submitted_date\" Fullname=\"submitted_date\"/><DisplayField Visible=\"true\" Shortname=\"submitted_by\" Fullname=\"submitted_by\"/><DisplayField Visible=\"true\" Shortname=\"updated_date\" Fullname=\"updated_date\"/><DisplayField Visible=\"true\" Shortname=\"updated_by\" Fullname=\"updated_by\"/><DisplayField Visible=\"true\" Shortname=\"closed_date\" Fullname=\"closed_date\"/><DisplayField Visible=\"true\" Shortname=\"closed_reason\" Fullname=\"bug.closed_reason\"/><DisplayField Visible=\"true\" Shortname=\"open_date\" Fullname=\"bug.open_date\"/></Display><SortOrder/></Query>'
    fields_position = query_xml_user.find('<Display Displayas=')
    query_xml_custom = query_xml_user[0:fields_position] + query_xml_fields
    query_xml = '{"queryXml": '+json.dumps(query_xml_custom)+'}'
    query_info = {'query_subject':query_subject, 'query_title':query_title, 'query_xml':query_xml, 'query_fields': query_fields}
    return query_info


def get_xml_query_data(query_xml,query_fields):
    headers = { 'Content-type': 'application/json' }
    url = 'https://hsdes-api.intel.com/rest/query/execution?include_text_fields=Y&start_at=1&fields='+query_fields
    response = requests.post(url, data = query_xml, verify=False, auth=HTTPKerberosAuth(), headers = headers)
    results = response.json()
    return results['data']  


def get_record_id_history(id):
    headers = { 'Content-type': 'application/json' }
    url = 'https://hsdes-api.intel.com/rest/article/'+id+'/history'
    response = requests.get(url, verify=False, auth=HTTPKerberosAuth(), headers = headers)
    results = response.json()
    record_history = {'id': id, 'data': results['data']}
    return record_history  

def set_data_fields_by_subject(query_subject):
    bug_data_fields = ['id', 'title', 'release', 'component', 'status', 'reason', 'priority', 'tag', 'submitted_date', 'submitted_by', 'updated_date', 'updated_by', 'closed_date', 'closed_reason', 'open_date', 'rev', 'tenant', 'subject']
    # todo: add more fields by subjects
    # now limited to bugs
    data_fields = bug_data_fields
    return data_fields

def update_config_file(config_wb,projects):
    ws = config_wb['hsdes']
    ws.cell(row=1, column=4, value='Query title')
    ws.cell(row=1, column=5, value='Query subject')
    for i in range(len(projects)):
        ws.cell(row=i+2, column=4, value=projects[i]['query_title'])
        ws.cell(row=i+2, column=4).style = "Hyperlink"
        ws.cell(row=i+2, column=4).hyperlink = projects[i]['query_link']
        ws.cell(row=i+2, column=5, value=projects[i]['query_subject'])
    config_wb.save('Config.cfg.xlsx')

def generate_project_ws(output_wb, projects):
    for project in projects:
        project_name = project['title']
        ws = output_wb.create_sheet(f'{project_name}')
        row=1
        col=0
        for column_name in project['column_names']:
            col+=1
            ws.cell(row=row, column=col, value=column_name)
        for result in project['data']:
            row+=1
            col=0
            for column_name in project['column_names']:
                col+=1
                if column_name == 'id':
                    ws.cell(row=row, column=col).value = f'=HYPERLINK("https://hsdes.intel.com/appstore/article/#/{result[column_name]}","{result[column_name]}")'
                else:
                    ws.cell(row=row, column=col, value=result[column_name])
        ws.auto_filter.ref = ws.dimensions
        set_width(ws)

def set_width(ws):
    column_widths = []
    max_num_lines = []

    for row_cnt, row in enumerate(ws.iter_rows()):
        max_num_lines.append(1)

        for cnt, cell in enumerate(row):
            # Ignore cells that don't wrap
            if cell.alignment.wrap_text == False:
                continue
            if isinstance(cell.value, str):
                if 'HYPERLINK' in cell.value:
                    pieces = cell.value.split(',')
                    val = pieces[-1].split('"')[1]
                    length = len(val)
                else:
                    if '\n' in cell.value:
                        max_num_lines[row_cnt] = max(max_num_lines[row_cnt], len(cell.value.split('\n')))
                        length = max([len(s) for s in cell.value.split('\n')])
                    else:
                        length = len(cell.value)
            elif isinstance(cell.value, bool):
                length = len('False')
            elif isinstance(cell.value, int):
                if cell.value == 0:
                    length = 1
                else:
                    if cell.value < 0:
                        val = -1 * cell.value
                    else:
                        val = cell.value
                    length = math.ceil(math.log10(abs(cell.value)))
            else:
                length = 0
            length += 2
            if len(column_widths) > cnt:
                if length > column_widths[cnt]:
                    column_widths[cnt] = length
            else:
                column_widths += [length]

    max_column_width_allowed = 60
    for cnt, column_width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(cnt + 1)].width = min(column_width, max_column_width_allowed)

    row_height = 14
    for cnt, num_lines in enumerate(max_num_lines, 1):
        if num_lines > 1:
            ws.row_dimensions[1].height = row_height * num_lines

def generate_detail_ws(output_wb, projects):
    ws = output_wb.create_sheet('detail')
    ws.cell(row=1, column=1, value='detail')

def generate_summary_ws(output_wb, projects):
    ws = output_wb.create_sheet('summary')
    row=1
    ws.cell(row=row, column=2, value='priority')
    ws.cell(row=row, column=7, value='priority %')
    row+=1
    ws.cell(row=row, column=1, value='Query title')
    ws.cell(row=row, column=2, value='P1')
    ws.cell(row=row, column=3, value='P2')
    ws.cell(row=row, column=4, value='P3')
    ws.cell(row=row, column=5, value='P4')
    ws.cell(row=row, column=7, value='P1 %')
    ws.cell(row=row, column=8, value='P2 %')
    ws.cell(row=row, column=9, value='P3 %')
    ws.cell(row=row, column=10, value='P4 %')
        
    for i in range(len(projects)):
        row+=1
        ws.cell(row=row, column=1, value=projects[i]['query_title'])
        priority_stat = list()
        for bug in range(len(projects[i]['data'])):
             priority_stat.append(projects[i]['data'][bug]['priority'])
        ws.cell(row=row, column=2, value=priority_stat.count('p1-showstopper'))
        ws.cell(row=row, column=7, value=round(priority_stat.count('p1-showstopper')/len(priority_stat),4))
        ws.cell(row=row, column=7).style = 'Percent'
        ws.cell(row=row, column=3, value=priority_stat.count('p2-high'))
        ws.cell(row=row, column=8, value=round(priority_stat.count('p2-high')/len(priority_stat),4))
        ws.cell(row=row, column=8).style = 'Percent'
        ws.cell(row=row, column=4, value=priority_stat.count('p3-medium'))
        ws.cell(row=row, column=9, value=round(priority_stat.count('p3-medium')/len(priority_stat),4))
        ws.cell(row=row, column=9).style = 'Percent'
        ws.cell(row=row, column=5, value=priority_stat.count('p4-low'))
        ws.cell(row=row, column=10, value=round(priority_stat.count('p4-low')/len(priority_stat),4))
        ws.cell(row=row, column=10).style = 'Percent'
    chart_data = Reference(worksheet=ws,
                 min_row=2,
                 max_row=2+len(projects),
                 min_col=2,
                 max_col=5)
    chart_titles = Reference(ws, min_col=1, min_row=3, max_row=2+len(projects))

    add_bar_chart_priority(ws, chart_data, chart_titles, 'M2')

def add_bar_chart_priority(ws,data,titles,position):
    chart = BarChart()
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(titles)
    # Change bar filling and line color 
    chart.series[0].graphicalProperties.solidFill = "ff0000" 
    chart.series[1].graphicalProperties.solidFill = "ff8800" 
    chart.series[2].graphicalProperties.solidFill = "ffbb00" 
    chart.series[3].graphicalProperties.solidFill = "ffee00" 
    ws.add_chart(chart, position)


def add_time_to_close(dict_data_set,colum_names_list):
    for i in range(len(dict_data_set)):
        time_to_close = ''
        if len(dict_data_set[i]['closed_date']) > 1:
            submitted_date = datetime.strptime(dict_data_set[i]['submitted_date'], '%Y-%m-%d %H:%M:%S.%f')
            closed_date = datetime.strptime(dict_data_set[i]['closed_date'], '%Y-%m-%d %H:%M:%S.%f')
            time_to_close = (closed_date - submitted_date).days
        dict_data_set[i].update({'time_to_close':time_to_close})
    colum_names_list.append('time_to_close')


def add_time_since_last_update(dict_data_set,colum_names_list):
    now = datetime.now()
    for i in range(len(dict_data_set)):
        updated_date = datetime.strptime(dict_data_set[i]['updated_date'], '%Y-%m-%d %H:%M:%S.%f')
        time_since_last_update = (now - updated_date).days
        dict_data_set[i].update({'time_since_last_update':time_since_last_update})
    colum_names_list.append('time_since_last_update')

def add_time_since_submitted_date(dict_data_set,colum_names_list):
    now = datetime.now()
    for i in range(len(dict_data_set)):
        submitted_date = datetime.strptime(dict_data_set[i]['submitted_date'], '%Y-%m-%d %H:%M:%S.%f')
        time_since_submitted_date = (now - submitted_date).days
        dict_data_set[i].update({'time_since_submitted_date':time_since_submitted_date})
    colum_names_list.append('time_since_submitted_date')


def get_max_open_time(project_data):
    now = datetime.now()
    days_since_open_stat = list()
    for bug_data in project_data:
        if bug_data['status'] == 'open':
            open_date = datetime.strptime(bug_data['open_date'], '%Y-%m-%d %H:%M:%S.%f')
            days_since_open_stat.append((now - open_date).days)
    return max(days_since_open_stat)

def get_max_closed_time(project_data):
    time_to_close_stat = list()
    max_time_to_close = '---'
    for bug_data in project_data:
        if type(bug_data['time_to_close']) == int:
            time_to_close_stat.append(bug_data['time_to_close'])
    if len(time_to_close_stat)>0:
        max_time_to_close = max(time_to_close_stat)
    return max_time_to_close

def get_min_closed_time(project_data):
    time_to_close_stat = list()
    min_time_to_close = '---'
    for bug_data in project_data:
        if type(bug_data['time_to_close']) == int:
            time_to_close_stat.append(bug_data['time_to_close'])
    if len(time_to_close_stat)>0:
        min_time_to_close = min(time_to_close_stat)
    return min_time_to_close

def get_average_closed_time(project_data):
    time_to_close_count = 0
    total_time_to_close = 0
    average_closed_time = '---'
    for bug_data in project_data:
        if type(bug_data['time_to_close']) == int:
            if bug_data['time_to_close'] > 0:
                total_time_to_close += int(bug_data['time_to_close'])
                time_to_close_count += 1
    if time_to_close_count > 0:
        average_closed_time = round(total_time_to_close/time_to_close_count,4)
    return average_closed_time

def generate_project_metrics_ws(output_wb, projects):
    for project in projects:
        project_name = project['title']
        ws = output_wb.create_sheet(f'{project_name}_metric')
        row=1
        col=1
        ws.cell(row=row, column=col, value='max open time') 
        col+=1
        ws.cell(row=row, column=col, value='max close time') 
        col+=1
        ws.cell(row=row, column=col, value='min close time')
        col+=1
        ws.cell(row=row, column=col, value='average close time') 
        col=1
        row=2
        ws.cell(row=row, column=col, value=get_max_open_time(project['data'])) 
        col+=1
        ws.cell(row=row, column=col, value=get_max_closed_time(project['data'])) 
        col+=1
        ws.cell(row=row, column=col, value=get_min_closed_time(project['data'])) 
        col+=1
        ws.cell(row=row, column=col, value=get_average_closed_time(project['data'])) 
        set_width(ws)

if __name__ == '__main__':
    main()
