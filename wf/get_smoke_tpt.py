import sys
import math
import re
import os.path
import argparse
import pickle
import gzip
from datetime import datetime
from datetime import timedelta
from operator import itemgetter
import time
import json
import pandas as pd
import requests
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList

def main(status, count, date_from, date_to, dag_stream, submitter, filename, offline_mode=False):
    ci_cycle_stat = list()
    ci_test_sessions_stat = list()

    #GETING DATA
    cache_file1_name = f'cycles_master_tmp.cache'
    cache_file2_name = f'sessions_master_tmp.cache'
    cache_wf = False
    #cache_wf = True
    if cache_wf and os.path.isfile(cache_file1_name) and os.path.isfile(cache_file2_name):
        with open(cache_file1_name, 'rb') as cache:
            ci_cycle_stat = pickle.load(cache)
        with open(cache_file2_name, 'rb') as cache:
            ci_test_sessions_stat = pickle.load(cache)
    else:
        if dag_stream.find(':') > 0:
            dag_name = dag_stream.split(':')[0]
            stream_name = dag_stream.split(':')[1]
            update_stats(ci_cycle_stat, ci_test_sessions_stat, get_cycles(dag_name, stream_name, submitter, status=status, count=count, date_from=date_from, date_to=date_to))
            print(f'{len(ci_cycle_stat)} cycles and {len(ci_test_sessions_stat)} smoke sessions found')
            update_sessions_with_tp_data(ci_test_sessions_stat)
            update_sessions_with_gtax_data(ci_test_sessions_stat, offline_mode)
            with open(cache_file1_name, 'w+b') as cache:
                pickle.dump(ci_cycle_stat, cache) 
            with open(cache_file2_name, 'w+b') as cache:
                pickle.dump(ci_test_sessions_stat, cache)
        else:
            print('wrong dag stream format. use DAG:stream exp: CI_SMOKE__gfx-driver__master:gfx-driver__master')
    
    # EXCEL OUTPUT
    output_wb = openpyxl.Workbook()
    ws = output_wb.active
    ws.title = 'smoke-ci_cycles'
    cycle_columns = ['cycle_id', 'build_name', 'cycle_type', 'dag_id', 'status', 'WW', 'start_date', 'end_date', 'duration', 'build_start_date', 'started_after_build', 'lgcb_bad', 'has_regressions', 'comparison_view']
    print_data_table(ws, ci_cycle_stat, cycle_columns, 1, 1)
    ws.auto_filter.ref = ws.dimensions
    set_width(ws, max_width=600)
    ws = output_wb.create_sheet('smoke-ci_sessions')
    session_columns = ['cycle_id', 'gtax_id', 'build_name', 'cycle_type', 'environment', 'test_plan_id', 'test_plan_name', 'status', 'progress', 'WW', 'start_date', 'end_date', 'duration', 'max_job_duration', 'duration_min', 'tpt_met', 'remote_name', 'lgcb_bad', 'has_notes', 'has_regressions', 'has_code_issue', 'false_alarm', 'has_comment_of_failure_reasons', 'setup_success', 'test_best', 'test_worst', 'test_best_worst', 'comparison_view', 'failure_reasons', 'comments']
    #session_columns = ['cycle_id', 'gtax_id', 'build_name', 'cycle_type', 'environment', 'test_plan_id', 'test_plan_name', 'status', 'progress', 'WW', 'start_date', 'end_date', 'duration', 'max_job_duration', 'duration_min', 'tpt_met', 'remote_name', 'lgcb_bad', 'has_notes', 'has_regressions', 'has_comment_of_failure_reasons', 'setup_success', 'test_best', 'test_worst', 'test_best_worst', 'comparison_view']
    #session_columns = ['cycle_id', 'gtax_id', 'build_name', 'cycle_type', 'environment', 'test_plan_id', 'test_plan_name', 'status', 'progress', 'WW', 'start_date', 'end_date', 'duration', 'max_job_duration', 'duration_min', 'tpt_met', 'remote_name', 'lgcb_bad', 'has_notes', 'has_regressions', 'setup_tasks_success', 'setup_tasks_total', 'setup_success', 'test_tasks_best_success', 'test_tasks_worst_success', 'test_tasks_total', 'test_best_worst', 'comparison_view']
    print_data_table(ws, ci_test_sessions_stat, session_columns, 1, 1)
    ws.auto_filter.ref = ws.dimensions
    set_width(ws, max_width=600)
    ws = output_wb.create_sheet('smoke-report')
    print_tpt_report_data_table(ws, prepare_tpt_report(ci_test_sessions_stat), 2, 2)
    set_width(ws, min_width=10, max_width=600)
    output_wb.save(filename)
    print(f'Saved the output file:{filename}')
    return 0

def prepare_tpt_report(ci_test_sessions_stat):
    tpt_report_dict = dict()
    pd_report_data = pd.DataFrame(ci_test_sessions_stat)
    pd_all_alarms = pd_report_data.loc[(pd_report_data['has_regressions'] == True) | (pd_report_data['status'] == 'TIMEOUT')]
    tpt_report_dict.update({'False-Alarms to all Failed (Simulation) [%]':mean_for_ww(pd_all_alarms, 'false_alarm', 'Simulation')})
    tpt_report_dict.update({'False-Alarms to all Failed (Silicon) [%]':mean_for_ww(pd_all_alarms, 'false_alarm', 'Silicon')})
    tpt_report_dict.update({'False-Alarms to all Smokes (Simulation) [%]':mean_for_ww(pd_report_data, 'false_alarm', 'Simulation')})
    tpt_report_dict.update({'False-Alarms to all Smokes (Silicon) [%]':mean_for_ww(pd_report_data, 'false_alarm', 'Silicon')})
    tpt_report_dict.update({'Job Setup Success Rate (Simulation) [%]':mean_for_ww(pd_report_data, 'setup_success', 'Simulation')})
    tpt_report_dict.update({'Job Setup Success Rate (Silicon) [%]':mean_for_ww(pd_report_data, 'setup_success', 'Silicon')})
    tpt_report_dict.update({'Test Best-Worst Success Rate (Simulation) [%]':mean_for_ww(pd_report_data, 'test_best_worst', 'Simulation')})
    tpt_report_dict.update({'Test Best-Worst Success Rate (Silicon) [%]':mean_for_ww(pd_report_data, 'test_best_worst', 'Silicon')})
    tpt_report_dict.update({'Average TPT (Simulation) [min]':mean_for_ww(pd_report_data, 'duration_min', 'Simulation')})
    tpt_report_dict.update({'Average TPT (Silicon) [min]':mean_for_ww(pd_report_data, 'duration_min', 'Silicon')})
    tpt_report_dict.update({'TPT Success Rate (Simulation) [%]':mean_for_ww(pd_report_data, 'tpt_met', 'Simulation')})
    tpt_report_dict.update({'TPT Success Rate (Silicon) [%]':mean_for_ww(pd_report_data, 'tpt_met', 'Silicon')})
    tpt_report_dict.update({'targets':['<25', '<25', '<5', '<5', '>98', '>98', '<5', '<5', '<=60', '<=60', '>90', '>90']})
    tpt_report_dict.update({'all_ww':pd_report_data['WW'].sort_index(axis=0, ascending=True).unique().tolist()})
    return tpt_report_dict

def mean_for_ww(pd_report_data, column_name, environment):
    pd_mean_data = pd.DataFrame(pd_report_data[pd_report_data['environment'] == environment].groupby(['WW']).mean()[[column_name]]).round(decimals = 2)
    pd_mean_data.sort_index(axis=0, ascending=False, inplace=True)
    return pd_mean_data[column_name].to_dict()

def print_tpt_report_data_table(ws, report_data, start_row, start_column):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    report_keys = list(report_data.keys())
    report_keys.remove('all_ww')
    report_keys.remove('targets')
    data_keys = report_data['all_ww']
    row = start_row + 1
    column = start_column
    for key in report_keys:
        ws.cell(row=row, column=column, value=key)
        ws.cell(row=row, column=column).font = Font(bold = True)
        ws.cell(row=row, column=column).border = thin_border
        row += 1
    start_column += 1
    row = start_row
    column = start_column
    ws.cell(row=row, column=column, value='Targets')
    ws.cell(row=row, column=column).font = Font(bold = True)
    ws.cell(row=row, column=column).border = thin_border
    row += 1
    column = start_column
    for target in report_data['targets']:
        ws.cell(row=row, column=column, value=target)
        ws.cell(row=row, column=column).font = Font(bold = True)
        ws.cell(row=row, column=column).border = thin_border
        row += 1
    start_column += 1
    row = start_row
    column = start_column
    for key in data_keys:
        ws.cell(row=row, column=column, value=key)
        ws.cell(row=row, column=column).font = Font(bold = True)
        ws.cell(row=row, column=column).border = thin_border
        column += 1
    for report_key in report_keys:
        row += 1
        column = start_column
        for key in data_keys:
            if key in report_data[report_key].keys():
                ws.cell(row=row, column=column, value=report_data[report_key][key])
            else:
                ws.cell(row=row, column=column, value=0)
            ws.cell(row=row, column=column).border = thin_border
            column += 1
    return row

def update_stats(ci_cycle_stat, ci_test_sessions_stat, cycles):
    ci_cycle_stat += get_ci_cycle_stat(cycles)
    ci_test_sessions_stat += get_ci_cycle_session_stat(cycles)
    return 0

def print_data_table(ws, report_data, data_keys, start_row, start_column):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    row = start_row
    column = start_column
    for key in data_keys:
        ws.cell(row=row, column=column, value=key)
        ws.cell(row=row, column=column).font = Font(bold = True)
        column += 1
    for data in report_data:
        row += 1
        column = start_column
        fill_color = set_fill_by_data_conditions(data)
        font_color = set_font_by_data_conditions(data)
        for key in data_keys:
            if key == 'cycle_id':
                ws.cell(row=row, column=column, value=data[key])
                ws.cell(row=row, column=column).style = 'Hyperlink'
                ws.cell(row=row, column=column).hyperlink = f"https://gta.intel.com/#/workflow/testruns?stream=&page=1&count=5&cols=stream,testRunId,dagId,startTime,finishTime,status,cycleType,testSessions,&sorting%5Bid%5D=desc&filter%5Bid%5D={data[key]}"
                ws.cell(row=row, column=column).font = Font(color='0000FF')
            elif key == 'test_plan_id':
                ws.cell(row=row, column=column, value=data[key])
                ws.cell(row=row, column=column).style = 'Hyperlink'
                ws.cell(row=row, column=column).hyperlink = f"https://gta.intel.com/#/testplanning/plan/{data[key]}?timestamp={data['test_plan_timestamp']}"
                ws.cell(row=row, column=column).font = Font(color='0000FF')
            elif key == 'build_name':
                ws.cell(row=row, column=column, value=data[key])
                ws.cell(row=row, column=column).style = 'Hyperlink'
                ws.cell(row=row, column=column).hyperlink = data['build_link']
                ws.cell(row=row, column=column).font = Font(color='0000FF')
            elif key == 'gtax_id':
                ws.cell(row=row, column=column, value=data[key])
                ws.cell(row=row, column=column).style = 'Hyperlink'
                ws.cell(row=row, column=column).hyperlink = data['url']
                ws.cell(row=row, column=column).font = Font(color='0000FF')
            elif key == 'comparison_view':
                ws.cell(row=row, column=column, value=key)
                ws.cell(row=row, column=column).style = 'Hyperlink'
                ws.cell(row=row, column=column).hyperlink = data[key]
                ws.cell(row=row, column=column).font = Font(color='0000FF')
            #elif key == 'comments':
            #    ws.cell(row=row, column=column, value=str(data[key]).encode('unicode_escape'))
            else:
                # skip for missing key (blank values)
                if key in data.keys():
                    ws.cell(row=row, column=column, value=data[key])
                    font_bold = False
                    if str(data[key]) in ['True', 'TIMEOUT', 'IN PROGRESS']:
                        font_bold = True
                    ws.cell(row=row, column=column).font = Font(color=font_color, bold=font_bold)
            ws.cell(row=row, column=column).border = thin_border
            ws.cell(row=row, column=column).fill = PatternFill("solid", fgColor=fill_color)
            column += 1
    return row

def set_fill_by_data_conditions(data):
    fill_color = 'FFFFFF'
    if get_stat_key_value(data, 'tpt_met', if_not_found=1) == 0:
        fill_color = 'E0E0E0'
    if get_stat_key_value(data, 'setup_success', if_not_found=100) < 95.0:
        fill_color = 'CCAAAA'
    if get_stat_key_value(data, 'test_best_worst', if_not_found=0) > 5.0:
        fill_color = 'FFAAAA'
    return fill_color

def set_font_by_data_conditions(data):
    font_color = '000000'
    if get_stat_key_value(data, 'setup_success', if_not_found=100) < 95.0:
        font_color = 'FFFF00'
    return font_color

def get_cycles(dag_id, stream, submitter=None, date_from=None, date_to=None, status=None, count=100):
    cycles_list = list()
    url = f'http://gta.intel.com/api/workflow/v2/test_runs_dashboard?count={count}&filter%5Bdag_id%5D={dag_id}'
    print(f'geting cycles  dag_id:{dag_id}', end=' ') 
    if submitter:
        url += f'&filter%5Bsubmitter%5D={submitter}'
        print(f'submitter:{submitter}', end=' ')
    if stream:
        url += f'&filter%5Bstream%5D={stream}'
        print(f'stream:{stream}', end=' ')
    if status in ['NEW', 'IN PROGRESS' ,'COMPLETED', 'ERROR', 'TIMEOUT', 'INCOMPLETE']:
        url += f'&filter%5Bstatus%5D={status}'
        print(f'status:{status}', end=' ')
    if status == 'DEFAULT':
        url += '&filter%5Bstatus%5D!=NEW&filter%5Bstatus%5D!=IN%20PROGRESS'
        print('status: !NEW AND !IN PROGRESS', end=' ')
    if date_from:
        url += f'&filter%5Btest_run_start_date%5D=%3E{date_from}T00%3A00%3A00Z'        
        print(f'date_from:{date_from}', end=' ')
    if date_to:
        url += f'&filter%5Btest_run_start_date%5D=%3C{date_to}T23%3A59%3A59Z'
        print(f'date_to:{date_to}', end=' ')
    url += '&sorting%5Bid%5D=desc&filter%5Bexact_dag_id%5D=true'
    print(' ')
    #print(url)
    cycles_list = get_gta_cycles_pages(url)
    return cycles_list


def get_ci_cycle_stat(ci_cycles):
    ci_cycle_stat = list()
    for cycle in ci_cycles:
        cycle_dict = dict()
        cycle_dict.update({'cycle_id':cycle['id']})
        for key in ['dag_id', 'submitter', 'start_date', 'status_date', 'status', 'cycle_type', 'label']:
            cycle_dict.update({key:cycle[key]})
        cycle_dict.update({'end_date':get_end_date(cycle['status_date'], cycle['status'])})    
        cycle_dict.update({'duration':get_duration(cycle['start_date'], cycle_dict['end_date'])})
        cycle_dict.update({'duration_seconds':get_duration(cycle['start_date'], cycle_dict['end_date'], seconds=True)})
        cycle_dict.update({'stream_name':cycle['builds'][0]['stream_name']})
        cycle_dict.update({'build_name':cycle['builds'][0]['name']})
        cycle_dict.update({'build_id':cycle['builds'][0]['external_id']})
        cycle_dict.update({'build_link':cycle['builds'][0]['url']})
        cycle_dict.update({'build_start_date':cycle['builds'][0]['start_date']})
        cycle_dict.update({'started_after_build':get_duration(cycle['builds'][0]['status_date'], cycle['start_date'])})
        cycle_dict.update({'started_after_build_seconds':get_duration(cycle['builds'][0]['status_date'], cycle['start_date'], seconds=True)})
        cycle_dict.update({'lgcb_bad':get_flag_status(cycle['test_sessions'], 'lgcb_bad')})
        cycle_dict.update({'has_regressions':get_flag_status(cycle['test_sessions'], 'has_regressions')})
        cycle_dict.update({'comparison_view':get_cycle_comparison_view_link(cycle['builds'][0]['name'], cycle['id'])})
        cycle_dict.update({'WW':convert_to_intel_ww(cycle['start_date'])})
        ci_cycle_stat.append(cycle_dict)
    return ci_cycle_stat

def get_flag_status(session_list, flag_name):
    flag_status = False
    for session in session_list:
        if session[flag_name]:
            flag_status = True
            break
    return flag_status

def get_ci_cycle_session_stat(ci_cycles):
    session_stat = list()
    for cycle in ci_cycles:
        for session in cycle['test_sessions']:
            session_dict = dict()
            session_dict.update({'stream_name':cycle['builds'][0]['stream_name']})
            session_dict.update({'build_name':cycle['builds'][0]['name']})
            session_dict.update({'build_id':cycle['builds'][0]['external_id']})
            session_dict.update({'build_link':cycle['builds'][0]['url']})
            session_dict.update({'cycle_id':cycle['id']})
            for key in session.keys():
                if key == 'id':
                    session_dict.update({'session_id':session[key]})
                elif key == 'external_id':
                    session_dict.update({'gtax_id':session[key]})
                elif key == 'url':
                    url = session[key]
                    session_dict.update({'url':url})
                    if url:
                        instance_start = url.find('//') + 2
                        instance_end = url.find('.intel.com')
                        session_dict.update({'gtax_instance':url[instance_start:instance_end]})
                    else:
                        session_dict.update({'gtax_instance':''})
                elif key == 'cycle_type':
                    session_dict.update({key:cycle[key]})
                elif key == 'test_session_failure':
                    failure_reasons_set = set()
                    failure_category_set = set()
                    if 'comment' in session['test_session_failure'].keys():
                        session_dict.update({'comments':str(session['test_session_failure']['comment']).encode('unicode_escape')})
                    else:
                        session_dict.update({'comments':'None'})
                    if 'categories' in session['test_session_failure'].keys():
                        for category in session['test_session_failure']['categories']:
                            if category['selected']:
                                failure_category_set.add(category['name'])
                                reason_set = set()
                                for reason in category['values']:
                                    if reason['selected']:
                                        # WA for "GTA (WF, RES, TP)" - > "GTA (WF-RES-TP)"
                                        reason_set.add(reason['name'].replace(', ','-'))
                                if len(reason_set) > 0:
                                    failure_reasons_set.add(f"{category['name']}:{','.join(reason_set)}")
                                else:
                                    failure_reasons_set.add(category['name'])
                    if len(failure_reasons_set) > 0:
                        session_dict.update({'failure_reasons':','.join(failure_reasons_set)})
                        if 'Code Issue' in failure_category_set:
                            session_dict.update({'has_code_issue':True})
                            session_dict.update({'false_alarm':0})
                        else:
                            session_dict.update({'has_code_issue':False})
                            if session['has_regressions'] == True or session['status'] == 'TIMEOUT':
                                session_dict.update({'false_alarm':100})
                            else:
                                session_dict.update({'false_alarm':0})
                    else:
                        session_dict.update({'failure_reasons':'None'})
                        session_dict.update({'has_code_issue':False})
                        session_dict.update({'false_alarm':0})
                else:
                    session_dict.update({key:session[key]})
            session_dict.update({'end_date':get_end_date(session['status_date'], session['status'])}) 
            session_dict.update({'duration':get_duration(session['start_date'], session_dict['end_date'])})
            session_dict.update({'duration_seconds':get_duration(session['start_date'], session_dict['end_date'], seconds=True)})
            if session_dict['duration_seconds']:
                session_dict.update({'duration_min':round(session_dict['duration_seconds']/60, 0)})
            else:
                session_dict.update({'duration_min':'N/A'})
            session_dict.update({'comparison_view':get_session_comparison_view_link(cycle['builds'][0]['name'], cycle['id'], session['id'])})
            session_dict.update({'WW':convert_to_intel_ww(cycle['start_date'])})
            session_stat.append(session_dict)
    return session_stat

def get_cycle_comparison_view_link(build_name, cycle_id, worst=False):
    url = 'https://gta.intel.com/#/reports/comparison-view?diffMode=false&reRuns[name]=all&table[page]=1&table[count]=25&visibleColumns[]=Item%20Name&visibleColumns[]=Args&visibleColumns[]=OS&visibleColumns[]=Platform&visibleColumns[]=Vertical&visibleColumns[]=Test%20Run&visibleColumns[]=Test%20Session&compareFields[]=compare_id&settingsFromUrl=true&tagExcept[0][name]=notAnIssue&tagExcept[1][name]=obsoleted&tagExcept[2][name]=iteration&tagExcept[3][name]=isolation'
    if worst:
        url += '&rerun=Worst'
    url += f'&globalFilterId=&builds[0][name]={build_name}&builds[0][key]={build_name}&filters[test_run][0][name]={cycle_id}'
    return url

def get_session_comparison_view_link(build_name, cycle_id, session_id):
    url = get_cycle_comparison_view_link(build_name, cycle_id, worst=True)
    url += f'&filters[test_session][0][name]={session_id}'
    return url

def update_sessions_with_tp_data(ci_test_sessions_stat):
    test_plan_dict = get_test_plan_dict(ci_test_sessions_stat)
    for session in ci_test_sessions_stat:
        if session['test_plan_id'] in test_plan_dict.keys():
            session.update({'test_plan_name':get_tp_name(session['test_plan_id'], test_plan_dict)})
            session.update({'environment':get_tp_env(session['test_plan_id'], test_plan_dict)})
        else:
            # for null tp id
            session.update({'test_plan_name':session['label']})
            session.update({'environment':'unknown'})

def get_end_date(status_date, status, time_format='%Y-%m-%dT%H:%M:%SZ'):
    end_date = status_date
    if status == 'IN PROGRESS':
        time_end = datetime.utcnow()
        end_date = time_end.strftime(time_format)
    return end_date

def get_tp_name(tp_id, test_plan_dict):
    tp_name = None
    if tp_id:
        tp_name = test_plan_dict[tp_id]['name']
    return tp_name

def get_tp_env(tp_id, test_plan_dict):
    tp_env = 'unknown'
    if tp_id:
        tp_env = test_plan_dict[tp_id]['env']
    return tp_env

def get_test_plan_dict(ci_session_stat):
    test_plan_dict = dict()
    tp_cache_file = 'smoke_stat_tp_dict.cache'
    if os.path.isfile(tp_cache_file):
        with open(tp_cache_file, 'rb') as cache:
            test_plan_dict = pickle.load(cache)
    test_plan_ids_set = set()
    for session in ci_session_stat:
        if session['test_plan_id']:
            test_plan_ids_set.add(session['test_plan_id'])
    for tp_id in test_plan_ids_set:
        if tp_id not in test_plan_dict.keys():
                tp_data = get_test_plan_data(tp_id)
                test_plan_dict.update({tp_id:{'name':tp_data['name'], 'env':tp_data['env']}})
    with open(tp_cache_file, 'w+b') as cache:
        pickle.dump(test_plan_dict, cache) 
    return test_plan_dict

def get_test_plan_data(test_plan_id):
    url = f'http://gta.intel.com/api/tp/v1/testplans/{test_plan_id}'
    print(f'geting test plan data for {test_plan_id}')
    response = get_gta_data(url)
    test_plan = response.json()
    tp_data = dict()
    tp_env = 'unknown'
    tp_data.update({'name':test_plan['name']})
    for attribute in test_plan['attributes']:
        if attribute['name'] == 'Environment':
            tp_env = attribute['resolvedValue']
    tp_data.update({'env':tp_env})
    return tp_data

def get_duration(start_date, end_date, seconds=False, time_format='%Y-%m-%dT%H:%M:%SZ', timedate_format=False):
    duration = None
    if start_date and end_date:
        if timedate_format:
            time_start = start_date
            time_end = end_date
        else:
            time_start = datetime.strptime(start_date, time_format)
            time_end = datetime.strptime(end_date, time_format)
        if seconds:
            duration = round((time_end - time_start).total_seconds())
        else:
            duration = str(time_end - time_start)
    return duration

def update_sessions_with_gtax_data(ci_test_sessions_stat, offline_mode=False):
    call_count = 0
    call_total = len(ci_test_sessions_stat)
    for session in ci_test_sessions_stat:
        call_count += 1
        if session['gtax_id']:
            print(f'getting gtax data {round(call_count/call_total*100,2)}%   ', end='\r', flush=True)
            jobs_full_info = get_gtax_jobs_full_info(session['gtax_instance'], session['gtax_id'], offline_mode)
            session_task_dict = get_session_tasks_dict(jobs_full_info)
            session_jobs_dict = get_session_jobs_dict(jobs_full_info)
            # init setup success rate
            init_setup_stat = get_gtax_task_stat(session_task_dict, 'setup', submission_type=['init'])
            session.update({'setup_tasks_total':get_stat_key_value(init_setup_stat, 'total')})
            session.update({'setup_tasks_success':get_stat_key_value(init_setup_stat, 'passed') + get_stat_key_value(init_setup_stat, 'blocked') + get_stat_key_value(init_setup_stat, 'ignored')})
            if session['setup_tasks_total'] > 0:
                session.update({'setup_success':round(session['setup_tasks_success']/session['setup_tasks_total']*100, 2)})
            else:
                session.update({'setup_success':'no setup tasks'})
            # test success rate (best-worst)
            init_test_stat = get_gtax_task_stat(session_task_dict, 'test', submission_type=['init'])
            rerun_test_stat = get_gtax_task_stat(session_task_dict, 'test', submission_type=['resume', 'rerun'])
            session.update({'test_tasks_total':get_stat_key_value(init_test_stat, 'total')})
            session.update({'test_tasks_worst_success':get_stat_key_value(init_test_stat, 'passed') + get_stat_key_value(init_test_stat, 'blocked')})
            session.update({'test_tasks_best_success':get_stat_key_value(init_test_stat, 'passed') + get_stat_key_value(rerun_test_stat, 'passed') + get_stat_key_value(init_test_stat, 'blocked')})
            if session['test_tasks_total'] > 0:
                session.update({'test_best':round(session['test_tasks_best_success']/session['test_tasks_total']*100, 2)})
                session.update({'test_worst':round(session['test_tasks_worst_success']/session['test_tasks_total']*100, 2)})
                session.update({'test_best_worst':round((session['test_tasks_best_success']-session['test_tasks_worst_success'])/session['test_tasks_total']*100, 2)})
            else:
                session.update({'test_best':'no test tasks'})
                session.update({'test_worst':'no test tasks'})
                session.update({'test_best_worst':'no test tasks'})
            # max job duration in min.
            max_job_duration = get_max_job_duration(session_jobs_dict)
            session.update({'max_job_duration':max_job_duration})
            session.update({'tpt_met':get_tpt_met(session['duration_min'], 45)})
        else:
            for key in ['setup_tasks_total', 'setup_tasks_success', 'setup_success', 'test_tasks_total', 'test_tasks_worst_success', 'test_tasks_best_success', 'test_best', 'test_worst', 'test_best_worst', 'max_job_duration', 'tpt_met']:
                session.update({key:'N/A'})
    print('\n')

def get_max_job_duration(session_jobs_dict):
    max_duration = 0
    for job in session_jobs_dict:
        if isinstance(session_jobs_dict[job]['duration'], (float, int)):
            if int(session_jobs_dict[job]['duration']) > max_duration:
                max_duration = round(int(session_jobs_dict[job]['duration'])/60, 0)
    return max_duration

def get_tpt_met(max_job_duration, limit):
    if max_job_duration > limit:
        tpt_met = 0
    else:
        tpt_met = 100
    return tpt_met

   
def get_gtax_task_stat(session_task_dict, phase, submission_type=[]):
    task_stat_list = list()
    task_status_set = set()
    for task in session_task_dict:
        if session_task_dict[task]['phase'] == phase:
            if len(submission_type) > 0:
                if session_task_dict[task]['submission_type'] in submission_type:
                    task_stat_list.append(session_task_dict[task]['status'])
                    task_status_set.add(session_task_dict[task]['status'])
            else:
                task_stat_list.append(session_task_dict[task]['status'])
                task_status_set.add(session_task_dict[task]['status'])
    task_stat_dict = dict()
    for status in task_status_set:
        task_stat_dict.update({f'{status}':task_stat_list.count(status)})
    task_stat_dict.update({f'total':len(task_stat_list)})
    return(task_stat_dict)

def get_stat_key_value(stat_dict, key, if_not_found=0):
    stat_key_value = if_not_found
    if key in stat_dict.keys():
        stat_key_value = stat_dict[key]
    if not isinstance(stat_key_value, (float, int)):
        stat_key_value = if_not_found
    return stat_key_value
        
def get_session_tasks_dict(session_jobs_data):
    session_task_dict = dict()
    for job_data in session_jobs_data:
        for task_data in job_data['tasks']:
            task_dict = dict()
            for key in ['id', 'job_id', 'task_result_id', 'task_definition_id']:
                task_dict.update({key:task_data[key]})
            if job_data['client']:
                task_dict.update({'client_name':job_data['client']['name']})
                task_dict.update({'client_id':job_data['client']['id']})
            else:
                task_dict.update({'client_name':'unassigned'})
                task_dict.update({'client_id':'unassigned'})
            if 'task_result' in task_data.keys():
                for key in ['submission_type', 'gta_result_key', 'started_date', 'completed_date', 'status', 'gta_status', 'primary_reason', 'publishing_dumps']:
                    if key == 'gta_result_key':
                        task_dict.update({'phase':get_task_phase(task_data['task_result']['gta_result_key'])})
                        task_dict.update({key:task_data['task_result'][key]})
                    else:
                        task_dict.update({key:task_data['task_result'][key]})
            if 'task_definition' in task_data.keys():
                task_dict.update({'command':task_data['task_definition']['command']})
            if 'started_date' in task_dict.keys() and 'completed_date' in task_dict.keys():
                task_dict.update({'duration':get_duration(task_dict['started_date'], task_dict['completed_date'], time_format='%Y-%m-%d %H:%M:%S.%f')})
                task_dict.update({'duration_seconds':get_duration(task_dict['started_date'], task_dict['completed_date'], seconds=True, time_format='%Y-%m-%d %H:%M:%S.%f')})
            session_task_dict.update({task_data['id']:task_dict})
    return session_task_dict

def get_session_jobs_dict(session_jobs_data):
    session_jobs_dict = dict()
    for job_data in session_jobs_data:
        job_dict = dict()
        for key in ['id', 'status', 'name', 'result', 'submission_type', 'submitted_date', 'started_date', 'completed_date', 'duration', 'scheduler_account_name', 'dumps_produced']:
            job_dict.update({key:job_data[key]})
        if 'csq' in job_data.keys():
            csq = job_data['csq']['query']
            job_dict.update({'csq':csq})
            job_dict.update({'platform':get_client_platform(csq)})
            job_dict.update({'pool':get_client_pool(csq)})
        session_jobs_dict.update({job_data['id']:job_dict})
    return session_jobs_dict    

def get_client_platform(csq):
    platform_regex = re.compile(r"('platform' = '.*?')") 
    platform = 'unknown'
    if platform_regex.search(csq):
        platform = platform_regex.search(csq).group(0)[14:-1]
    return platform

def get_client_pool(csq):
    platform_regex = re.compile(r"('pool' = '.*?')") 
    platform = 'unknown'
    if platform_regex.search(csq):
        platform = platform_regex.search(csq).group(0)[10:-1]
    return platform   

def get_task_phase(gta_result_key):
    phase = 'unknown'
    if gta_result_key: 
        for phase_name in ['teardown', 'setup', 'test']:
            if gta_result_key.find(phase_name) > 0:
                phase = phase_name
    else:
        phase = ''
    return phase

def get_gtax_jobs_full_info_old(gtax_instance, job_set_session_id, offline_mode=False):
    app_path = os.path.dirname(os.path.abspath(sys.argv[0]))
    cache_path = os.path.join(app_path, '_cache')
    cache_file_name = os.path.join(cache_path, f"gtax_{job_set_session_id}_full_info.cache")
    if offline_mode and os.path.isfile(cache_file_name):
        with open(cache_file_name, 'rb') as cache:
            jobs_full_info = pickle.load(cache)
    else:
        #print(f'be patient!! - getting full info of tasks for job set session: {job_set_session_id}')
        url = f'http://{gtax_instance}.intel.com/api/v1/jobs?include_tasks=true&include_csq=true&include_phases_task_counts=true&full_info=true&jobset_session_ids={job_set_session_id}'
        #print(url)
        data = get_gtax_data(url)
        jobs_full_info = data['data']
        if not os.path.exists(cache_path):
            os.makedirs(cache_path)
        with open(cache_file_name, 'w+b') as cache:
            pickle.dump(jobs_full_info, cache) 
    return jobs_full_info

def get_gtax_jobs_full_info(gtax_instance, job_set_session_id, offline_mode=False):
    app_path = os.path.dirname(os.path.abspath(sys.argv[0]))
    cache_path = os.path.join(app_path, '_cache')
    cache_file_name = os.path.join(cache_path, f"gtax_{job_set_session_id}_full_info.gzip")
    if offline_mode and os.path.isfile(cache_file_name):
        jobs_full_info = cache_json_load(cache_file_name)
    else:
        #print(f'be patient!! - getting full info of tasks for job set session: {job_set_session_id}')
        url = f'http://{gtax_instance}.intel.com/api/v1/jobs?include_tasks=true&include_csq=true&include_phases_task_counts=true&full_info=true&jobset_session_ids={job_set_session_id}'
        #print(url)
        data = get_gtax_data(url)
        jobs_full_info = data['data']
        if not os.path.exists(cache_path):
            os.makedirs(cache_path)
        cache_json_save(jobs_full_info, cache_file_name)
    return jobs_full_info

def post_gta_data(url, data):
    output = None
    headers = { 'Content-type': 'application/json' }
    response = requests.put(url, data, headers=headers, proxies={'http': 'http://proxy-chain.intel.com:911', 'https': 'http://proxy-chain.intel.com:912' })
    if response.status_code == 200:
        output = response
    else:
        print(response.status_code)
        print(response)
    return output

def get_gta_data(url, page=None):
    output = None
    headers = { 'Content-type': 'application/json' }
    if page:
        url += f'&page={page}'
    response = requests.get(url, headers=headers, proxies={'http': 'http://proxy-chain.intel.com:911', 'https': 'http://proxy-chain.intel.com:912' })
    if response.status_code == 200:
        output = response
    else:
        print(response.status_code)
        print(response)
    return output

def get_gta_cycles_pages(url):
    cycles_list = list()
    page = 1
    response = get_gta_data(url, page=page)
    cycles = response.json()
    cycles_list.extend(cycles['items'])
    pages = cycles['pages']
    if pages > 1:
        for i in range(page+1, pages+1):
            response = get_gta_data(url, page=i)
            cycles = response.json()
            cycles_list.extend(cycles['items'])
    return cycles_list

def get_gtax_data(url):
    headers = { 'Content-type': 'application/json' }
    response = requests.get(url, headers=headers, proxies={'http': 'http://proxy-chain.intel.com:911', 'https': 'http://proxy-chain.intel.com:912' })
    return response.json()

def set_width(ws, min_width=0, max_width=100):
    column_widths = []
    column_heights = []
    for row_cnt, row in enumerate(ws.iter_rows()):
        column_heights.append(0)
        for col_cnt, cell in enumerate(row):
            if len(column_widths) <= col_cnt:
                column_widths.append(min_width)
            # Ignore cells that don't wrap
            if cell.alignment.wrap_text is False:
                continue
            if is_merged_cell(ws, cell):
                continue
            length, num_lines = calculate_cell_length(cell)
            if cell.alignment.text_rotation == 0:
                column_widths[col_cnt] = max(column_widths[col_cnt], length)
                row_height = 14 * num_lines
                column_heights[row_cnt] = max(column_heights[row_cnt], row_height)
            else:
                column_heights[row_cnt] = max(column_heights[row_cnt], 5 * length)
                col_width = 14 * num_lines
                column_widths[col_cnt] = max(column_widths[col_cnt], col_width)
    max_column_width_allowed = max_width
    for cnt, column_width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(cnt)].width = min(column_width, max_column_width_allowed)
    for cnt, row_height in enumerate(column_heights, 1):
        ws.row_dimensions[cnt].height = row_height

def is_merged_cell(ws, cell):
    for cell_range in ws.merged_cells:
        if cell.column >= cell_range.min_col and cell.column <= cell_range.max_col and \
           cell.row    >= cell_range.min_row and cell.row    <= cell_range.max_row:
            return True
    return False

def calculate_cell_length(cell):
    num_lines = 1
    if isinstance(cell.value, str):
        if 'HYPERLINK' in cell.value:
            pieces = cell.value.split(',')
            val = pieces[-1].split('"')[1]
            length = len(val)
        else:
            if '\n' in cell.value:
                num_lines = len(cell.value.split('\n'))
                length = max([len(s) for s in cell.value.split('\n')])
            else:
                length = len(cell.value)
    elif isinstance(cell.value, bool):
        length = len('False')
    elif isinstance(cell.value, int):
        if cell.value == 0:
            length = 1
        else:
            if cell.value < 0:
                val = -1 * cell.value
            else:
                val = cell.value
            length = math.ceil(math.log10(abs(cell.value)))
    else:
        length = 0
    length += 2
    return length, num_lines

def get_filename_timestamp():
    ts = str(datetime.now())
    ts = ts.replace(':', '_')
    return ts[5:-7]

def get_date_n_days_ago(days_ago):
    ts = str(datetime.now()- timedelta(days=days_ago))
    return ts[0:10]

def check_user_input(user_input):
    try:
        datetime.strptime(user_input, '%Y-%m-%d')
    except ValueError:
        return False
    return True

def convert_to_intel_ww(date_string):
    #correct to Intel WW
    date_time = datetime.strptime(date_string, '%Y-%m-%dT%H:%M:%SZ')
    iso_year = date_time.isocalendar()[0]
    iso_ww = date_time.isocalendar()[1]
    max_iso_ww = datetime.strptime(f'{iso_year}-12-31', '%Y-%m-%d').isocalendar()[1]
    first_day_of_the_year = int(datetime.strptime(f'{iso_year}-01-01', '%Y-%m-%d').strftime("%w"))
    last_day_of_the_year = int(datetime.strptime(f'{iso_year}-12-31', '%Y-%m-%d').strftime("%w"))
    day_of_year = int(date_time.strftime("%j"))
    if first_day_of_the_year in [0, 1, 2, 3, 4]:
        intel_ww = iso_ww
    else:
        intel_ww = iso_ww + 1
    if iso_ww == max_iso_ww:
        if last_day_of_the_year != 6:
            intel_ww = 1
    if day_of_year == 1 and intel_ww > 1:
        intel_ww = 1
    return f'{str(iso_year)[2:]}WW{str(intel_ww).zfill(2)}'

def cache_json_save(data, gzip_file_name):
    with gzip.open(gzip_file_name, 'w') as f:
        f.write(json.dumps(data).encode('utf-8')) 
    return 0

def cache_json_load(gzip_file_name):
    with gzip.open(gzip_file_name, 'r') as f:
        data = json.loads(f.read().decode('utf-8'))
    return data

if __name__ == '__main__':
    ts = get_filename_timestamp()
    filename = f'smoke_tpt_report_{ts}.xlsx'
    ap = argparse.ArgumentParser()
    ap.add_argument('-s', '-status', default='DEFAULT', help='status filter values: [NEW, IN PROGRESS ,COMPLETED, ERROR, TIMEOUT, INCOMPLETE]', required=False)
    #ap.add_argument('-c', '-count', default='100', help='max cycles count per gta request', required=False)
    ap.add_argument('-i', '-interactive', action='store_true', default=False, help='interactive mode', required=False)
    ap.add_argument('-f', '-date_from', default=get_date_n_days_ago(7), help='date from format yyyy-mm-dd', required=False)
    ap.add_argument('-t', '-date_to', default=get_date_n_days_ago(1), help='date to format yyyy-mm-dd', required=False)
    ap.add_argument('-submitter', default='sys_gtawf', help='submitter default:sys_gtawf', required=False)
    ap.add_argument('-dag', default='CI_SMOKE__gfx-driver__master:gfx-driver__master', help='DAG stream default: CI_SMOKE__gfx-driver__master:gfx-driver__master', required=False)
    ap.add_argument('-o', '-output', default=filename, help='output file name', required=False)
    #ap.add_argument('-no_cache', action='store_false', default=True, help='no cache for gtax api calls (could be use to update cache)', required=False)
    parsed = ap.parse_args()
    date_from = parsed.f
    date_to = parsed.t
    if parsed.i:
        date_from = input(f'provide from date (YYYY-MM-DD) [enter use default={get_date_n_days_ago(7)}]:')
        if not check_user_input(date_from):
            date_from = get_date_n_days_ago(7)
        date_to = input(f'provide to date (YYYY-MM-DD) [enter use default={get_date_n_days_ago(1)}]:')
        if not check_user_input(date_to):
            date_to = get_date_n_days_ago(1)
    main(parsed.s, 100, date_from, date_to, parsed.dag, parsed.submitter, parsed.o, True)
