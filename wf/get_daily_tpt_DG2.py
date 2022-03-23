import sys
import math
import re
import os.path
import argparse
import pickle
from datetime import datetime
from datetime import timedelta
from operator import itemgetter
import time
import json
import requests
import openpyxl
import pandas as pd
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList

def main(status, count, date_from, date_to, dag_stream, submitter, filename, offline_mode=False):
    ci_cycle_stat = list()
    ci_test_sessions_stat = list()
    ci_jobs_data_dict = dict()

    #GETING DATA
    cache_file1_name = f'cycles_master_tmp.cache'
    cache_file2_name = f'sessions_master_tmp.cache'
    cache_file3_name = f'jobs_dict_tmp.cache'
    cache_wf = False
    #cache_wf = True

    if cache_wf and os.path.isfile(cache_file1_name) and os.path.isfile(cache_file2_name):
        with open(cache_file1_name, 'rb') as cache:
            ci_cycle_stat = pickle.load(cache)
        with open(cache_file2_name, 'rb') as cache:
            ci_test_sessions_stat = pickle.load(cache)
        with open(cache_file3_name, 'rb') as cache:
            ci_jobs_data_dict = pickle.load(cache)    
    else:
        if dag_stream.find(':') > 0:
            dag_name = dag_stream.split(':')[0]
            stream_name = dag_stream.split(':')[1]
            update_stats(ci_cycle_stat, ci_test_sessions_stat, get_cycles(dag_name, stream_name, submitter, status=status, count=count, date_from=date_from, date_to=date_to))
            print(f'{len(ci_cycle_stat)} cycles and {len(ci_test_sessions_stat)} sessions found')
            update_sessions_with_tp_data(ci_test_sessions_stat)
            update_sessions_with_gtax_data(ci_test_sessions_stat, ci_jobs_data_dict, offline_mode)
            with open(cache_file1_name, 'w+b') as cache:
                pickle.dump(ci_cycle_stat, cache) 
            with open(cache_file2_name, 'w+b') as cache:
                pickle.dump(ci_test_sessions_stat, cache)
            with open(cache_file3_name, 'w+b') as cache:
                pickle.dump(ci_jobs_data_dict, cache)
        else:
            print('wrong dag stream format. use DAG:stream exp: CI_SMOKE__gfx-driver__master:gfx-driver__master')
    
    print(len(ci_jobs_data_dict))

    print(ci_jobs_data_dict[list(ci_jobs_data_dict.keys())[0]])

    print(get_key_set(ci_jobs_data_dict, 'platform'))

    print('\n\n')

    #print(get_key_set(ci_jobs_data_dict, 'pool'))

    get_client_property_by_client_id('894', 'platform', 'gtax-igk')

    print('\n\n')

    for ci_job in ci_jobs_data_dict:
        if ci_jobs_data_dict[ci_job]['platform'] == 'Not Found in MasterData':
        #if ci_jobs_data_dict[ci_job]['platform'] == 'not_found':
            #print(ci_jobs_data_dict[ci_job])
            print(f"http://{ci_jobs_data_dict[ci_job]['gtax_instance']}.intel.com/#/jobs/{ci_jobs_data_dict[ci_job]['id']}?tab=taskml {ci_jobs_data_dict[ci_job]['client_name']} {ci_jobs_data_dict[ci_job]['pool']} {ci_jobs_data_dict[ci_job]['csq']}")


    print('\n\n')
    return 0

'''
    # EXCEL OUTPUT
    output_wb = openpyxl.Workbook()
    ws = output_wb.active
    ws.title = 'smoke-ci_cycles'
    cycle_columns = ['cycle_id', 'build_name', 'cycle_type', 'dag_id', 'status', 'WW', 'start_date', 'end_date', 'duration', 'build_start_date', 'started_after_build', 'lgcb_bad', 'has_regressions', 'comparison_view']
    print_data_table(ws, ci_cycle_stat, cycle_columns, 1, 1)
    ws.auto_filter.ref = ws.dimensions
    set_width(ws, max_width=600)
    ws = output_wb.create_sheet('smoke-ci_sessions')
    session_columns = ['cycle_id', 'gtax_id', 'build_name', 'cycle_type', 'environment', 'test_plan_id', 'test_plan_name', 'status', 'progress', 'WW', 'start_date', 'end_date', 'duration', 'max_job_duration', 'duration_min', 'tpt_met', 'remote_name', 'lgcb_bad', 'has_notes', 'has_regressions', 'has_comment_of_failure_reasons', 'comparison_view']
    #session_columns = ['cycle_id', 'gtax_id', 'build_name', 'cycle_type', 'environment', 'test_plan_id', 'test_plan_name', 'status', 'progress', 'WW', 'start_date', 'end_date', 'duration', 'max_job_duration', 'duration_min', 'tpt_met', 'remote_name', 'lgcb_bad', 'has_notes', 'has_regressions', 'has_comment_of_failure_reasons', 'setup_success', 'test_best', 'test_worst', 'test_best_worst', 'comparison_view']
    #session_columns = ['cycle_id', 'gtax_id', 'build_name', 'cycle_type', 'environment', 'test_plan_id', 'test_plan_name', 'status', 'progress', 'WW', 'start_date', 'end_date', 'duration', 'max_job_duration', 'duration_min', 'tpt_met', 'remote_name', 'lgcb_bad', 'has_notes', 'has_regressions', 'setup_tasks_success', 'setup_tasks_total', 'setup_success', 'test_tasks_best_success', 'test_tasks_worst_success', 'test_tasks_total', 'test_best_worst', 'comparison_view']
    print_data_table(ws, ci_test_sessions_stat, session_columns, 1, 1)
    ws.auto_filter.ref = ws.dimensions
    set_width(ws, max_width=600)
    output_wb.save(filename)
    print(f'Saved the output file:{filename}')
'''


def get_avr_by_ww_for_key(report_data, key, additional_filter=None):
    avr_by_ww = dict()
    if additional_filter:
        report_data = filer_report_data(report_data, [additional_filter])
    for ww in get_key_set(report_data, 'WW'):
        avr_by_ww.update({ww:get_avr_from_report_data(filer_report_data(report_data, [['WW', ww, 'match']]), key)})
    return avr_by_ww

def get_key_set(report_data, key):
    value_set = set()
    for data in report_data:
        value_set.add(report_data[data][key])
    return value_set

def check_condition(value, condition_value, condition_func):
    check = False
    if condition_func == 'eql':
        if float(value) == float(condition_value):
            check = True
    elif condition_func == 'n_eql':
        if float(value) != float(condition_value):
            check = True
    elif condition_func == 'greater_eql':
        if float(value) >= float(condition_value):
            check = True
    elif condition_func == 'greater':
        if float(value) > float(condition_value):
            check = True
    elif condition_func == 'less':
        if float(value) < float(condition_value):
            check = True
    elif condition_func == 'less_eql':
        if float(value) <= float(condition_value):
            check = True
    elif condition_func == 'match':
        if str(value) == str(condition_value):
            check = True
    elif condition_func == 'contain':
        if str(value).find(str(condition_value)) >= 0:
            check = True
    else:
        check = False
    return check

def get_avr_from_report_data(report_data, key):
    total = 0
    avr = 0
    for data in report_data:
        if isinstance(report_data[data][key], float) or isinstance(report_data[data][key], int):
            total += report_data[data][key]
    if len(report_data) > 0:
        avr = round(total/len(report_data), 2)
    return avr

def filer_report_data(report_data, condition_list, match_all=True):
    # format condition_list : [[condition_key1, condition_key1_value, condition_key1_func], [condition_key2, condition_key2_value, condition_key2_func]]
    # func: eql, n_eql, greater, greater_eql, less, less_eql, match, contain  
    report_data_filtered = dict()
    if match_all:
        for bug in report_data:
            condition_match = True
            for condition in condition_list:
                if check_condition(report_data[bug][condition[0]], condition[1], condition[2]) and condition_match == True:
                    condition_match = True
                else:
                    condition_match = False
            if condition_match == True:
                report_data_filtered.update({bug:report_data[bug]})
    else:
        for bug in report_data:
            condition_match = False
            for condition in condition_list:
                if check_condition(report_data[bug][condition[0]], condition[1], condition[2]):
                    condition_match = True
            if condition_match == True:
                report_data_filtered.update({bug:report_data[bug]})        
    return report_data_filtered

def update_stats(ci_cycle_stat, ci_test_sessions_stat, cycles):
    ci_cycle_stat += get_ci_cycle_stat(cycles)
    ci_test_sessions_stat += get_ci_cycle_session_stat(cycles)
    return 0

def print_data_table(ws, report_data, data_keys, start_row, start_column):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    row = start_row
    column = start_column
    for key in data_keys:
        ws.cell(row=row, column=column, value=key)
        ws.cell(row=row, column=column).font = Font(bold = True)
        column += 1
    for data in report_data:
        row += 1
        column = start_column
        fill_color = set_fill_by_data_conditions(data)
        font_color = set_font_by_data_conditions(data)
        for key in data_keys:
            if key == 'cycle_id':
                ws.cell(row=row, column=column, value=data[key])
                ws.cell(row=row, column=column).style = 'Hyperlink'
                ws.cell(row=row, column=column).hyperlink = f"https://gta.intel.com/#/workflow/testruns?stream=&page=1&count=5&cols=stream,testRunId,dagId,startTime,finishTime,status,cycleType,testSessions,&sorting%5Bid%5D=desc&filter%5Bid%5D={data[key]}"
                ws.cell(row=row, column=column).font = Font(color='0000FF')
            elif key == 'test_plan_id':
                ws.cell(row=row, column=column, value=data[key])
                ws.cell(row=row, column=column).style = 'Hyperlink'
                ws.cell(row=row, column=column).hyperlink = f"https://gta.intel.com/#/testplanning/plan/{data[key]}?timestamp={data['test_plan_timestamp']}"
                ws.cell(row=row, column=column).font = Font(color='0000FF')
            elif key == 'build_name':
                ws.cell(row=row, column=column, value=data[key])
                ws.cell(row=row, column=column).style = 'Hyperlink'
                ws.cell(row=row, column=column).hyperlink = data['build_link']
                ws.cell(row=row, column=column).font = Font(color='0000FF')
            elif key == 'gtax_id':
                ws.cell(row=row, column=column, value=data[key])
                ws.cell(row=row, column=column).style = 'Hyperlink'
                ws.cell(row=row, column=column).hyperlink = data['url']
                ws.cell(row=row, column=column).font = Font(color='0000FF')
            elif key == 'comparison_view':
                ws.cell(row=row, column=column, value=key)
                ws.cell(row=row, column=column).style = 'Hyperlink'
                ws.cell(row=row, column=column).hyperlink = data[key]
                ws.cell(row=row, column=column).font = Font(color='0000FF')
            else:
                ws.cell(row=row, column=column, value=data[key])
                font_bold = False
                if str(data[key]) in ['True', 'TIMEOUT', 'IN PROGRESS']:
                    font_bold = True
                ws.cell(row=row, column=column).font = Font(color=font_color, bold=font_bold)
            ws.cell(row=row, column=column).border = thin_border
            ws.cell(row=row, column=column).fill = PatternFill("solid", fgColor=fill_color)
            column += 1
    return row

def set_fill_by_data_conditions(data):
    fill_color = 'FFFFFF'
    if get_stat_key_value(data, 'tpt_met', if_not_found=1) == 0:
        fill_color = 'E0E0E0'
    if get_stat_key_value(data, 'setup_success', if_not_found=100) < 95.0:
        fill_color = 'CCAAAA'
    if get_stat_key_value(data, 'test_best_worst', if_not_found=0) > 5.0:
        fill_color = 'FFAAAA'
    return fill_color

def set_font_by_data_conditions(data):
    font_color = '000000'
    if get_stat_key_value(data, 'setup_success', if_not_found=100) < 95.0:
        font_color = 'FFFF00'
    return font_color

def get_cycles(dag_id, stream, submitter=None, date_from=None, date_to=None, status=None, count=100):
    cycles_list = list()
    url = f'http://gta.intel.com/api/workflow/v2/test_runs_dashboard?count={count}&filter%5Bdag_id%5D={dag_id}'
    print(f'geting cycles  dag_id:{dag_id}', end=' ') 
    if submitter:
        url += f'&filter%5Bsubmitter%5D={submitter}'
        print(f'submitter:{submitter}', end=' ')
    if stream:
        url += f'&filter%5Bstream%5D={stream}'
        print(f'stream:{stream}', end=' ')
    if status in ['NEW', 'IN PROGRESS' ,'COMPLETED', 'ERROR', 'TIMEOUT', 'INCOMPLETE']:
        url += f'&filter%5Bstatus%5D={status}'
        print(f'status:{status}', end=' ')
    if status == 'DEFAULT':
        url += '&filter%5Bstatus%5D!=NEW&filter%5Bstatus%5D!=IN%20PROGRESS'
        print('status: !NEW AND !IN PROGRESS', end=' ')
    if date_from:
        url += f'&filter%5Btest_run_start_date%5D=%3E{date_from}T00%3A00%3A00Z'        
        print(f'date_from:{date_from}', end=' ')
    if date_to:
        url += f'&filter%5Btest_run_start_date%5D=%3C{date_to}T23%3A59%3A59Z'
        print(f'date_to:{date_to}', end=' ')
    url += '&sorting%5Bid%5D=desc&filter%5Bexact_dag_id%5D=true'
    print(' ')
    #print(url)
    cycles_list = get_gta_cycles_pages(url)
    return cycles_list


def get_ci_cycle_stat(ci_cycles):
    ci_cycle_stat = list()
    for cycle in ci_cycles:
        cycle_dict = dict()
        cycle_dict.update({'cycle_id':cycle['id']})
        for key in ['dag_id', 'submitter', 'start_date', 'status_date', 'status', 'cycle_type', 'label']:
            cycle_dict.update({key:cycle[key]})
        cycle_dict.update({'end_date':get_end_date(cycle['status_date'], cycle['status'])})    
        cycle_dict.update({'duration':get_duration(cycle['start_date'], cycle_dict['end_date'])})
        cycle_dict.update({'duration_seconds':get_duration(cycle['start_date'], cycle_dict['end_date'], seconds=True)})
        cycle_dict.update({'stream_name':cycle['builds'][0]['stream_name']})
        cycle_dict.update({'build_name':cycle['builds'][0]['name']})
        cycle_dict.update({'build_id':cycle['builds'][0]['external_id']})
        cycle_dict.update({'build_link':cycle['builds'][0]['url']})
        cycle_dict.update({'build_start_date':cycle['builds'][0]['start_date']})
        cycle_dict.update({'started_after_build':get_duration(cycle['builds'][0]['status_date'], cycle['start_date'])})
        cycle_dict.update({'started_after_build_seconds':get_duration(cycle['builds'][0]['status_date'], cycle['start_date'], seconds=True)})
        cycle_dict.update({'lgcb_bad':get_flag_status(cycle['test_sessions'], 'lgcb_bad')})
        cycle_dict.update({'has_regressions':get_flag_status(cycle['test_sessions'], 'has_regressions')})
        cycle_dict.update({'comparison_view':get_cycle_comparison_view_link(cycle['builds'][0]['name'], cycle['id'])})
        cycle_dict.update({'WW':convert_to_intel_ww(cycle['start_date'])})
        ci_cycle_stat.append(cycle_dict)
    return ci_cycle_stat

def get_flag_status(session_list, flag_name):
    flag_status = False
    for session in session_list:
        if session[flag_name]:
            flag_status = True
            break
    return flag_status

def get_ci_cycle_session_stat(ci_cycles):
    session_stat = list()
    for cycle in ci_cycles:
        for session in cycle['test_sessions']:
            session_dict = dict()
            session_dict.update({'stream_name':cycle['builds'][0]['stream_name']})
            session_dict.update({'build_name':cycle['builds'][0]['name']})
            session_dict.update({'build_id':cycle['builds'][0]['external_id']})
            session_dict.update({'build_link':cycle['builds'][0]['url']})
            session_dict.update({'cycle_id':cycle['id']})
            for key in session.keys():
                if key == 'id':
                    session_dict.update({'session_id':session[key]})
                elif key == 'external_id':
                    session_dict.update({'gtax_id':session[key]})
                elif key == 'url':
                    url = session[key]
                    session_dict.update({'url':url})
                    if url:
                        instance_start = url.find('//') + 2
                        instance_end = url.find('.intel.com')
                        session_dict.update({'gtax_instance':url[instance_start:instance_end]})
                    else:
                        session_dict.update({'gtax_instance':''})
                elif key == 'cycle_type':
                    session_dict.update({key:cycle[key]})
                elif key == 'test_session_failure':
                    #http://gta.intel.com/api/workflow/v2/test_runs_dashboard?page=1&count=10&filter%5Bdag_id%5D=CI_SMOKE__gfx-driver__master&filter%5Bbuild_name%5D=gfx-driver-ci-master-7952&filter%5Bstatus%5D=COMPLETED
                    session_dict.update({'comments':str(session['test_session_failure']['comment'])})
                    if session['test_session_failure']['failure_reasons_set']:
                        failure_reasons_set = set()
                        for reasons in session['test_session_failure']['failure_reasons_set']:
                            failure_reasons_set.add(reasons['name'] )
                        session_dict.update({'failure_reasons':','.join(failure_reasons_set)})    
                    else:
                        session_dict.update({'failure_reasons':'None'}) 
                else:
                    session_dict.update({key:session[key]})
            session_dict.update({'end_date':get_end_date(session['status_date'], session['status'])}) 
            session_dict.update({'duration':get_duration(session['start_date'], session_dict['end_date'])})
            session_dict.update({'duration_seconds':get_duration(session['start_date'], session_dict['end_date'], seconds=True)})
            if session_dict['duration_seconds']:
                session_dict.update({'duration_min':round(session_dict['duration_seconds']/60, 0)})
            else:
                session_dict.update({'duration_min':'N/A'})
            session_dict.update({'comparison_view':get_session_comparison_view_link(cycle['builds'][0]['name'], cycle['id'], session['id'])})
            session_dict.update({'WW':convert_to_intel_ww(cycle['start_date'])})
            session_stat.append(session_dict)
    #print(session_stat[0].keys())
    return session_stat

def get_cycle_comparison_view_link(build_name, cycle_id, worst=False):
    url = 'https://gta.intel.com/#/reports/comparison-view?diffMode=false&reRuns[name]=all&table[page]=1&table[count]=25&visibleColumns[]=Item%20Name&visibleColumns[]=Args&visibleColumns[]=OS&visibleColumns[]=Platform&visibleColumns[]=Vertical&visibleColumns[]=Test%20Run&visibleColumns[]=Test%20Session&compareFields[]=compare_id&settingsFromUrl=true&tagExcept[0][name]=notAnIssue&tagExcept[1][name]=obsoleted&tagExcept[2][name]=iteration&tagExcept[3][name]=isolation'
    if worst:
        url += '&rerun=Worst'
    url += f'&globalFilterId=&builds[0][name]={build_name}&builds[0][key]={build_name}&filters[test_run][0][name]={cycle_id}'
    return url

def get_session_comparison_view_link(build_name, cycle_id, session_id):
    url = get_cycle_comparison_view_link(build_name, cycle_id, worst=True)
    url += f'&filters[test_session][0][name]={session_id}'
    return url

def update_sessions_with_tp_data(ci_test_sessions_stat):
    test_plan_dict = get_test_plan_dict(ci_test_sessions_stat)
    for session in ci_test_sessions_stat:
        if session['test_plan_id'] in test_plan_dict.keys():
            session.update({'test_plan_name':get_tp_name(session['test_plan_id'], test_plan_dict)})
            session.update({'environment':get_tp_env(session['test_plan_id'], test_plan_dict)})
        else:
            # for null tp id
            session.update({'test_plan_name':session['label']})
            session.update({'environment':'unknown'})

def get_end_date(status_date, status, time_format='%Y-%m-%dT%H:%M:%SZ'):
    end_date = status_date
    if status == 'IN PROGRESS':
        time_end = datetime.utcnow()
        end_date = time_end.strftime(time_format)
    return end_date

def get_tp_name(tp_id, test_plan_dict):
    tp_name = None
    if tp_id:
        tp_name = test_plan_dict[tp_id]['name']
    return tp_name

def get_tp_env(tp_id, test_plan_dict):
    tp_env = 'unknown'
    if tp_id:
        tp_env = test_plan_dict[tp_id]['env']
    return tp_env

def get_test_plan_dict(ci_session_stat):
    test_plan_dict = dict()
    tp_cache_file = 'smoke_stat_tp_dict.cache'
    if os.path.isfile(tp_cache_file):
        with open(tp_cache_file, 'rb') as cache:
            test_plan_dict = pickle.load(cache)
    test_plan_ids_set = set()
    for session in ci_session_stat:
        if session['test_plan_id']:
            test_plan_ids_set.add(session['test_plan_id'])
    for tp_id in test_plan_ids_set:
        if tp_id not in test_plan_dict.keys():
                tp_data = get_test_plan_data(tp_id)
                test_plan_dict.update({tp_id:{'name':tp_data['name'], 'env':tp_data['env']}})
    with open(tp_cache_file, 'w+b') as cache:
        pickle.dump(test_plan_dict, cache) 
    return test_plan_dict

def get_test_plan_data(test_plan_id):
    url = f'http://gta.intel.com/api/tp/v1/testplans/{test_plan_id}'
    print(f'geting test plan data for {test_plan_id}')
    response = get_gta_data(url)
    test_plan = response.json()
    tp_data = dict()
    tp_env = 'unknown'
    tp_data.update({'name':test_plan['name']})
    for attribute in test_plan['attributes']:
        if attribute['name'] == 'Environment':
            tp_env = attribute['resolvedValue']
    tp_data.update({'env':tp_env})
    return tp_data

def get_duration(start_date, end_date, seconds=False, time_format='%Y-%m-%dT%H:%M:%SZ', timedate_format=False):
    duration = None
    if start_date and end_date:
        if timedate_format:
            time_start = start_date
            time_end = end_date
        else:
            time_start = datetime.strptime(start_date, time_format)
            time_end = datetime.strptime(end_date, time_format)
        if seconds:
            duration = round((time_end - time_start).total_seconds())
        else:
            duration = str(time_end - time_start)
    return duration

def update_sessions_with_gtax_data(ci_test_sessions_stat, ci_jobs_data_dict, offline_mode=False):
    call_count = 0
    call_total = len(ci_test_sessions_stat)
    for session in ci_test_sessions_stat:
        call_count += 1
        if session['gtax_id']:
            print(f'getting gtax data {round(call_count/call_total*100,2)}%   ', end='\r', flush=True)
            #get_gtax_jobs_ids(session['gtax_instance'], session['gtax_id'], offline_mode)
            #jobs_full_info = get_gtax_jobs_full_info(session['gtax_instance'], session['gtax_id'], offline_mode)
            #jobs_full_info = list()

            jobs_full_info = get_gtax_jobs_full_info(session['gtax_instance'], session['gtax_id'], offline_mode)
            #session_task_dict = get_session_tasks_dict(jobs_full_info)
            session_jobs_dict = get_session_jobs_dict(jobs_full_info, session['gtax_instance'], session['build_name'])
            ci_jobs_data_dict.update(session_jobs_dict)
            # init setup success rate
            #init_setup_stat = get_gtax_task_stat(session_task_dict, 'setup', submission_type=['init'])
            #session.update({'setup_tasks_total':get_stat_key_value(init_setup_stat, 'total')})
            #session.update({'setup_tasks_success':get_stat_key_value(init_setup_stat, 'passed') + get_stat_key_value(init_setup_stat, 'blocked')})
            #if session['setup_tasks_total'] > 0:
            #    session.update({'setup_success':round(session['setup_tasks_success']/session['setup_tasks_total']*100, 2)})
            #else:
            #    session.update({'setup_success':'no setup tasks'})
            # test success rate (best-worst)
            #init_test_stat = get_gtax_task_stat(session_task_dict, 'test', submission_type=['init'])
            #rerun_test_stat = get_gtax_task_stat(session_task_dict, 'test', submission_type=['resume', 'rerun'])
            #session.update({'test_tasks_total':get_stat_key_value(init_test_stat, 'total')})
            #session.update({'test_tasks_worst_success':get_stat_key_value(init_test_stat, 'passed') + get_stat_key_value(init_test_stat, 'blocked')})
            #session.update({'test_tasks_best_success':get_stat_key_value(init_test_stat, 'passed') + get_stat_key_value(rerun_test_stat, 'passed') + get_stat_key_value(init_test_stat, 'blocked')})
            #if session['test_tasks_total'] > 0:
            #    session.update({'test_best':round(session['test_tasks_best_success']/session['test_tasks_total']*100, 2)})
            #    session.update({'test_worst':round(session['test_tasks_worst_success']/session['test_tasks_total']*100, 2)})
            #    session.update({'test_best_worst':round((session['test_tasks_best_success']-session['test_tasks_worst_success'])/session['test_tasks_total']*100, 2)})
            #else:
            #    session.update({'test_best':'no test tasks'})
            #    session.update({'test_worst':'no test tasks'})
            #    session.update({'test_best_worst':'no test tasks'})
            # max job duration in min.
            max_job_duration = get_max_job_duration(session_jobs_dict)
            session.update({'max_job_duration':max_job_duration})
            session.update({'tpt_met':get_tpt_met(session['duration_min'], 60)})
        else:
            for key in ['setup_tasks_total', 'setup_tasks_success', 'setup_success', 'test_tasks_total', 'test_tasks_worst_success', 'test_tasks_best_success', 'test_best', 'test_worst', 'test_best_worst', 'max_job_duration', 'tpt_met']:
                session.update({key:'N/A'})
    print('\n')

def get_max_job_duration(session_jobs_dict):
    max_duration = 0
    for job in session_jobs_dict:
        if isinstance(session_jobs_dict[job]['duration'], (float, int)):
            if int(session_jobs_dict[job]['duration']) > max_duration:
                max_duration = round(int(session_jobs_dict[job]['duration'])/60, 0)
    return max_duration

def get_tpt_met(max_job_duration, limit):
    if max_job_duration > limit:
        tpt_met = 0
    else:
        tpt_met = 100
    return tpt_met

   
def get_gtax_task_stat(session_task_dict, phase, submission_type=[]):
    task_stat_list = list()
    task_status_set = set()
    for task in session_task_dict:
        if session_task_dict[task]['phase'] == phase:
            if len(submission_type) > 0:
                if session_task_dict[task]['submission_type'] in submission_type:
                    task_stat_list.append(session_task_dict[task]['status'])
                    task_status_set.add(session_task_dict[task]['status'])
            else:
                task_stat_list.append(session_task_dict[task]['status'])
                task_status_set.add(session_task_dict[task]['status'])
    task_stat_dict = dict()
    for status in task_status_set:
        task_stat_dict.update({f'{status}':task_stat_list.count(status)})
    task_stat_dict.update({f'total':len(task_stat_list)})
    return(task_stat_dict)

def get_stat_key_value(stat_dict, key, if_not_found=0):
    stat_key_value = if_not_found
    if key in stat_dict.keys():
        stat_key_value = stat_dict[key]
    if not isinstance(stat_key_value, (float, int)):
        stat_key_value = if_not_found
    return stat_key_value
        
def get_session_tasks_dict(session_jobs_data):
    session_task_dict = dict()
    for job_data in session_jobs_data:
        for task_data in job_data['tasks']:
            task_dict = dict()
            for key in ['id', 'job_id', 'task_result_id', 'task_definition_id']:
                task_dict.update({key:task_data[key]})
            if job_data['client']:
                task_dict.update({'client_name':job_data['client']['name']})
                task_dict.update({'client_id':job_data['client']['id']})
            else:
                task_dict.update({'client_name':'unassigned'})
                task_dict.update({'client_id':'unassigned'})
            if 'task_result' in task_data.keys():
                for key in ['submission_type', 'gta_result_key', 'started_date', 'completed_date', 'status', 'gta_status', 'primary_reason', 'publishing_dumps']:
                    if key == 'gta_result_key':
                        task_dict.update({'phase':get_task_phase(task_data['task_result']['gta_result_key'])})
                        task_dict.update({key:task_data['task_result'][key]})
                    else:
                        task_dict.update({key:task_data['task_result'][key]})
            if 'task_definition' in task_data.keys():
                task_dict.update({'command':task_data['task_definition']['command']})
            if 'started_date' in task_dict.keys() and 'completed_date' in task_dict.keys():
                task_dict.update({'duration':get_duration(task_dict['started_date'], task_dict['completed_date'], time_format='%Y-%m-%d %H:%M:%S.%f')})
                task_dict.update({'duration_seconds':get_duration(task_dict['started_date'], task_dict['completed_date'], seconds=True, time_format='%Y-%m-%d %H:%M:%S.%f')})
            session_task_dict.update({task_data['id']:task_dict})
    return session_task_dict

def get_session_jobs_dict(session_jobs_data, gtax_instance, build_name):
    session_jobs_dict = dict()
    for job_data in session_jobs_data:
        job_dict = dict()
        for key in ['id', 'status', 'name', 'result', 'submission_type', 'submitted_date', 'started_date', 'completed_date', 'duration', 'scheduler_account_name', 'dumps_produced']:
            job_dict.update({key:job_data[key]})
        job_dict.update({'build_name':build_name})
        for key in ['client']:
            for sub_key in ['name','id']:
                key_name = f'{key}_{sub_key}'
                job_dict.update({key_name:job_data[key][sub_key]})
            if 'csq' in job_data.keys():
                job_dict.update({'platform':get_client_platform(job_data['csq']['query'])})
                job_dict.update({'pool':get_client_pool(job_data['csq']['query'])})
                if job_dict['platform'] == 'unknown':
                    if job_dict['client_id']:
                        job_dict.update({'platform':get_client_property_by_client_id(job_dict['client_id'], 'platform', gtax_instance)})
                #if job_dict['pool'] == 'unknown':
                #    if job_dict['client_id']:
                #        job_dict.update({'pool':get_client_property_by_client_id(job_dict['client_id'], 'pool', gtax_instance)})
            else:
                job_dict.update({'platform':get_client_property_by_client_id(job_dict['client_id'], 'platform', gtax_instance)})
                job_dict.update({'pool':get_client_property_by_client_id(job_dict['client_id'], 'pool', gtax_instance)})
        if 'results_summary' not in job_data.keys():
            print(job_data)
        for key in ['results_summary']:
            for sub_key in job_data[key].keys():
                key_name = f'{sub_key}'
                job_dict.update({key_name:job_data[key][sub_key]})
        if 'csq' in job_data.keys():
            job_dict.update({'csq':job_data['csq']['query']})
        else:
            job_dict.update({'csq':'not_found'})
        job_dict.update({'gtax_instance':gtax_instance})
        session_jobs_dict.update({job_data['id']:job_dict})
    return session_jobs_dict


""" def get_session_jobs_dict(session_jobs_data):
    session_jobs_dict = dict()
    for job_data in session_jobs_data:
        job_dict = dict()
        for key in ['id', 'status', 'name', 'result', 'submission_type', 'submitted_date', 'started_date', 'completed_date', 'duration', 'scheduler_account_name', 'dumps_produced']:
            job_dict.update({key:job_data[key]})
        if 'client' in job_data.keys():
            key = 'client'
            for sub_key in ['name','id']:
                key_name = f'{key}_{sub_key}'
                if type(job_data[key][sub_key]) == 'NoneType':
                    print(job_data)
                else:
                    job_dict.update({key_name:job_data[key][sub_key]}) 
                #job_dict.update({key_name:job_data[key][sub_key]})
        if 'csq' in job_data.keys():
            job_dict.update({'platform':get_client_platform(job_data['csq']['query'])})
            job_dict.update({'pool':get_client_pool(job_data['csq']['query'])})
        if 'results_summary' in job_data.keys():
            key = 'results_summary'
            if job_data[key].keys() == None:
                print(job_data[key].keys())
            for sub_key in job_data[key].keys():
                key_name = f'{sub_key}'
                if type(job_data[key][sub_key]) == 'NoneType':
                    print(job_data)
                else: 
                    job_dict.update({key_name:job_data[key][sub_key]})        
        if 'csq' in job_data.keys():
            csq = job_data['csq']['query']
            job_dict.update({'csq':csq})
        session_jobs_dict.update({job_data['id']:job_dict})
    return session_jobs_dict    """ 

def get_client_platform(csq):
    platform_regex = re.compile(r"('platform' = '.*?')") 
    platform = 'unknown'
    if platform_regex.search(csq):
        platform = platform_regex.search(csq).group(0)[14:-1]
    return platform

def get_client_property_by_client_id(client_id, property_name, gtax_instance):
    property_value = 'unknown'
    app_path = os.path.dirname(os.path.abspath(sys.argv[0]))
    cache_path = os.path.join(app_path, '_cache')
    cache_file_name = os.path.join(cache_path, f"client_{gtax_instance}_{client_id}_properties.cache")
    if os.path.isfile(cache_file_name):
        with open(cache_file_name, 'rb') as cache:
            client_properties = pickle.load(cache)
    else:
        url = f'http://{gtax_instance}.intel.com/api/v1/clients/{client_id}/properties'
        client_properties = get_gtax_data(url)
        if not os.path.exists(cache_path):
            os.makedirs(cache_path)
        with open(cache_file_name, 'w+b') as cache:
            pickle.dump(client_properties, cache)
    if property_name in client_properties['user_defined'].keys():
        property_value = client_properties['user_defined'][property_name]
    else:
        if property_name in client_properties['auto_detected'].keys():
            property_value = client_properties['auto_detected'][property_name]
    return property_value

def get_gtax_jobs_full___info(gtax_instance, job_set_session_id, offline_mode=False):
    app_path = os.path.dirname(os.path.abspath(sys.argv[0]))
    cache_path = os.path.join(app_path, '_cache')
    cache_file_name = os.path.join(cache_path, f"gtax_{job_set_session_id}_full_info.cache")
    url = f'http://{gtax_instance}.intel.com/api/v1/jobs?include_tasks=false&include_taskml=false&include_csq=true&full_info=false&jobset_session_ids={job_set_session_id}'
    if offline_mode and os.path.isfile(cache_file_name):
        with open(cache_file_name, 'rb') as cache:
            jobs_full_info = pickle.load(cache)
        #print(cache_file_name)
    else:
        #print(f'be patient!! - getting full info of tasks for job set session: {job_set_session_id}')
        #url = f'http://{gtax_instance}.intel.com/api/v1/jobs?include_tasks=true&include_csq=true&include_phases_task_counts=true&full_info=true&jobset_session_ids={job_set_session_id}'
        url = f'http://{gtax_instance}.intel.com/api/v1/jobs?include_tasks=false&include_taskml=false&include_csq=true&full_info=false&jobset_session_ids={job_set_session_id}'
        print(url)
        data = get_gtax_data(url)
        jobs_full_info = data['data']
        if not os.path.exists(cache_path):
            os.makedirs(cache_path)
        with open(cache_file_name, 'w+b') as cache:
            pickle.dump(jobs_full_info, cache) 
    return jobs_full_info

def get_client_pool(csq):
    platform_regex = re.compile(r"('pool' = '.*?')") 
    platform = 'unknown'
    if platform_regex.search(csq):
        platform = platform_regex.search(csq).group(0)[10:-1]
    return platform   

def get_task_phase(gta_result_key):
    phase = 'unknown'
    if gta_result_key: 
        for phase_name in ['teardown', 'setup', 'test']:
            if gta_result_key.find(phase_name) > 0:
                phase = phase_name
    else:
        phase = ''
    return phase

def get_gtax_jobs_full_info(gtax_instance, job_set_session_id, offline_mode=False):
    app_path = os.path.dirname(os.path.abspath(sys.argv[0]))
    cache_path = os.path.join(app_path, '_cache')
    cache_file_name = os.path.join(cache_path, f"gtax_{job_set_session_id}_full_info.cache")
    url = f'http://{gtax_instance}.intel.com/api/v1/jobs?include_tasks=false&include_taskml=false&include_csq=true&full_info=false&jobset_session_ids={job_set_session_id}'
    if offline_mode and os.path.isfile(cache_file_name):
        with open(cache_file_name, 'rb') as cache:
            jobs_full_info = pickle.load(cache)
        #print(cache_file_name)
    else:
        #print(f'be patient!! - getting full info of tasks for job set session: {job_set_session_id}')
        #url = f'http://{gtax_instance}.intel.com/api/v1/jobs?include_tasks=true&include_csq=true&include_phases_task_counts=true&full_info=true&jobset_session_ids={job_set_session_id}'
        url = f'http://{gtax_instance}.intel.com/api/v1/jobs?include_tasks=false&include_taskml=false&include_csq=true&full_info=false&jobset_session_ids={job_set_session_id}'
        print(url)
        data = get_gtax_data(url)
        jobs_full_info = data['data']
        if not os.path.exists(cache_path):
            os.makedirs(cache_path)
        with open(cache_file_name, 'w+b') as cache:
            pickle.dump(jobs_full_info, cache) 
    return jobs_full_info

def post_gta_data(url, data):
    output = None
    headers = { 'Content-type': 'application/json' }
    response = requests.put(url, data, headers=headers, proxies={'http': 'http://proxy-chain.intel.com:911', 'https': 'http://proxy-chain.intel.com:912' })
    if response.status_code == 200:
        output = response
    else:
        print(response.status_code)
        print(response)
    return output

def get_gta_data(url, page=None):
    output = None
    headers = { 'Content-type': 'application/json' }
    if page:
        url += f'&page={page}'
    response = requests.get(url, headers=headers, proxies={'http': 'http://proxy-chain.intel.com:911', 'https': 'http://proxy-chain.intel.com:912' })
    if response.status_code == 200:
        output = response
    else:
        print(response.status_code)
        print(response)
    return output

def get_gta_cycles_pages(url):
    cycles_list = list()
    page = 1
    response = get_gta_data(url, page=page)
    cycles = response.json()
    cycles_list.extend(cycles['items'])
    pages = cycles['pages']
    if pages > 1:
        for i in range(page+1, pages+1):
            response = get_gta_data(url, page=i)
            cycles = response.json()
            cycles_list.extend(cycles['items'])
    return cycles_list

def get_gtax_data(url):
    headers = { 'Content-type': 'application/json' }
    response = requests.get(url, headers=headers, proxies={'http': 'http://proxy-chain.intel.com:911', 'https': 'http://proxy-chain.intel.com:912' })
    return response.json()

def set_width(ws, min_width=0, max_width=100):
    column_widths = []
    column_heights = []
    for row_cnt, row in enumerate(ws.iter_rows()):
        column_heights.append(0)
        for col_cnt, cell in enumerate(row):
            if len(column_widths) <= col_cnt:
                column_widths.append(min_width)
            # Ignore cells that don't wrap
            if cell.alignment.wrap_text is False:
                continue
            if is_merged_cell(ws, cell):
                continue
            length, num_lines = calculate_cell_length(cell)
            if cell.alignment.text_rotation == 0:
                column_widths[col_cnt] = max(column_widths[col_cnt], length)
                row_height = 14 * num_lines
                column_heights[row_cnt] = max(column_heights[row_cnt], row_height)
            else:
                column_heights[row_cnt] = max(column_heights[row_cnt], 5 * length)
                col_width = 14 * num_lines
                column_widths[col_cnt] = max(column_widths[col_cnt], col_width)
    max_column_width_allowed = max_width
    for cnt, column_width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(cnt)].width = min(column_width, max_column_width_allowed)
    for cnt, row_height in enumerate(column_heights, 1):
        ws.row_dimensions[cnt].height = row_height

def is_merged_cell(ws, cell):
    for cell_range in ws.merged_cells:
        if cell.column >= cell_range.min_col and cell.column <= cell_range.max_col and \
           cell.row    >= cell_range.min_row and cell.row    <= cell_range.max_row:
            return True
    return False

def calculate_cell_length(cell):
    num_lines = 1
    if isinstance(cell.value, str):
        if 'HYPERLINK' in cell.value:
            pieces = cell.value.split(',')
            val = pieces[-1].split('"')[1]
            length = len(val)
        else:
            if '\n' in cell.value:
                num_lines = len(cell.value.split('\n'))
                length = max([len(s) for s in cell.value.split('\n')])
            else:
                length = len(cell.value)
    elif isinstance(cell.value, bool):
        length = len('False')
    elif isinstance(cell.value, int):
        if cell.value == 0:
            length = 1
        else:
            if cell.value < 0:
                val = -1 * cell.value
            else:
                val = cell.value
            length = math.ceil(math.log10(abs(cell.value)))
    else:
        length = 0
    length += 2
    return length, num_lines

def get_filename_timestamp():
    ts = str(datetime.now())
    ts = ts.replace(':', '_')
    return ts[5:-7]

def get_date_n_days_ago(days_ago):
    ts = str(datetime.now()- timedelta(days=days_ago))
    return ts[0:10]

def check_user_input(user_input):
    try:
        datetime.strptime(user_input, '%Y-%m-%d')
    except ValueError:
        return False
    return True

def convert_to_intel_ww(date_string):
    #correct to Intel WW
    date_time = datetime.strptime(date_string, '%Y-%m-%dT%H:%M:%SZ')
    iso_year = date_time.isocalendar()[0]
    iso_ww = date_time.isocalendar()[1]
    max_iso_ww = datetime.strptime(f'{iso_year}-12-31', '%Y-%m-%d').isocalendar()[1]
    first_day_of_the_year = int(datetime.strptime(f'{iso_year}-01-01', '%Y-%m-%d').strftime("%w"))
    last_day_of_the_year = int(datetime.strptime(f'{iso_year}-12-31', '%Y-%m-%d').strftime("%w"))
    day_of_year = int(date_time.strftime("%j"))
    if first_day_of_the_year in [0, 1, 2, 3, 4]:
        intel_ww = iso_ww
    else:
        intel_ww = iso_ww + 1
    if iso_ww == max_iso_ww:
        if last_day_of_the_year != 6:
            intel_ww = 1
    if day_of_year == 1 and intel_ww > 1:
        intel_ww = 1
    return f'{str(iso_year)[2:]}WW{str(intel_ww).zfill(2)}'

if __name__ == '__main__':
    ts = get_filename_timestamp()
    filename = f'smoke_tpt_report_{ts}.xlsx'
    ap = argparse.ArgumentParser()
    ap.add_argument('-s', '-status', default='DEFAULT', help='status filter values: [NEW, IN PROGRESS ,COMPLETED, ERROR, TIMEOUT, INCOMPLETE]', required=False)
    #ap.add_argument('-c', '-count', default='100', help='max cycles count per gta request', required=False)
    ap.add_argument('-i', '-interactive', action='store_true', default=False, help='interactive mode', required=False)
    ap.add_argument('-f', '-date_from', default=get_date_n_days_ago(3), help='date from format yyyy-mm-dd', required=False)
    ap.add_argument('-t', '-date_to', default=get_date_n_days_ago(1), help='date to format yyyy-mm-dd', required=False)
    ap.add_argument('-submitter', default='sys_gtawf', help='submitter default:sys_gtawf', required=False)
    ap.add_argument('-dag', default='CI_DAILY__gfx-driver__master__silicon:gfx-driver__master', help='DAG stream default: CI_DAILY__gfx-driver__master__silicon:gfx-driver__master', required=False)
    ap.add_argument('-o', '-output', default=filename, help='output file name', required=False)
    #ap.add_argument('-no_cache', action='store_false', default=True, help='no cache for gtax api calls (could be use to update cache)', required=False)
    parsed = ap.parse_args()
    date_from = parsed.f
    date_to = parsed.t
    if parsed.i:
        date_from = input(f'provide from date (YYYY-MM-DD) [enter use default={get_date_n_days_ago(7)}]:')
        if not check_user_input(date_from):
            date_from = get_date_n_days_ago(7)
        date_to = input(f'provide to date (YYYY-MM-DD) [enter use default={get_date_n_days_ago(1)}]:')
        if not check_user_input(date_to):
            date_to = get_date_n_days_ago(1)
    main(parsed.s, 100, date_from, date_to, parsed.dag, parsed.submitter, parsed.o, True)
