import sys
import math
import os.path
import argparse
import pickle
from datetime import datetime
from datetime import timedelta
from operator import itemgetter
import time
import json
import requests
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList

def main(status, filename, branches=False, offline_mode=False):
    ci_cycle_stat = list()
    ci_test_sessions_stat = list()

    #MASTER
    cache_file1_name = f'cycles_master_tmp.cache'
    cache_file2_name = f'sessions_master_tmp.cache'
    if offline_mode and os.path.isfile(cache_file1_name) and os.path.isfile(cache_file2_name):
        with open(cache_file1_name, 'rb') as cache:
            ci_cycle_stat = pickle.load(cache)
        with open(cache_file2_name, 'rb') as cache:
            ci_test_sessions_stat = pickle.load(cache)
    else:
        update_stats(ci_cycle_stat, ci_test_sessions_stat, get_cycles('CI_SMOKE__gfx-driver__master', 'gfx-driver__master', 'sys_gtawf', status, 30))
        update_stats(ci_cycle_stat, ci_test_sessions_stat, get_cycles('CI_DAILY__gfx-driver__master__silicon', 'gfx-driver__master', 'sys_gtawf', status, 20))
        update_stats(ci_cycle_stat, ci_test_sessions_stat, get_cycles('CI_WEEKLY__gfx-driver__master__uber_gft', 'gfx-driver__master', 'sys_gtawf', status, 10))
        update_stats(ci_cycle_stat, ci_test_sessions_stat, get_cycles('CI_WEEKLY__gfx-driver__master__silicon-ehl', 'gfx-driver__master', 'sys_gtawf', status, 10))
        with open(cache_file1_name, 'w+b') as cache:
            pickle.dump(ci_cycle_stat, cache) 
        with open(cache_file2_name, 'w+b') as cache:
            pickle.dump(ci_test_sessions_stat, cache) 

    update_sessions_with_tp_name(ci_test_sessions_stat)

    output_wb = openpyxl.Workbook()
    ws = output_wb.active
    ws.title = 'ci_cycles-master'
    cycle_columns = ['cycle_id', 'build_name', 'cycle_type', 'dag_id', 'status', 'start_date', 'end_date', 'duration', 'build_start_date', 'started_after_build', 'lgcb_bad', 'has_regressions', 'comparison_view']
    print_data_table(ws, ci_cycle_stat, cycle_columns, 1, 1)
    ws.auto_filter.ref = ws.dimensions
    set_width(ws, max_width=600)
    ws = output_wb.create_sheet('ci_sessions-master')
    session_columns = ['gtax_id', 'build_name', 'cycle_type', 'test_plan_id', 'test_plan_name', 'status', 'progress', 'start_date', 'end_date', 'duration', 'remote_name', 'lgcb_bad', 'has_notes', 'has_regressions', 'cycle_id', 'comparison_view']
    print_data_table(ws, ci_test_sessions_stat, session_columns, 1, 1)
    ws.auto_filter.ref = ws.dimensions
    set_width(ws, max_width=600)

    #BRANCHES
    if branches:
        cache_file1_name = f'cycles_branches_tmp.cache'
        cache_file2_name = f'sessions_branches_tmp.cache'
        if offline_mode and os.path.isfile(cache_file1_name) and os.path.isfile(cache_file2_name):
            with open(cache_file1_name, 'rb') as cache:
                ci_cycle_stat = pickle.load(cache)
            with open(cache_file2_name, 'rb') as cache:
                ci_test_sessions_stat = pickle.load(cache)
        else:
            ci_cycle_stat = list()
            ci_test_sessions_stat = list()
            update_stats(ci_cycle_stat, ci_test_sessions_stat, get_cycles('ON_DEMAND_TESTS__gfx-driver__comp_media__Smoke', 'gfx-driver__comp_media', None, status, 30))

            update_stats(ci_cycle_stat, ci_test_sessions_stat, get_cycles('CI_smoke__ci-neo_master', 'ci-neo_master', 'sys_gtawf', status, 30))
            update_stats(ci_cycle_stat, ci_test_sessions_stat, get_cycles('CI_SMOKE__gfx-driver__comp_glsl', 'gfx-driver__comp_glsl', 'sys_gtawf', status, 30))
            update_stats(ci_cycle_stat, ci_test_sessions_stat, get_cycles('CI_smoke__gfx-driver__comp_igc', 'gfx-driver__comp_igc', 'sys_gtawf', status, 30))
            update_stats(ci_cycle_stat, ci_test_sessions_stat, get_cycles('CI_SMOKE__open-linux-driver-ci-dev_igc', 'open-linux-driver-ci-dev_igc', 'sys_gtawf', status, 30))
            update_stats(ci_cycle_stat, ci_test_sessions_stat, get_cycles('CI_SMOKE__ci-neo_embargo', 'ci-neo_embargo', 'sys_gtawf', status, 30))
            update_stats(ci_cycle_stat, ci_test_sessions_stat, get_cycles('CI_SMOKE__gfx-driver__comp_ogl', 'gfx-driver__comp_ogl', 'sys_gtawf', status, 30))
            update_stats(ci_cycle_stat, ci_test_sessions_stat, get_cycles('CI_SMOKE__gfx-driver__comp_vulkan', 'gfx-driver__comp_vulkan', 'sys_gtawf', status, 30))

            update_stats(ci_cycle_stat, ci_test_sessions_stat, get_cycles('CI_daily__ci-neo_master', 'ci-neo_master', 'sys_gtawf', status, 20))
            update_stats(ci_cycle_stat, ci_test_sessions_stat, get_cycles('CI_daily__gfx-driver__comp_glsl', 'gfx-driver__comp_glsl', 'sys_gtawf', status, 20))
            update_stats(ci_cycle_stat, ci_test_sessions_stat, get_cycles('CI_daily__gfx-driver__comp_igc', 'gfx-driver__comp_igc', 'sys_gtawf', status, 20))
            update_stats(ci_cycle_stat, ci_test_sessions_stat, get_cycles('CI_DAILY__open-linux-driver-ci-dev_igc', 'open-linux-driver-ci-dev_igc', 'sys_gtawf', status, 20))
            update_stats(ci_cycle_stat, ci_test_sessions_stat, get_cycles('CI_daily__ci-neo_embargo', 'ci-neo_embargo', 'sys_gtawf', status, 20))
            update_stats(ci_cycle_stat, ci_test_sessions_stat, get_cycles('CI_daily__gfx-driver__comp_ogl', 'gfx-driver__comp_ogl', 'sys_gtawf', status, 20))
            update_stats(ci_cycle_stat, ci_test_sessions_stat, get_cycles('CI_DAILY__gfx-driver__comp_vulkan', 'gfx-driver__comp_vulkan', 'sys_gtawf', status, 20))
            
            update_stats(ci_cycle_stat, ci_test_sessions_stat, get_cycles('CI_WEEKLY__ci-neo_master', 'ci-neo_master', 'sys_gtawf', status, 10))
            update_stats(ci_cycle_stat, ci_test_sessions_stat, get_cycles('CI_WEEKLY__ci-neo_embargo', 'ci-neo_embargo', 'sys_gtawf', status, 10))
            update_stats(ci_cycle_stat, ci_test_sessions_stat, get_cycles('CI_WEEKLY__gfx-driver__comp_ogl', 'gfx-driver__comp_ogl', 'sys_gtawf', status, 10))
            update_stats(ci_cycle_stat, ci_test_sessions_stat, get_cycles('CI_WEEKLY__gfx-driver__comp_vulkan', 'gfx-driver__comp_vulkan', 'sys_gtawf', status, 10))
            with open(cache_file1_name, 'w+b') as cache:
                pickle.dump(ci_cycle_stat, cache) 
            with open(cache_file2_name, 'w+b') as cache:
                pickle.dump(ci_test_sessions_stat, cache)

        update_sessions_with_tp_name(ci_test_sessions_stat)
        
        ws = output_wb.create_sheet('ci_cycles-branches')
        cycle_columns = ['cycle_id', 'build_name', 'stream_name', 'cycle_type', 'dag_id', 'status', 'start_date', 'end_date', 'duration', 'build_start_date', 'started_after_build', 'lgcb_bad', 'has_regressions', 'comparison_view']
        print_data_table(ws, ci_cycle_stat, cycle_columns, 1, 1)
        ws.auto_filter.ref = ws.dimensions
        set_width(ws, max_width=600)
        ws = output_wb.create_sheet('ci_sessions-branches')
        session_columns = ['gtax_id', 'build_name', 'stream_name', 'cycle_type', 'test_plan_id', 'test_plan_name', 'status', 'progress', 'start_date', 'end_date', 'duration', 'remote_name', 'lgcb_bad', 'has_notes', 'has_regressions', 'cycle_id', 'comparison_view']
        print_data_table(ws, ci_test_sessions_stat, session_columns, 1, 1)
        ws.auto_filter.ref = ws.dimensions
        set_width(ws, max_width=600)

    output_wb.save(filename)
    print(f'Saved the output file:{filename}')
    return 0

def update_stats(ci_cycle_stat, ci_test_sessions_stat, cycles):
    ci_cycle_stat += get_ci_cycle_stat(cycles)
    ci_test_sessions_stat += get_ci_cycle_session_stat(cycles)
    return 0

def print_data_table(ws, report_data, data_keys, start_row, start_column):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    row = start_row
    column = start_column
    for key in data_keys:
        ws.cell(row=row, column=column, value=key)
        ws.cell(row=row, column=column).font = Font(bold = True)
        column += 1
    for data in report_data:
        row += 1
        column = start_column
        fill_color = set_fill_by_data_conditions(data)
        font_color = set_font_by_data_conditions(data)
        for key in data_keys:
            if key == 'cycle_id':
                ws.cell(row=row, column=column, value=data[key])
                ws.cell(row=row, column=column).style = 'Hyperlink'
                ws.cell(row=row, column=column).hyperlink = f"https://gta.intel.com/#/workflow/testruns?stream=&page=1&count=5&cols=stream,testRunId,dagId,startTime,finishTime,status,cycleType,testSessions,&sorting%5Bid%5D=desc&filter%5Bid%5D={data[key]}"
                ws.cell(row=row, column=column).font = Font(color='0000FF')
            elif key == 'test_plan_id':
                ws.cell(row=row, column=column, value=data[key])
                ws.cell(row=row, column=column).style = 'Hyperlink'
                ws.cell(row=row, column=column).hyperlink = f"https://gta.intel.com/#/testplanning/plan/{data[key]}?timestamp={data['test_plan_timestamp']}"
                ws.cell(row=row, column=column).font = Font(color='0000FF')
            elif key == 'build_name':
                ws.cell(row=row, column=column, value=data[key])
                ws.cell(row=row, column=column).style = 'Hyperlink'
                ws.cell(row=row, column=column).hyperlink = data['build_link']
                ws.cell(row=row, column=column).font = Font(color='0000FF')
            elif key == 'gtax_id':
                ws.cell(row=row, column=column, value=data[key])
                ws.cell(row=row, column=column).style = 'Hyperlink'
                ws.cell(row=row, column=column).hyperlink = data['url']
                ws.cell(row=row, column=column).font = Font(color='0000FF')
            elif key == 'comparison_view':
                ws.cell(row=row, column=column, value=key)
                ws.cell(row=row, column=column).style = 'Hyperlink'
                ws.cell(row=row, column=column).hyperlink = data[key]
                ws.cell(row=row, column=column).font = Font(color='0000FF')
            else:
                ws.cell(row=row, column=column, value=data[key])
                font_bold = False
                if str(data[key]) in ['True', 'TIMEOUT', 'IN PROGRESS']:
                    font_bold = True
                ws.cell(row=row, column=column).font = Font(color=font_color, bold=font_bold)
            ws.cell(row=row, column=column).border = thin_border
            ws.cell(row=row, column=column).fill = PatternFill("solid", fgColor=fill_color)
            column += 1
    return row

def set_fill_by_data_conditions(data):
    fill_color = 'FFFFFF'
    if data['status'] == 'IN PROGRESS':
        fill_color = 'EFEFEF'
    if data['status'] == 'TIMEOUT':
        fill_color = 'FFFF00'
    if data['lgcb_bad'] == True:
        fill_color = 'CC0000'
    return fill_color

def set_font_by_data_conditions(data):
    font_color = '000000'
    if data['lgcb_bad'] == True:
        font_color = 'FFFF00'
    return font_color

def get_cycles(dag_id, stream, submitter=None, status=None, count=10):
    url = f'http://gta.intel.com/api/workflow/v2/test_runs_dashboard?page=1&count={count}&filter%5Bdag_id%5D={dag_id}'
    if submitter:
        url += f'&filter%5Bsubmitter%5D={submitter}'
    if stream:
        url += f'&filter%5Bstream%5D={stream}'
    if status in ['NEW', 'IN PROGRESS' ,'COMPLETED', 'ERROR', 'TIMEOUT', 'INCOMPLETE']:
        url += f'&filter%5Bstatus%5D={status}'
    url += '&sorting%5Bid%5D=desc&filter%5Bexact_dag_id%5D=true'
    print(f'geting cycles: {dag_id} {stream} status:{status}')
    #print(url)
    response = get_gta_data(url)
    cycles = response.json()
    return cycles['items']


def get_ci_cycle_stat(ci_cycles):
    ci_cycle_stat = list()
    for cycle in ci_cycles:
        cycle_dict = dict()
        cycle_dict.update({'cycle_id':cycle['id']})
        for key in ['dag_id', 'submitter', 'start_date', 'status_date', 'status', 'cycle_type', 'label']:
            cycle_dict.update({key:cycle[key]})
        cycle_dict.update({'end_date':get_end_date(cycle['status_date'], cycle['status'])})    
        cycle_dict.update({'duration':get_duration(cycle['start_date'], cycle_dict['end_date'])})
        cycle_dict.update({'duration_seconds':get_duration(cycle['start_date'], cycle_dict['end_date'], seconds=True)})
        cycle_dict.update({'stream_name':cycle['builds'][0]['stream_name']})
        cycle_dict.update({'build_name':cycle['builds'][0]['name']})
        cycle_dict.update({'build_id':cycle['builds'][0]['external_id']})
        cycle_dict.update({'build_link':cycle['builds'][0]['url']})
        cycle_dict.update({'build_start_date':cycle['builds'][0]['start_date']})
        cycle_dict.update({'started_after_build':get_duration(cycle['builds'][0]['status_date'], cycle['start_date'])})
        cycle_dict.update({'started_after_build_seconds':get_duration(cycle['builds'][0]['status_date'], cycle['start_date'], seconds=True)})
        cycle_dict.update({'lgcb_bad':get_flag_status(cycle['test_sessions'], 'lgcb_bad')})
        cycle_dict.update({'has_regressions':get_flag_status(cycle['test_sessions'], 'has_regressions')})
        cycle_dict.update({'comparison_view':get_cycle_comparison_view_link(cycle['builds'][0]['name'], cycle['id'])})
        ci_cycle_stat.append(cycle_dict)
    return ci_cycle_stat

def get_flag_status(session_list, flag_name):
    flag_status = False
    for session in session_list:
        if session[flag_name]:
            flag_status = True
            break
    return flag_status

def get_ci_cycle_session_stat(ci_cycles):
    session_stat = list()
    for cycle in ci_cycles:
        for session in cycle['test_sessions']:
            session_dict = dict()
            session_dict.update({'stream_name':cycle['builds'][0]['stream_name']})
            session_dict.update({'build_name':cycle['builds'][0]['name']})
            session_dict.update({'build_id':cycle['builds'][0]['external_id']})
            session_dict.update({'build_link':cycle['builds'][0]['url']})
            session_dict.update({'cycle_id':cycle['id']})
            for key in session.keys():
                if key == 'id':
                    session_dict.update({'session_id':session[key]})
                elif key == 'external_id':
                    session_dict.update({'gtax_id':session[key]})
                elif key == 'cycle_type':
                    session_dict.update({key:cycle[key]})
                else:
                    session_dict.update({key:session[key]})
            session_dict.update({'end_date':get_end_date(session['status_date'], session['status'])}) 
            session_dict.update({'duration':get_duration(session['start_date'], session_dict['end_date'])})
            session_dict.update({'duration_seconds':get_duration(session['start_date'], session_dict['end_date'], seconds=True)})
            session_dict.update({'comparison_view':get_session_comparison_view_link(cycle['builds'][0]['name'], cycle['id'], session['id'])})
            session_stat.append(session_dict)
    return session_stat

def get_cycle_comparison_view_link(build_name, cycle_id):
    url = f'https://gta.intel.com/#/reports/comparison-view?diffMode=false&reRuns[name]=all&table[page]=1&table[count]=25&visibleColumns[]=Item%20Name&visibleColumns[]=Args&visibleColumns[]=OS&visibleColumns[]=Platform&visibleColumns[]=Vertical&visibleColumns[]=Test%20Run&visibleColumns[]=Test%20Session&compareFields[]=compare_id&settingsFromUrl=true&tagExcept[0][name]=notAnIssue&tagExcept[1][name]=obsoleted&tagExcept[2][name]=iteration&tagExcept[3][name]=isolation&builds[0][name]={build_name}&builds[0][key]={build_name}&filters[test_run][0][name]={cycle_id}'
    return url

def get_session_comparison_view_link(build_name, cycle_id, session_id):
    url = get_cycle_comparison_view_link(build_name, cycle_id)
    url += f'&filters[test_session][0][name]={session_id}'
    return url

def update_sessions_with_tp_name(ci_test_sessions_stat):
    test_plan_dict = get_test_plan_dict(ci_test_sessions_stat)
    for session in ci_test_sessions_stat:
        test_plan_name = get_tp_name(session['test_plan_id'], test_plan_dict)
        if test_plan_name:
            session.update({'test_plan_name':test_plan_name})
        else:
            # for null tp id
            session.update({'test_plan_name':session['label']})

def get_end_date(status_date, status, time_format='%Y-%m-%dT%H:%M:%SZ'):
    end_date = status_date
    if status == 'IN PROGRESS':
        time_end = datetime.utcnow()
        end_date = time_end.strftime(time_format)
    return end_date

def get_tp_name(tp_id, test_plan_dict):
    tp_name = None
    if tp_id:
        tp_name = test_plan_dict[tp_id]
    return tp_name

def get_test_plan_dict(ci_session_stat):
    test_plan_dict = dict()
    tp_cache_file = 'tp_names.cache'
    if os.path.isfile(tp_cache_file):
        with open(tp_cache_file, 'rb') as cache:
            test_plan_dict = pickle.load(cache)
    test_plan_ids_set = set()
    for session in ci_session_stat:
        if session['test_plan_id']:
            test_plan_ids_set.add(session['test_plan_id'])
    for tp_id in test_plan_ids_set:
        if tp_id not in test_plan_dict.keys():
                test_plan_dict.update({tp_id:get_test_plan_name(tp_id)})
    with open(tp_cache_file, 'w+b') as cache:
        pickle.dump(test_plan_dict, cache) 
    return test_plan_dict

def get_test_plan_name(test_plan_id):
    url = f'http://gta.intel.com/api/tp/v1/testplans/{test_plan_id}'
    print(f'geting test plan name for {test_plan_id}')
    response = get_gta_data(url)
    test_plan = response.json()
    return test_plan['name']

def get_duration(start_date, end_date, seconds=False, time_format='%Y-%m-%dT%H:%M:%SZ'):
    duration = None
    if start_date and end_date:
        time_start = datetime.strptime(start_date, time_format)
        time_end = datetime.strptime(end_date, time_format)
        if seconds:
            duration = round((time_end - time_start).total_seconds())
        else:
            duration = str(time_end - time_start)
    return duration

def post_gta_data(url, data):
    output = None
    headers = { 'Content-type': 'application/json' }
    response = requests.put(url, data, headers=headers, proxies={'http': 'http://proxy-chain.intel.com:911', 'https': 'http://proxy-chain.intel.com:912' })
    if response.status_code == 200:
        output = response
    else:
        print(response.status_code)
        print(response)
    return output

def get_gta_data(url):
    output = None
    headers = { 'Content-type': 'application/json' }
    response = requests.get(url, headers=headers, proxies={'http': 'http://proxy-chain.intel.com:911', 'https': 'http://proxy-chain.intel.com:912' })
    if response.status_code == 200:
        output = response
    else:
        print(response.status_code)
        print(response)
    return output

def set_width(ws, min_width=0, max_width=100):
    column_widths = []
    column_heights = []
    for row_cnt, row in enumerate(ws.iter_rows()):
        column_heights.append(0)
        for col_cnt, cell in enumerate(row):
            if len(column_widths) <= col_cnt:
                column_widths.append(min_width)
            # Ignore cells that don't wrap
            if cell.alignment.wrap_text is False:
                continue
            if is_merged_cell(ws, cell):
                continue
            length, num_lines = calculate_cell_length(cell)
            if cell.alignment.text_rotation == 0:
                column_widths[col_cnt] = max(column_widths[col_cnt], length)
                row_height = 14 * num_lines
                column_heights[row_cnt] = max(column_heights[row_cnt], row_height)
            else:
                column_heights[row_cnt] = max(column_heights[row_cnt], 5 * length)
                col_width = 14 * num_lines
                column_widths[col_cnt] = max(column_widths[col_cnt], col_width)
    max_column_width_allowed = max_width
    for cnt, column_width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(cnt)].width = min(column_width, max_column_width_allowed)
    for cnt, row_height in enumerate(column_heights, 1):
        ws.row_dimensions[cnt].height = row_height

def is_merged_cell(ws, cell):
    for cell_range in ws.merged_cells:
        if cell.column >= cell_range.min_col and cell.column <= cell_range.max_col and \
           cell.row    >= cell_range.min_row and cell.row    <= cell_range.max_row:
            return True
    return False

def calculate_cell_length(cell):
    num_lines = 1
    if isinstance(cell.value, str):
        if 'HYPERLINK' in cell.value:
            pieces = cell.value.split(',')
            val = pieces[-1].split('"')[1]
            length = len(val)
        else:
            if '\n' in cell.value:
                num_lines = len(cell.value.split('\n'))
                length = max([len(s) for s in cell.value.split('\n')])
            else:
                length = len(cell.value)
    elif isinstance(cell.value, bool):
        length = len('False')
    elif isinstance(cell.value, int):
        if cell.value == 0:
            length = 1
        else:
            if cell.value < 0:
                val = -1 * cell.value
            else:
                val = cell.value
            length = math.ceil(math.log10(abs(cell.value)))
    else:
        length = 0
    length += 2
    return length, num_lines

def get_filename_timestamp():
    ts = str(datetime.now())
    ts = ts.replace(':', '_')
    return ts[5:-7]

if __name__ == '__main__':
    ts = get_filename_timestamp()
    filename = f'ci_cycles_{ts}.xlsx'
    ap = argparse.ArgumentParser()
    ap.add_argument('-s', '-status', default='ALL', help='status filter', required=False)
    ap.add_argument('-b', '-branches', action='store_true', default=False, help='offline mode', required=False)
    ap.add_argument('-o', '-output', default=filename, help='output file name', required=False)
    ap.add_argument('-t', '-test_mode', action='store_true', default=False, help='offline mode', required=False)
    parsed = ap.parse_args()
    main(parsed.s, parsed.o, parsed.b, parsed.t)
