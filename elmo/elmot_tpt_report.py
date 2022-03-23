# pylint: disable=line-too-long,C0103,C0111

import sys
import os.path
import argparse
from datetime import datetime
from operator import itemgetter
import json
import pickle
import time
import pyodbc
import openpyxl
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
import statistics 


def main(offline_mode):
    bug_data = get_elmo_bugs(offline_mode)
    builds_data = get_elmo_builds(get_gfx_driver_version_set(bug_data), offline_mode)
    report_data = get_report_data(bug_data, builds_data)
    save_report(report_data)
    #test_print(report_data)

def test_print(report_data):
    #print(get_key_stat(report_data, 'test_cycle', condition_list=[['submitted_by', 'sys_gkvpgci', 'match']], total=True))

    #print(sort_dict(get_median_by_ww_for_key(report_data, 'tpt_regression_build'),value_sort=True))
    #print(sort_dict(get_median_by_ww_for_key(report_data, 'tpt_regression_build')))

    #print(sort_dict(get_median_by_ww_for_key(report_data, 'tpt_master_build')))

    #print(sort_dict(get_avr_by_ww_for_key(report_data, 'tpt_regression_build'),value_sort=True))
    #print(sort_dict(get_avr_by_ww_for_key(report_data, 'tpt_regression_build')))

    #print(sort_dict(get_avr_by_ww_for_key(report_data, 'tpt_master_build')))

    time_condition_list = [['<=1', 0, 1], ['1-2', 1, 2], ['2-3', 2, 3], ['3-4', 3, 4], ['4-5', 4, 5], ['5-6', 5, 6], ['6-7', 6, 7], ['7-8', 7, 8], ['8-9', 8, 9], ['9-10', 9, 10], ['10-11', 10, 11], ['11-12', 11, 12], ['12-13', 12, 13], ['13-14', 13, 14], ['14-15', 14, 15], ['15-16', 15, 16], ['16-17', 16, 17], ['17-18', 17, 18], ['18-19', 18, 19], ['19-20', 19, 20], ['20-21', 20, 21], ['21-22', 21, 22], ['22-23', 22, 23], ['23-24', 23, 24], ['24-25', 24, 25], ['25-26', 25, 26], ['26-27', 26, 27], ['27-28', 27, 28], ['>28', 28, 1000]]
    print(get_distribution_for_key(report_data, 'tpt_regression_build', time_condition_list))

    #for i in range(0,10):
    #    print(f"['0{i}-0{i+1}', {i}, {i+1}],", end=' ')
    #['00-01', 0, 1], ['01-02', 1, 2], ['02-03', 2, 3], ['03-04', 3, 4], ['04-05', 4, 5], ['05-06', 5, 6], ['06-07', 6, 7], ['07-08', 7, 8], ['08-09', 8, 9], ['09-010', 9, 10]

    #print(sort_dict(get_median_by_ww_for_key(report_data, 'tpt_regression_build', ['env_found', 'silicon', 'match'])))
    #print(sort_dict(get_median_by_ww_for_key(report_data, 'tpt_regression_build', ['env_found', 'simulation', 'match'])))

    #print(sort_dict(get_median_by_ww_for_key(report_data, 'tpt_master_build', ['env_found', 'silicon', 'match'])))
    #print(sort_dict(get_median_by_ww_for_key(report_data, 'tpt_master_build', ['env_found', 'simulation', 'match'])))

    #print(sort_dict(get_avr_by_ww_for_key(report_data, 'tpt_regression_build', ['env_found', 'silicon', 'match'])))
    #print(sort_dict(get_avr_by_ww_for_key(report_data, 'tpt_regression_build', ['env_found', 'simulation', 'match'])))

    #print(sort_dict(get_avr_by_ww_for_key(report_data, 'tpt_master_build', ['env_found', 'silicon', 'match'])))
    #print(sort_dict(get_avr_by_ww_for_key(report_data, 'tpt_master_build', ['env_found', 'simulation', 'match'])))
    
    #print(sort_dict(get_median_by_ww_for_key(report_data, 'tpt_regression_build', ['submitted_by', 'sys_gkvpgci', 'match'])))

def save_report(report_data):
    ts = get_filename_timestamp()
    filename = f'TPT-{ts}-report.xlsx'
    output_wb = openpyxl.Workbook()
    ws = output_wb.active
    ws.title = f'TPT-{ts}'
    print_report_data_table(ws, report_data, 1, 1)
    ws.auto_filter.ref = ws.dimensions
    set_width(ws)
    ws = output_wb.create_sheet('TPT graphs')
    row = 2
    col = 3
    # median
    print_bar_graph(ws, 'CI Daily Bug TPT - Checkin->Development Simulation', get_median_by_ww_for_key(report_data, 'tpt_regression_build', ['env_found', 'simulation', 'match']), 200, col, f'B{row}', 25, 10)
    print_bar_graph(ws, 'CI Daily Bug TPT - Checkin->Development Silicon', get_median_by_ww_for_key(report_data, 'tpt_regression_build', ['env_found', 'silicon', 'match']), 200, col+40, f'R{row}', 25, 10)
    row +=20
    print_bar_graph(ws, 'CI Daily Bug TPT - Checkin->Development Simulation [master]', get_median_by_ww_for_key(report_data, 'tpt_master_build', ['env_found', 'simulation', 'match']), 200, col+10, f'B{row}', 25, 10)
    print_bar_graph(ws, 'CI Daily Bug TPT - Checkin->Development Silicon [master]', get_median_by_ww_for_key(report_data, 'tpt_master_build', ['env_found', 'silicon', 'match']), 200, col+50, f'R{row}', 25, 10)
    row +=20
    # avr
    print_bar_graph(ws, 'CI Daily Bug TPT - Checkin->Development Simulation (AVR)', get_avr_by_ww_for_key(report_data, 'tpt_regression_build', ['env_found', 'simulation', 'match']), 200, col+20, f'B{row}', 25, 10)
    print_bar_graph(ws, 'CI Daily Bug TPT - Checkin->Development Silicon (AVR)', get_avr_by_ww_for_key(report_data, 'tpt_regression_build', ['env_found', 'silicon', 'match']), 200, col+60, f'R{row}', 25, 10)
    row +=20
    print_bar_graph(ws, 'CI Daily Bug TPT - Checkin->Development Simulation [master] (AVR)', get_avr_by_ww_for_key(report_data, 'tpt_master_build', ['env_found', 'simulation', 'match']), 200, col+30, f'B{row}', 25, 10)
    print_bar_graph(ws, 'CI Daily Bug TPT - Checkin->Development Silicon [master] (AVR)', get_avr_by_ww_for_key(report_data, 'tpt_master_build', ['env_found', 'silicon', 'match']), 200, col+70, f'R{row}', 25, 10)
    row +=20

    # TPT distribution
    # time range list
    time_condition_list = [['00-01', 0, 1], ['01-02', 1, 2], ['02-03', 2, 3], ['03-04', 3, 4], ['04-05', 4, 5], ['05-06', 5, 6], ['06-07', 6, 7], ['07-08', 7, 8], ['08-09', 8, 9], ['09-10', 9, 10], ['10-11', 10, 11], ['11-12', 11, 12], ['12-13', 12, 13], ['13-14', 13, 14], ['14-15', 14, 15], ['15-16', 15, 16], ['16-17', 16, 17], ['17-18', 17, 18], ['18-19', 18, 19], ['19-20', 19, 20], ['20-21', 20, 21], ['21-22', 21, 22], ['22-23', 22, 23], ['23-24', 23, 24], ['24-25', 24, 25], ['25-26', 25, 26], ['26-27', 26, 27], ['27-28', 27, 28], ['>28', 28, 1000]]

    print_bar_graph(ws, 'CI Daily Bug TPT distribution - Checkin->Development Simulation', get_distribution_for_key(filer_report_data(report_data, [['env_found', 'simulation', 'match']]), 'tpt_regression_build', time_condition_list), 260, col, f'B{row}', 25, 10)
    print_bar_graph(ws, 'CI Daily Bug TPT distribution - Checkin->Development Silicon', get_distribution_for_key(filer_report_data(report_data, [['env_found', 'silicon', 'match']]), 'tpt_regression_build', time_condition_list), 260, col+20, f'R{row}', 25, 10)
    row +=20
    print_bar_graph(ws, 'CI Daily Bug TPT distribution - Checkin->Development Simulation [master]', get_distribution_for_key(filer_report_data(report_data, [['env_found', 'simulation', 'match']]), 'tpt_master_build', time_condition_list), 260, col+10, f'B{row}', 25, 10)
    print_bar_graph(ws, 'CI Daily Bug TPT distribution - Checkin->Development Silicon [master]', get_distribution_for_key(filer_report_data(report_data, [['env_found', 'silicon', 'match']]), 'tpt_master_build', time_condition_list), 260, col+30, f'R{row}', 25, 10)
    row +=20
    pie_graph_size = 12
    print_pie_graph(ws, 'test_cycle', get_key_stat(report_data, 'test_cycle'), row, col, f'B{row}', pie_graph_size, pie_graph_size, value_sort=True)
    print_pie_graph(ws, 'env_found', get_key_stat(report_data, 'env_found'), row, col+8, f'J{row}', pie_graph_size, pie_graph_size)
    print_pie_graph(ws, 'component', get_key_stat(report_data, 'env_found'), row, col+16, f'R{row}', pie_graph_size, pie_graph_size)
    print_pie_graph(ws, 'platform', get_key_stat(report_data, 'env_found'), row, col+24, f'Z{row}', pie_graph_size, pie_graph_size)

    ws = output_wb.create_sheet('TPT stat')
    row = 2
    col = 2
    # median
    print_stat_table(ws, 'CI Daily Bug TPT - Checkin->Development Simulation', get_median_by_ww_for_key(report_data, 'tpt_regression_build', ['env_found', 'simulation', 'match']), row, col, horizontal=True)
    row +=5
    print_stat_table(ws, 'CI Daily Bug TPT - Checkin->Development Silicon', get_median_by_ww_for_key(report_data, 'tpt_regression_build', ['env_found', 'silicon', 'match']), row, col, horizontal=True)
    row +=5
    print_stat_table(ws, 'CI Daily Bug TPT - Checkin->Development Simulation [master]', get_median_by_ww_for_key(report_data, 'tpt_master_build', ['env_found', 'simulation', 'match']), row, col, horizontal=True)
    row +=5
    print_stat_table(ws, 'CI Daily Bug TPT - Checkin->Development Silicon [master]', get_median_by_ww_for_key(report_data, 'tpt_master_build', ['env_found', 'silicon', 'match']), row, col, horizontal=True)
    row +=5
    # avr
    print_stat_table(ws, 'CI Daily Bug TPT - Checkin->Development Simulation (AVR)', get_avr_by_ww_for_key(report_data, 'tpt_regression_build', ['env_found', 'simulation', 'match']), row, col, horizontal=True)
    row +=5
    print_stat_table(ws, 'CI Daily Bug TPT - Checkin->Development Silicon (AVR)', get_avr_by_ww_for_key(report_data, 'tpt_regression_build', ['env_found', 'silicon', 'match']), row, col, horizontal=True)
    row +=5
    print_stat_table(ws, 'CI Daily Bug TPT - Checkin->Development Simulation [master] (AVR)', get_avr_by_ww_for_key(report_data, 'tpt_master_build', ['env_found', 'simulation', 'match']), row, col, horizontal=True)
    row +=5
    print_stat_table(ws, 'CI Daily Bug TPT - Checkin->Development Silicon [master] (AVR)', get_avr_by_ww_for_key(report_data, 'tpt_master_build', ['env_found', 'silicon', 'match']), row, col, horizontal=True)
    row +=5

    output_wb.save(filename)
    print(f'bugs:{len(report_data)}')
    print(f'Saved the output file:{filename}')

#get_distribution_for_key(report_data, 'tpt_regression_build', [[0,1],[1,2],[2,3],[3,4],[4,5],[5,1000]])

def get_distribution_for_key(report_data, key, range_list):
    #range_list = [['range_1_name', from_int, to_int], ['range_2_name', from_int, to_int]]
    distribution_dict = dict()
    for time_range in range_list: #[[0,1],[1,2],[2,3],[3,4],[4,5],[5,1000]]:
        distribution_dict.update({time_range[0]:len(filer_report_data(report_data, [[key, time_range[1], 'greater'], [key, time_range[2], 'less_eql']]))})
    return distribution_dict


def get_avr_by_ww_for_key(report_data, key, additional_filter=None):
    avr_by_ww = dict()
    if additional_filter:
        report_data = filer_report_data(report_data, [additional_filter])
    for ww in get_key_set(report_data, 'submitted_date_ww'):
        avr_by_ww.update({ww:get_avr_from_report_data(filer_report_data(report_data, [['submitted_date_ww', ww, 'match']]), key)})
    return avr_by_ww

def get_median_by_ww_for_key(report_data, key, additional_filter=None):
    avr_by_ww = dict()
    if additional_filter:
        report_data = filer_report_data(report_data, [additional_filter])
    for ww in get_key_set(report_data, 'submitted_date_ww'):
        avr_by_ww.update({ww:get_median_from_report_data(filer_report_data(report_data, [['submitted_date_ww', ww, 'match']]), key)})
    return avr_by_ww

def get_elmo_bugs(offline_mode=False):
    if offline_mode and os.path.isfile('elmo_bugs_tmp.cache'):
        with open('elmo_bugs_tmp.cache', 'rb') as cache:
            elmo_bugs = pickle.load(cache)
    else:
        conn = pyodbc.connect(Driver='{SQL Server}', Server='GKISQL1601.ger.corp.intel.com,3180', Database='vpgci', Trusted_Connection='Yes')
        cursor = conn.cursor()
        query = "select bug_id, regression, gfx_baseline, component_affected, submitted_by, gfx_driver_version, TRY_CAST(regression_build_label AS XML).value('(/dict/a)[1]', 'varchar(max)') AS regression_build_label_parsed,"
        query += "test_cycle, begin_date, submitted_date, rejected_date, implemented_date, resolution_type, dbo.elmo_hsdes_bug.status, submitter_org, ccn_regression_id, operating_system, platform, platform_affected, problem_classification, begin_date_int, submitted_date_int, component, env_found, gfx_branch,"
        query += "DATEDIFF(HOUR, Convert(datetime, submitted_date), Convert(datetime, rejected_date)) as TPT_DP_resolved_with_weekends,"
        query += "DATEDIFF(HOUR, Convert(datetime, submitted_date), Convert(datetime, implemented_date)) as TPT_DP_implemented_with_weekends,"
        query += "DATEDIFF(HOUR, Convert(datetime, submitted_date), Convert(datetime, GETDATE())) as TPT_DP_open_with_weekends "
        query += "FROM dbo.elmo_hsdes_bug  inner join elmo_gfx_build on TRY_CAST(regression_build_label AS XML).value('(/dict/a)[1]', 'varchar(max)') = elmo_gfx_build.label "
        query += "where submitted_date > '2020-01-01 00:00:00.000'"
        query += "and is_parent=1 and build_type = 'ci'"
        query += "and regression = 'yes'"
        query += "and test_cycle in ('ci_smoke','ci_daily','ci_weekly','ci_dynamic')"
        query += "and gfx_branch = 'gfx-driver__master'"
        query += "order by submitted_date desc"
        cursor.execute(query)
        elmo_bugs = cursor.fetchall()
        with open('elmo_bugs_tmp.cache', 'w+b') as cache:
            pickle.dump(elmo_bugs, cache) 
    return elmo_bugs

def get_elmo_builds(gfx_driver_version_set, offline_mode=False):
    if offline_mode and os.path.isfile('elmo_builds_tmp.cache'):
        with open('elmo_builds_tmp.cache', 'rb') as cache:
            elmo_builds = pickle.load(cache)
    else:
        conn = pyodbc.connect(Driver='{SQL Server}', Server='GKISQL1601.ger.corp.intel.com,3180', Database='vpgci', Trusted_Connection='Yes')
        cursor = conn.cursor()
        query = "select label, qb_id, build_type, status, build_owner, clientspec_branches, begin_date, begin_date_int, duration FROM dbo.elmo_gfx_build where is_parent=1 and build_type = 'ci'"
        query += "and label in ('"
        query += "','".join(gfx_driver_version_set)
        query += "') order by label"
        cursor.execute(query)
        elmo_builds = cursor.fetchall()
        with open('elmo_builds_tmp.cache', 'w+b') as cache:
            pickle.dump(elmo_builds, cache) 
    return elmo_builds

def get_gfx_driver_version_set(query_data):
    gfx_driver_version_set = set()
    for row in query_data:
        gfx_driver_version_set.add(get_first_master_build(row.gfx_driver_version))
    return gfx_driver_version_set

def get_first_master_build(gfx_driver_version):
    builds_list = gfx_driver_version_to_list(gfx_driver_version)
    for build in builds_list:
        if build.find('-master-') < 0:
            builds_list.remove(build)
    if len(builds_list) == 0:
        builds_list = gfx_driver_version_to_list(gfx_driver_version)
    builds_list.sort()
    return builds_list[0]

def gfx_driver_version_to_list(gfx_driver_version):
    gfx_driver_version_list = list()
    gfx_driver_version = gfx_driver_version.replace(';', ',')
    gfx_driver_version = gfx_driver_version.replace('\n', ',')
    gfx_driver_version = gfx_driver_version.replace('\r', ',')
    if gfx_driver_version.find(',') < 0 and gfx_driver_version.find(' ') > 0:
        gfx_driver_version = gfx_driver_version.replace(' ', ',')
    gfx_driver_version = gfx_driver_version.replace(' ', '')
    if gfx_driver_version.find(',') > 0:
        gfx_driver_version_list = gfx_driver_version.split(',')
    else:
        gfx_driver_version_list.append(gfx_driver_version)
    return gfx_driver_version_list

def get_build_begin_date_int(build_label, builds_data):
    master_build_begin_date_int = 0
    for build in builds_data:
        if build.label == build_label:
            master_build_begin_date_int = build.begin_date_int
            break
    return master_build_begin_date_int

def get_first_master_build_begin_date_int(gfx_driver_version, builds_data):
    return get_build_begin_date_int(get_first_master_build(gfx_driver_version), builds_data)

def get_report_data(bug_data, builds_data):
    report_data = dict()    
    for bug in bug_data:
        report_data.update({bug.bug_id:{'id':bug.bug_id, 'regression_build':bug.regression_build_label_parsed, 'master_build':get_first_master_build(bug.gfx_driver_version), 'submitted_date':bug.submitted_date, 'submitted_date_ww':convert_to_iso_ww(bug.submitted_date), 'submitted_by':bug.submitted_by, 'test_cycle':bug.test_cycle, 'env_found':bug.env_found, 'component':bug.component, 'platform':bug.platform, 'platform_affected':bug.platform_affected, 'tpt_regression_build':round((bug.submitted_date_int-bug.begin_date_int)/86400,1), 'tpt_master_build':round((bug.submitted_date_int-get_first_master_build_begin_date_int(bug.gfx_driver_version, builds_data))/86400,1)}})
    return report_data

def convert_to_iso_ww(date_time):
    return f'{str(date_time.isocalendar()[0])[2:]}WW{str(date_time.isocalendar()[1]).zfill(2)}'

def print_report_data_table(ws, report_data, start_row, start_column):
    report_keys = list(report_data.keys())
    data_keys = list(report_data[report_keys[0]].keys())
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    row = start_row
    column = start_column
    for key in data_keys:
        ws.cell(row=row, column=column, value=key)
        ws.cell(row=row, column=column).font = Font(bold = True)
        column += 1
    for data in report_data:
        row += 1
        column = start_column
        for key in data_keys:
            if key == 'id':
                ws.cell(row=row, column=column, value=f'=HYPERLINK("https://hsdes.intel.com/appstore/article/#/{report_data[data][key]}","{report_data[data][key]}")')
                ws.cell(row=row, column=column).font = Font(color='0000FF')
            else:
                ws.cell(row=row, column=column, value=report_data[data][key])
            ws.cell(row=row, column=column).border = thin_border
            column += 1
    return row

def print_stat_table(ws, title, stat_dict, start_row, start_column, value_sort=False, no_sort=False, horizontal=False):
    row = start_row
    column = start_column
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws.cell(row=row, column=column, value=title)
    ws.cell(row=row, column=column).font = Font(bold=True)
    if horizontal:
        table_row = row + 1
        column = start_column
        for stat in sort_dict(stat_dict, value_sort):
            row = table_row
            ws.cell(row=row, column=column, value=stat[0])
            ws.cell(row=row, column=column).border = thin_border
            row += 1
            ws.cell(row=row, column=column, value=stat[1])
            ws.cell(row=row, column=column).border = thin_border
            column += 1
    else:
        for stat in sort_dict(stat_dict, value_sort):
            row += 1
            column = start_column
            ws.cell(row=row, column=column, value=stat[0])
            ws.cell(row=row, column=column).border = thin_border
            column += 1
            ws.cell(row=row, column=column, value=stat[1])
            ws.cell(row=row, column=column).border = thin_border
    return row

def sort_dict(stat_dict, value_sort=False):
    stat_sort_list = []
    for stat in stat_dict:
        stat_sort_list.append([stat, stat_dict[stat]])
    if value_sort:
        stat_sort_list.sort(key=itemgetter(1), reverse=True)
    else:
        stat_sort_list.sort(key=itemgetter(0), reverse=False)
    return stat_sort_list

def print_pie_graph(ws, title, stat_dict, start_row, start_column, graph_poz, graph_w, graph_h, value_sort=False):
    chart = PieChart()
    chart.varyColors = True
    chart.title = title
    chart.height = graph_h
    chart.width = graph_w
    row = print_stat_table(ws, title, stat_dict, start_row, start_column, value_sort)
    chart_data = Reference(ws, min_row=start_row, max_row=row, min_col=start_column+1)
    chart_titles = Reference(ws, min_row=start_row+1, max_row=row, min_col=start_column)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.add_data(chart_data, titles_from_data=True)
    chart.set_categories(chart_titles)
    ws.add_chart(chart, graph_poz)
    return row

def print_bar_graph(ws, title, stat_dict, start_row, start_column, graph_poz, graph_w, graph_h, value_sort=False):
    chart = BarChart()
    chart.style = 11
    chart.shape = 4
    chart.type = "col"
    chart.legend = None
    chart.varyColors = True
    chart.title = title
    chart.height = graph_h
    chart.width = graph_w
    row = print_stat_table(ws, title, stat_dict, start_row, start_column, value_sort)
    chart_data = Reference(ws, min_row=start_row+1, max_row=row, min_col=start_column+1)
    chart_titles = Reference(ws, min_row=start_row+1, max_row=row, min_col=start_column)
    chart.add_data(chart_data, titles_from_data=False)
    chart.set_categories(chart_titles)
    ws.add_chart(chart, graph_poz)
    return row

def get_filename_timestamp():
    ts = str(datetime.now())
    ts = ts.replace(':', '_')
    return ts[:-7]

def get_key_stat(report_data, key, condition_list=None, percent=False, total=False):
    # format condition_list : [[condition_key1, condition_key1_value, condition_key1_func], [condition_key2, condition_key2_value, condition_key2_func]]
    # func: eql, n_eql, greater, greater_eql, less, less_eql, match, contain 
    value_stat_dict = dict()
    value_list = list()
    for bug in report_data:
        if condition_list != None:
            condition_match = True
            for condition in condition_list:
                if check_condition(report_data[bug][condition[0]], condition[1], condition[2]) and condition_match:
                    condition_match = True
                else:
                    condition_match = False
            if condition_match:
                value_list.append(report_data[bug][key])
        else:
            value_list.append(report_data[bug][key])
    value_set = set(value_list)
    for value in value_set:
        if percent:
            value_stat_dict.update({value : round(value_list.count(value)/len(value_list)*100, 2)})
        else:
            value_stat_dict.update({value : value_list.count(value)})
    if total:
        value_stat_dict.update({'total':len(value_list)})
    return value_stat_dict

def get_key_set(report_data, key):
    value_set = set()
    for data in report_data:
        value_set.add(report_data[data][key])
    return value_set

def check_condition(value, condition_value, condition_func):
    check = False
    if condition_func == 'eql':
        if float(value) == float(condition_value):
            check = True
    elif condition_func == 'n_eql':
        if float(value) != float(condition_value):
            check = True
    elif condition_func == 'greater_eql':
        if float(value) >= float(condition_value):
            check = True
    elif condition_func == 'greater':
        if float(value) > float(condition_value):
            check = True
    elif condition_func == 'less':
        if float(value) < float(condition_value):
            check = True
    elif condition_func == 'less_eql':
        if float(value) <= float(condition_value):
            check = True
    elif condition_func == 'match':
        if str(value) == str(condition_value):
            check = True
    elif condition_func == 'contain':
        if str(value).find(str(condition_value)) >= 0:
            check = True
    else:
        check = False
    return check

def get_avr_from_report_data(report_data, key):
    total = 0
    avr = 0
    for data in report_data:
        if isinstance(report_data[data][key], float) or isinstance(report_data[data][key], int):
            total += report_data[data][key]
    if len(report_data) > 0:
        avr = round(total/len(report_data), 2)
    return avr

def get_median_from_report_data(report_data, key):
    data_list = list()
    data_median = 0
    for data in report_data:
        if isinstance(report_data[data][key], float) or isinstance(report_data[data][key], int):
            data_list.append(report_data[data][key])
    if len(report_data) > 0:
        data_median = round(statistics.median(data_list), 2)
    return data_median

def filer_report_data(report_data, condition_list):
    # format condition_list : [[condition_key1, condition_key1_value, condition_key1_func], [condition_key2, condition_key2_value, condition_key2_func]]
    # func: eql, n_eql, greater, greater_eql, less, less_eql, match, contain  
    report_data_filtered = dict()
    for bug in report_data:
        condition_match = True
        for condition in condition_list:
            if check_condition(report_data[bug][condition[0]], condition[1], condition[2]) and condition_match == True:
                condition_match = True
            else:
                condition_match = False
        if condition_match == True:
            report_data_filtered.update({bug:report_data[bug]})
    return report_data_filtered

def set_width(ws, min_width=0):
    column_widths = []
    column_heights = []
    for row_cnt, row in enumerate(ws.iter_rows()):
        column_heights.append(0)
        for col_cnt, cell in enumerate(row):
            if len(column_widths) <= col_cnt:
                column_widths.append(min_width)
            # Ignore cells that don't wrap
            if cell.alignment.wrap_text is False:
                continue
            if is_merged_cell(ws, cell):
                continue
            length, num_lines = calculate_cell_length(cell)
            if cell.alignment.text_rotation == 0:
                column_widths[col_cnt] = max(column_widths[col_cnt], length)
                row_height = 14 * num_lines
                column_heights[row_cnt] = max(column_heights[row_cnt], row_height)
            else:
                column_heights[row_cnt] = max(column_heights[row_cnt], 5 * length)
                col_width = 14 * num_lines
                column_widths[col_cnt] = max(column_widths[col_cnt], col_width)
    max_column_width_allowed = 60
    for cnt, column_width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(cnt)].width = min(column_width, max_column_width_allowed)
    for cnt, row_height in enumerate(column_heights, 1):
        ws.row_dimensions[cnt].height = row_height

def is_merged_cell(ws, cell):
    for cell_range in ws.merged_cells:
        if cell.column >= cell_range.min_col and cell.column <= cell_range.max_col and \
           cell.row    >= cell_range.min_row and cell.row    <= cell_range.max_row:
            return True
    return False

def calculate_cell_length(cell):
    num_lines = 1
    if isinstance(cell.value, str):
        if 'HYPERLINK' in cell.value:
            pieces = cell.value.split(',')
            val = pieces[-1].split('"')[1]
            length = len(val)
        else:
            if '\n' in cell.value:
                num_lines = len(cell.value.split('\n'))
                length = max([len(s) for s in cell.value.split('\n')])
            else:
                length = len(cell.value)
    elif isinstance(cell.value, bool):
        length = len('False')
    elif isinstance(cell.value, int):
        if cell.value == 0:
            length = 1
        else:
            if cell.value < 0:
                val = -1 * cell.value
            else:
                val = cell.value
            length = math.ceil(math.log10(abs(cell.value)))
    else:
        length = 0
    length += 2
    return length, num_lines

if __name__ == '__main__':
    ap = argparse.ArgumentParser()
    ap.add_argument('-o', '-offline', action='store_true', default=False, help='offline mode', required=False)
    parsed = ap.parse_args()
    #print(parsed)
    if parsed.o:
        print('------------ offline mode! ---------------')
    main(parsed.o)
