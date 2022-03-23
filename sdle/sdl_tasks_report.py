# SDL PROJECT ID
# 1145 - DMP
# 842 - PSMS
# 102 - RSD 2.4
# 1389 - PSME -SIM
# 1973 - ICO
# 198 - Workload Collector
# 2399 Performance Framework
import os
import sys
import requests
import urllib3
import json
import openpyxl
from openpyxl.styles import Font 
from openpyxl.styles import PatternFill
from datetime import datetime
# this is to ignore the ssl insecure warning as we are passing in 'verify=false'
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

def main(argv):
    if len(sys.argv) > 1:
        project_id = sys.argv[1]
        sdl_data = get_project_data(project_id)
        #print_tasks(sdl_data)
        print_summary(sdl_data)
        save_report(sdl_data)
        print_potential_na_testing(sdl_data)
        #print_relations(sdl_data)
    else:
        print('no project id')

def get_filename_timestamp():
    ts = str(datetime.now())
    ts = ts.replace(':', '_')
    return ts[:-7]

def color_row_for_status(ws,row,cell_start,cell_end,status):
    dict_status_to_color = {'Complete':'008000','Incomplete':'FF0000','In Progress':'FFFF00','Not Applicable':'A0A0A0'}
    for col in range(cell_start,cell_end+1):
                    ws.cell(row=row,column=col).fill = PatternFill(start_color=dict_status_to_color[status], fill_type = "solid")

def add_task_row(ws,row,task_data,col_start,col_end):
    for i in range (col_start+1, col_end+1):
        ws.cell(row=row, column=i, value=task_data[i-1])

def add_task_column_names(ws,row,column_names,col_start,col_end):
    col=0
    for i in range(col_start,col_end):
        col+=1
        ws.cell(row=row, column=col, value=column_names[i])
        ws.cell(row=row, column=col).font = Font(bold = True)

def format_columns(ws):
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 100
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 200

def format_columns_relations(ws):
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 21
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 100
    ws.column_dimensions['E'].width = 11
    ws.column_dimensions['F'].width = 21
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 100
    ws.column_dimensions['I'].width = 20


def get_project_name(project_id):
    headers = {'Authorization': 'Token 981050d9b8eba8ef73509451e4b42563b63d17bf'}
    url = 'https://sdp-prod.intel.com/api/v2/projects/'+project_id
    response = requests.get(url, verify=False, headers=headers)
    json.project = response.json()
    project_name = json.project['name']
    return project_name

def get_project_data(project_id):
    headers = {'Authorization': 'Token 981050d9b8eba8ef73509451e4b42563b63d17bf'}
    #https://docs.sdelements.com/release/4.19/api/docs/tasks/
    url = 'https://sdp-prod.intel.com/api/v2/projects/'+project_id+'/tasks/?include=tags,related&expand=status,tags'
    response = requests.get(url, verify=False, headers=headers)
    json.all_tasks = response.json()
    phase_dict = {"CX1":"Project Management","X1":"Requirements","X2":"Architecture & Design","X3":"Development","X4":"Testing","X5":"Deployment"}
    phase_dict_short = {"CX1":"PM","X1":"Req","X2":"A&D","X3":"Dev","X4":"T","X5":"Dep"}
    sdl_tasks = []
    sdl_tasks_relation = []
    sdl_statuses = []
    for task in json.all_tasks['results']:
        related_tasks = []
        related_tasks_data = []
        if 'related_tasks' in task.keys():
            for i in range(len(task['related_tasks'])):
                related_tasks.append('['+task['related_tasks'][i]['id']+'-'+phase_dict_short.get(get_related_task_phase(task['related_tasks'][i]['id'],json.all_tasks['results']))+'-'+get_related_task_status(task['related_tasks'][i]['id'],json.all_tasks['results'])+']')
                related_tasks_data.append([task['related_tasks'][i]['id'],phase_dict.get(get_related_task_phase(task['related_tasks'][i]['id'],json.all_tasks['results'])),get_related_task_status(task['related_tasks'][i]['id'],json.all_tasks['results']),task['related_tasks'][i]['title'],task['related_tasks'][i]['url']])
        related_tasks.sort()
        related_tasks_data.sort()
        sdl_tasks.append([task['task_id'],task['title'],task['status']['name'],phase_dict.get(task['phase']),task['priority'],' '.join(related_tasks),task['tags']['library_tags']])
        sdl_statuses.append(task['status']['name'])
        sdl_tasks_relation.append([task['task_id'],task['title'],task['status']['name'],phase_dict.get(task['phase']),task['priority'],related_tasks_data])
    sdl_tasks.sort()
    sdl_tasks_relation.sort()
    project_name = get_project_name(project_id).replace(' ', '_')
    column_names = ['task_id','title','status','phase','priority','related_tasks']
    sdl_data = {'project_name':project_name.lower(), 'sdl_tasks':sdl_tasks, 'sdl_statuses':sdl_statuses, 'sdl_column_names':column_names, 'sdl_tasks_relation':sdl_tasks_relation}
    return sdl_data

def get_related_task_status(related_task_id,all_sdl_tasks_list):
    related_task_status = '???'
    for task in all_sdl_tasks_list:
        if task['task_id'] == related_task_id:
            related_task_status = task['status']['name']
            break
    return related_task_status

def get_related_task_phase(related_task_id,all_sdl_tasks_list):
    related_task_phase = '???'
    for task in all_sdl_tasks_list:
        if task['task_id'] == related_task_id:
            related_task_phase = task['phase']
            break
    return related_task_phase

def save_report(sdl_data):
    output_wb = openpyxl.Workbook()
    ts = get_filename_timestamp()
    name = sdl_data['project_name']
    sdl_data_tags_position = 6
    filename = f'{name}-{ts}-SDL_report.xlsx'
    ws = output_wb.active
    ws.title = 'mandatory'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws.cell(row=1, column=1, value='Pre-prod Complete')
    ws.cell(row=1, column=1).font = Font(bold = True)
    row=2
    add_task_column_names(ws,row,sdl_data['sdl_column_names'],0,5)
    for task in sdl_data['sdl_tasks']:
        if 'PreprodComplete' in task[sdl_data_tags_position]:
            row+=1
            add_task_row(ws,row,task,0,5)
            color_row_for_status(ws,row,1,5,task[2])

    row+=1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value='Pre-prod In progress')
    ws.cell(row=row, column=1).font = Font(bold = True)
    row+=1
    add_task_column_names(ws,row,sdl_data['sdl_column_names'],0,5)
    for task in sdl_data['sdl_tasks']:
        if 'PreprodInprogress' in task[sdl_data_tags_position]:
            row+=1
            add_task_row(ws,row,task,0,5)
            color_row_for_status(ws,row,1,5,task[2])
    row+=1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value='Other need to be completed')
    ws.cell(row=row, column=1).font = Font(bold = True)
    row+=1
    add_task_column_names(ws,row,sdl_data['sdl_column_names'],0,5)
    for task in sdl_data['sdl_tasks']:
        if 'SWS4' in task[sdl_data_tags_position] and 'PreprodInprogress' not in task[sdl_data_tags_position] and 'PreprodComplete' not in task[sdl_data_tags_position]:
            row+=1
            add_task_row(ws,row,task,0,5)
            color_row_for_status(ws,row,1,5,task[2])
    format_columns(ws)

    ws = output_wb.create_sheet('all')
    row=1
    add_task_column_names(ws,row,sdl_data['sdl_column_names'],0,6)
    for task in sdl_data['sdl_tasks']:
            row+=1
            add_task_row(ws,row,task,0,6)
            color_row_for_status(ws,row,1,sdl_data_tags_position,task[2])
    ws.auto_filter.ref = ws.dimensions
    format_columns(ws)

    ws = output_wb.create_sheet('relations')
    add_task_column_names(ws,1,['task id','task phase','task status', 'task title', 'related id','related phase','related status','related title', 'related url'],0,9)
    row=2
    for task in sdl_data['sdl_tasks_relation']:
        if len(task[5])>0:
            for related in task[5]:
                ws.cell(row=row, column=1, value=task[0])
                ws.cell(row=row, column=2, value=task[3])
                ws.cell(row=row, column=3, value=task[2])
                ws.cell(row=row, column=4, value=task[1])
                ws.cell(row=row, column=5, value=related[0])
                ws.cell(row=row, column=6, value=related[1])
                ws.cell(row=row, column=7, value=related[2])
                ws.cell(row=row, column=8, value=related[3])
                ws.cell(row=row, column=9, value= f'=HYPERLINK("{related[4]}","LINK")')
                row+=1
    ws.auto_filter.ref = ws.dimensions
    format_columns_relations(ws)
        
    ws = output_wb.create_sheet('summary')
    ws.cell(row=1, column=1, value='Incomplete:')
    ws.cell(row=1, column=2, value=sdl_data['sdl_statuses'].count('Incomplete'))
    ws.cell(row=1, column=3, value=round(sdl_data['sdl_statuses'].count('Incomplete')/len(sdl_data['sdl_statuses']),4))
    ws.cell(row=1, column=3).style = 'Percent'
    ws.cell(row=2, column=1, value='In Progress:')
    ws.cell(row=2, column=2, value=sdl_data['sdl_statuses'].count('In Progress'))
    ws.cell(row=2, column=3, value=round(sdl_data['sdl_statuses'].count('In Progress')/len(sdl_data['sdl_statuses']),4))  
    ws.cell(row=2, column=3).style = 'Percent' 
    ws.cell(row=3, column=1, value='Not Applicable:')
    ws.cell(row=3, column=2, value=sdl_data['sdl_statuses'].count('Not Applicable'))
    ws.cell(row=3, column=3, value=round(sdl_data['sdl_statuses'].count('Not Applicable')/len(sdl_data['sdl_statuses']),4))   
    ws.cell(row=3, column=3).style = 'Percent'
    ws.cell(row=4, column=1, value='Complete:')
    ws.cell(row=4, column=2, value=sdl_data['sdl_statuses'].count('Complete'))
    ws.cell(row=4, column=3, value=round(sdl_data['sdl_statuses'].count('Complete')/len(sdl_data['sdl_statuses']),4))
    ws.cell(row=4, column=3).style = 'Percent'
    ws.cell(row=5, column=1, value='Total:')
    ws.cell(row=5, column=2, value=len(sdl_data['sdl_statuses']))
    ws.cell(row=6, column=1, value='Total Applicable:')
    ws.cell(row=6, column=2, value=len(sdl_data['sdl_statuses'])-sdl_data['sdl_statuses'].count('Not Applicable'))
    ws.cell(row=7, column=1, value='Complete + N/A:')
    ws.cell(row=7, column=2, value=sdl_data['sdl_statuses'].count('Complete')+sdl_data['sdl_statuses'].count('Not Applicable'))
    ws.cell(row=7, column=3, value=round((sdl_data['sdl_statuses'].count('Complete')+sdl_data['sdl_statuses'].count('Not Applicable'))/len(sdl_data['sdl_statuses']),4))
    ws.cell(row=7, column=3).style = 'Percent'
    ws.cell(row=8, column=1, value='Target:')
    ws.cell(row=8, column=2, value=round(len(sdl_data['sdl_statuses'])*.75,0))
    ws.cell(row=8, column=3, value=0.75)
    ws.cell(row=8, column=3).style = 'Percent'
    ws.column_dimensions['A'].width = 17

    output_wb.save(filename)
    print(f'COMPLETED: Saved the output file:{filename}')


def print_tasks(sdl_data):
    print('--------------------------------------------------------------------------------------------------------------------')
    print('--   PreprodComplete   ---------------------------------------------------------------------------------------------')
    print('--------------------------------------------------------------------------------------------------------------------')
    for task in sdl_data['sdl_tasks']:
        if 'PreprodComplete' in task[5]:
            print(f"{task[0]};{task[1]};{task[2]};{task[3]};{task[4]};{task[5]}")

    print('--------------------------------------------------------------------------------------------------------------------')
    print('--   PreprodInprogress   -------------------------------------------------------------------------------------------')
    print('--------------------------------------------------------------------------------------------------------------------')

    for task in sdl_data['sdl_tasks']:
        if 'PreprodInprogress' in task[5]:
            print(f"{task[0]};{task[1]};{task[2]};{task[3]};{task[4]};{task[5]}")

    print('--------------------------------------------------------------------------------------------------------------------')
    print('--   Prod Other   --------------------------------------------------------------------------------------------------')
    print('--------------------------------------------------------------------------------------------------------------------')

    for task in sdl_data['sdl_tasks']:
        if 'SWS4' in task[5] and 'PreprodInprogress' not in task[5] and 'PreprodComplete' not in task[5] :
            print(f"{task[0]};{task[1]};{task[2]};{task[3]};{task[4]};{task[5]}")

    print('--------------------------------------------------------------------------------------------------------------------')

    for task in sdl_data['sdl_tasks']:
        if 'PreprodInprogress' not in task[5] and 'PreprodComplete' not in task[5] and 'SWS4' not in  task[5]:
            print(f"{task[0]};{task[1]};{task[2]};{task[3]};{task[4]};{task[5]}")
    
    print('--------------------------------------------------------------------------------------------------------------------')  

def print_summary(sdl_data):
    print('--------------------------------------------------------------------------------------------------------------------') 
    print(sdl_data['project_name']) 
    print('--------------------------------------------------------------------------------------------------------------------') 
    print(f"      Incomplete:{sdl_data['sdl_statuses'].count('Incomplete')} - {round(sdl_data['sdl_statuses'].count('Incomplete')/len(sdl_data['sdl_statuses'])*100,2)}%") 
    print(f"     In Progress:{sdl_data['sdl_statuses'].count('In Progress')} - {round(sdl_data['sdl_statuses'].count('In Progress')/len(sdl_data['sdl_statuses'])*100,2)}%") 
    print(f"  Not Applicable:{sdl_data['sdl_statuses'].count('Not Applicable')} - {round(sdl_data['sdl_statuses'].count('Not Applicable')/len(sdl_data['sdl_statuses'])*100,2)}%") 
    print(f"        Complete:{sdl_data['sdl_statuses'].count('Complete')} - {round(sdl_data['sdl_statuses'].count('Complete')/len(sdl_data['sdl_statuses'])*100,2)}%") 
    print(f"           Total:{len(sdl_data['sdl_statuses'])}") 
    print(f"Total Applicable:{len(sdl_data['sdl_statuses'])-sdl_data['sdl_statuses'].count('Not Applicable')}") 
    print(f"  Complete + N/A:{sdl_data['sdl_statuses'].count('Complete')+sdl_data['sdl_statuses'].count('Not Applicable')} - {round((sdl_data['sdl_statuses'].count('Complete')+sdl_data['sdl_statuses'].count('Not Applicable'))/len(sdl_data['sdl_statuses'])*100,2)}%")
    print(f"          Target:{round(len(sdl_data['sdl_statuses'])*.75,0)} - 75%")
    print('--------------------------------------------------------------------------------------------------------------------') 

def print_potential_na_testing(sdl_data):
    print('--------------------------------------------------------------------------------------------------------------------') 
    print('Potential N/A testing tasks:') 
    for task in sdl_data['sdl_tasks_relation']:
        if task[2] == 'Not Applicable' and len(task[5])>0:
            for related in task[5]:
                if related[1] == 'Testing' and related[2] != 'Not Applicable':
                    print(f'{task[0]}-{task[3]} -> {related[0]}-{related[1]} ({related[2]}) {related[3]} [{related[4]}]')
    print('--------------------------------------------------------------------------------------------------------------------') 

def print_relations(sdl_data):
    print('--------------------------------------------------------------------------------------------------------------------') 
    print('Tasks relations:') 
    for task in sdl_data['sdl_tasks_relation']:
        if len(task[5])>0:
            for related in task[5]:
                print(f'{task[0]}-{task[3]} ({task[2]})-> {related[0]}-{related[1]} ({related[2]}) {related[3]}')
    print('--------------------------------------------------------------------------------------------------------------------') 


if __name__ == '__main__':
    main(sys.argv[1:])