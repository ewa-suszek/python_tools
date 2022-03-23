import sys
import os.path
import argparse
import requests
import urllib.parse
from datetime import datetime
from operator import itemgetter
import re
import json
import pickle
import time
import openpyxl
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
import statistics 
import math

#http://gtax-gcmxd-fm.intel.com/api/v1/

def main(gtax_instance, job_set_session_id, client_search, offline_mode, task_info):
    session_jobs_dict = dict()
    session_tasks_dict = dict()
    session_jobs_dict = get_jobs_dict_for_session_id(gtax_instance, job_set_session_id, client_search, offline_mode=offline_mode)
    print(f'{len(session_jobs_dict)} jobs and {get_task_count_for_jobs_dict(session_jobs_dict)} tasks')
    if task_info:
        if client_search:
            session_tasks_dict = get_tasks_dict_for_session_id_and_client_search(gtax_instance, job_set_session_id, client_search, offline_mode=offline_mode)
        else:
            #session_tasks_dict = get_tasks_dict_for_session_jobs_dict(gtax_instance, job_set_session_id, session_jobs_dict, offline_mode=offline_mode)
            session_tasks_dict = get_tasks_dict_for_session_id(gtax_instance, job_set_session_id, offline_mode=offline_mode)
    print_report(gtax_instance, job_set_session_id, session_jobs_dict, session_tasks_dict, task_info)

def get_task_count_for_jobs_dict(jobs_dict):
    tasks_count = 0
    for job in jobs_dict:
        tasks_count += int(jobs_dict[job]['total'])
    return tasks_count

def print_report(gtax_instance, job_set_session_id, session_jobs_dict, session_tasks_dict, task_info):
    # -----------------------------------------------------------
    # JOBS TAB
    write_tasks_to_large_xlsx_file(f'{job_set_session_id}-jobs.xlsx', f'jobs-{job_set_session_id}', session_jobs_dict, gtax_instance, tasks=False)
    #ws.title = f'jobs-{job_set_session_id}'
    #print_data_table(ws, gtax_instance, session_jobs_dict, 1, 1)
    #ws.auto_filter.ref = ws.dimensions
    #set_width(ws, max_width=600)
    

    # -----------------------------------------------------------
    # CSQ TAB
    filename = f'{job_set_session_id}-stat.xlsx'
    output_wb = openpyxl.Workbook()
    ws = output_wb.active
    ws.title = f'csq-{job_set_session_id}'
    csq_dict = get_csq_dict(session_jobs_dict, gtax_instance)
    print_data_table(ws, gtax_instance, csq_dict, 1, 1)
    ws.auto_filter.ref = ws.dimensions
    set_width(ws, max_width=600)

    # -----------------------------------------------------------
    # TASKS TAB

    if task_info:
        write_tasks_to_large_xlsx_file(f'{job_set_session_id}-tasks.xlsx', f'tasks-{job_set_session_id}', session_tasks_dict, gtax_instance, tasks=True)
        #ws = output_wb.create_sheet(f'tasks-{job_set_session_id}')
        #print_data_table(ws, gtax_instance, session_tasks_dict, 1, 1, tasks=True)
        #ws.auto_filter.ref = ws.dimensions
        #set_width(ws, max_width=600)


    # -----------------------------------------------------------
    # GRAPHS TAB JOBS
    ws = output_wb.create_sheet('jobs stat')
    row = 2
    col = 3   
    pie_graph_size = 12
    print_pie_graph(ws, 'status', get_key_stat(session_tasks_dict,'status'), row, col, f'B{row}', pie_graph_size, pie_graph_size, value_sort=True)
    print_pie_graph(ws, 'tasks results', get_tasks_stat(session_jobs_dict), row, col+8, f'J{row}', pie_graph_size, pie_graph_size, value_sort=True)
    print_pie_graph(ws, 'submission_type', get_key_stat(session_jobs_dict, 'submission_type'), row, col+16, f'R{row}', pie_graph_size, pie_graph_size, value_sort=True)
    row += 24
    filtered_jobs  = filer_report_data(session_jobs_dict, [['result', 'aborted in setup', 'match']])
    jobs_stat = get_key_stat(filtered_jobs,'client_name')
    bar_graph_size = len(jobs_stat)
    print_bar_graph(ws, 'aborted in setup', jobs_stat, row, col, f'B{row}', 25, bar_graph_size, value_sort=True)

    row += 24
    filtered_jobs  = filer_report_data(session_jobs_dict, [['submission_type', 'init', 'match']])
    jobs_stat = get_key_stat(filtered_jobs, 'csq')
    bar_graph_size = len(jobs_stat)
    print_bar_graph(ws, 'csq', jobs_stat, row, col, f'B{row}', 25, bar_graph_size, value_sort=True)

    # -----------------------------------------------------------
    # GRAPHS TAB TASKS
    if task_info:
        ws = output_wb.create_sheet('tasks stat')
        stat_limit = 3
        row = 2
        col = 3   
        pie_graph_size = 12
        bar_graph_size = 12
        print_pie_graph(ws, 'gta_status', get_key_stat(session_tasks_dict, 'gta_status'), row, col, f'B{row}', pie_graph_size, pie_graph_size, value_sort=True)
        print_pie_graph(ws, 'status', get_key_stat(session_tasks_dict, 'status'), row, col+8, f'J{row}', pie_graph_size, pie_graph_size, value_sort=True)
        filtered_tasks = filer_report_data(session_tasks_dict, [['gta_status', 'Error', 'match']])
        tasks_stat = get_key_stat(filtered_tasks,'primary_reason')
        print_pie_graph(ws, 'primary_reason for error tasks', tasks_stat, row, col+16, f'R{row}', pie_graph_size, pie_graph_size, value_sort=True)
        row += (pie_graph_size * 2)
        print_pie_graph(ws, 'gta_status', get_key_stat(session_tasks_dict, 'gta_status'), row, col, f'B{row}', pie_graph_size, pie_graph_size, value_sort=True)
        print_pie_graph(ws, 'status', get_key_stat(session_tasks_dict, 'status'), row, col+8, f'J{row}', pie_graph_size, pie_graph_size, value_sort=True)
        print_pie_graph(ws, 'submission_type', get_key_stat(session_tasks_dict, 'submission_type'), row, col+16, f'R{row}', pie_graph_size, pie_graph_size, value_sort=True)
        row += (pie_graph_size * 2)

        filtered_tasks = filer_report_data(session_tasks_dict, [['submission_type', 'init', 'match'], ['phase', 'test', 'match'], ['status', 'passed', 'not_match']])
        tasks_stat = get_key_stat(filtered_tasks,'command', limit=stat_limit)
        print_bar_graph(ws, f'init test not passed by command', tasks_stat, row, col, f'B{row}', 25, bar_graph_size, value_sort=True)

        filtered_tasks = filer_report_data(session_tasks_dict, [['submission_type', 'rerun', 'match'], ['phase', 'test', 'match'], ['status', 'passed', 'not_match']])
        tasks_stat = get_key_stat(filtered_tasks,'command', limit=stat_limit)
        row = print_bar_graph(ws, f'reruns not passed by command', tasks_stat, row, col+16, f'Q{row}', 25, bar_graph_size, value_sort=True)
        
        stat_limit = 2
        row += (bar_graph_size * 2)
        filtered_tasks = filer_report_data(session_tasks_dict, [['status', 'passed', 'not_match'],['phase', 'test', 'match']])
        tasks_stat = get_key_stat(filtered_tasks,'client_name', limit=stat_limit)
        print_bar_graph(ws, f'not passed by client_name', tasks_stat, row, col, f'B{row}', 25, bar_graph_size, value_sort=True)

        filtered_tasks = filer_report_data(session_tasks_dict, [['status', 'passed', 'match'],['phase', 'test', 'match']])
        tasks_stat = get_key_stat(filtered_tasks,'client_name', limit=stat_limit)
        print_bar_graph(ws, f'passed by client_name', tasks_stat, row, col+16, f'Q{row}', 25, bar_graph_size, value_sort=True)

        row += (bar_graph_size * 2)
        filtered_tasks = filer_report_data(session_tasks_dict, [['status', 'timed out', 'match'],['phase', 'test', 'match']])
        tasks_stat = get_key_stat(filtered_tasks,'client_name', limit=stat_limit)
        print_bar_graph(ws, f'timed out by client_name', tasks_stat, row, col, f'B{row}', 25, bar_graph_size, value_sort=True)

        filtered_tasks = filer_report_data(session_tasks_dict, [['status', 'failed', 'match'],['phase', 'test', 'match']])
        tasks_stat = get_key_stat(filtered_tasks,'client_name', limit=stat_limit)
        print_bar_graph(ws, f'fail by client_name', tasks_stat, row, col+16, f'Q{row}', 25, bar_graph_size, value_sort=True)

        row += (bar_graph_size * 2)
        filtered_tasks = filer_report_data(session_tasks_dict, [['status', 'aborted', 'match'],['phase', 'test', 'match']])
        tasks_stat = get_key_stat(filtered_tasks,'client_name', limit=stat_limit)
        print_bar_graph(ws, f'aborted in test by client_name', tasks_stat, row, col, f'B{row}', 25, bar_graph_size, value_sort=True)

        filtered_tasks = filer_report_data(session_tasks_dict, [['status', 'aborted', 'match'],['phase', 'setup', 'match']])
        tasks_stat = get_key_stat(filtered_tasks,'client_name', limit=stat_limit)
        print_bar_graph(ws, f'aborted in setup by client_name', tasks_stat, row, col+16, f'Q{row}', 25, bar_graph_size, value_sort=True)

        row += (bar_graph_size * 2)
        filtered_tasks = filer_report_data(session_tasks_dict, [['phase', 'test', 'match']])
        tasks_stat = get_key_stat(filtered_tasks,'client_name', limit=stat_limit)
        print_bar_graph(ws, f'all tests by client_name', tasks_stat, row, col, f'B{row}', 25, bar_graph_size, value_sort=True)

        filtered_tasks = filer_report_data(session_tasks_dict, [['submission_type', 'init', 'match'],['phase', 'setup', 'match']])
        tasks_stat = get_key_stat(filtered_tasks,'client_name', limit=stat_limit)
        print_bar_graph(ws, f'all init tests by client_name', tasks_stat, row, col+16, f'Q{row}', 25, bar_graph_size, value_sort=True)
    
    # -----------------------------------------------------------

    output_wb.save(filename)
    print(f'Saved the output file:{filename}')

def print_data_table(ws, gtax_instance, report_data, start_row, start_column, tasks=False):
    report_keys = list(report_data.keys())
    if len(report_data) > 0:
        data_keys = list(report_data[report_keys[0]].keys())
    else:
        data_keys = list()
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    row = start_row
    column = start_column
    for key in data_keys:
        ws.cell(row=row, column=column, value=key)
        ws.cell(row=row, column=column).font = Font(bold = True)
        column += 1
    for data in report_data:
        row += 1
        column = start_column
        for key in data_keys:
            if key == 'id':
                if tasks:
                    url = f"http://{gtax_instance}.intel.com/#/jobs/{report_data[data]['job_id']}/task_results/{report_data[data]['id']}"
                    ws.cell(row=row, column=column, value=f'=HYPERLINK("{url}","{report_data[data][key]}")')
                else:
                    ws.cell(row=row, column=column, value=f'=HYPERLINK("http://{gtax_instance}.intel.com/#/jobs/{report_data[data][key]}","{report_data[data][key]}")')
                ws.cell(row=row, column=column).font = Font(color='0000FF')
            elif key == 'csq_link':
                #ws.cell(row=row, column=column, value=f'=HYPERLINK("{report_data[data][key]}","gtax_link")')
                ws.cell(row=row, column=column, value='gtax_link')
                ws.cell(row=row, column=column).style = 'Hyperlink'
                ws.cell(row=row, column=column).hyperlink = report_data[data][key]
                ws.cell(row=row, column=column).font = Font(color='0000FF')
            else:
                ws.cell(row=row, column=column, value=report_data[data][key])
            ws.cell(row=row, column=column).border = thin_border
            column += 1
    return row

def write_tasks_to_large_xlsx_file(file_name, sheet_name, data_dict, gtax_instance, tasks=False):
    letter_index = get_letter_index_list()
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    wb = openpyxl.Workbook(write_only = True)
    ws = wb.create_sheet(sheet_name)
    if len(data_dict) > 0:
        report_keys = list(data_dict.keys())
        data_keys = list(data_dict[report_keys[0]].keys())
    else:
        data_keys = list()
    row_count = 1
    row_cells = list()
    for key in data_keys:
        cell = WriteOnlyCell(ws, value=key)
        cell.font = Font(bold = True)
        cell.border = thin_border
        row_cells.append(cell)
    ws.append(row_cells)
    for data in data_dict:
        row_count += 1
        row_cells = list()
        for key in data_keys:      
            if key == 'id':
                cell = WriteOnlyCell(ws, value=data_dict[data][key])
                cell.style = 'Hyperlink'
                cell.font = Font(color='0000FF')
                if tasks:
                    cell.hyperlink = f"http://{gtax_instance}.intel.com/#/jobs/{data_dict[data]['job_id']}/task_results/{data_dict[data]['id']}"
                else:
                    cell.hyperlink = f"http://{gtax_instance}.intel.com/#/jobs/{data_dict[data][key]}"
            elif key == 'csq_link':
                #ws.cell(row=row, column=column, value=f'=HYPERLINK("{report_data[data][key]}","gtax_link")')
                cell = WriteOnlyCell(ws, value='gtax_link')
                cell.style = 'Hyperlink'
                cell.font = Font(color='0000FF')
                cell.hyperlink = data_dict[data][key]
            else:
                if data_dict[data][key]:
                    cell = WriteOnlyCell(ws, value=data_dict[data][key])
                else:
                    cell = WriteOnlyCell(ws, value='')
            cell.border = thin_border
            row_cells.append(cell)
        ws.append(row_cells)
    ws.auto_filter.ref = f'A1:{letter_index[len(data_keys)]}{row_count}'
    wb.save(file_name)
    print(f'Saved the output file:{file_name}')

def get_gtax_data(url):
    headers = { 'Content-type': 'application/json' }
    response = requests.get(url, headers=headers, proxies={'http': 'http://proxy-chain.intel.com:911', 'https': 'http://proxy-chain.intel.com:912'})
    return response.json()

def get_jobs_dict_for_session_id(gtax_instance, job_set_session_id, client_search, offline_mode=False):
    session_jobs_dict = dict()
    cache_file_name = f'gtax_{job_set_session_id}_tmp.cache'
    if offline_mode and os.path.isfile(cache_file_name):
        with open(cache_file_name, 'rb') as cache:
            session_jobs_dict = pickle.load(cache)
    else:
        print(f'getting data for job set session: {job_set_session_id}')
        if client_search:
            url = f'http://{gtax_instance}.intel.com/api/v1/jobs?include_tasks=false&include_taskml=false&include_csq=true&full_info=false&client_name={client_search}&jobset_session_ids={job_set_session_id}'
        else:
            url = f'http://{gtax_instance}.intel.com/api/v1/jobs?include_tasks=false&include_taskml=false&include_csq=true&full_info=false&jobset_session_ids={job_set_session_id}'
        print(url)
        data = get_gtax_data(url)
        session_jobs_dict = get_session_jobs_dict(data['data'])
        with open(cache_file_name, 'w+b') as cache:
            pickle.dump(session_jobs_dict, cache) 
    return session_jobs_dict

def get_session_jobs_dict(session_jobs_data):
    session_jobs_dict = dict()
    for job_data in session_jobs_data:
        job_dict = dict()
        for key in ['id', 'status', 'name', 'result', 'submission_type', 'submitted_date', 'started_date', 'completed_date', 'duration', 'scheduler_account_name', 'dumps_produced']:
            job_dict.update({key:job_data[key]})
        for key in ['client']:
            for sub_key in ['name','id']:
                key_name = f'{key}_{sub_key}'
                job_dict.update({key_name:job_data[key][sub_key]})
                if key_name == 'client_name':
                    job_dict.update({'client_group':get_client_group(job_data['client']['name'])})
                if 'csq' in job_data.keys():
                    job_dict.update({'platform':get_client_platform(job_data['csq']['query'])})
                    job_dict.update({'pool':get_client_pool(job_data['csq']['query'])})
        for key in ['results_summary']:
            for sub_key in job_data[key].keys():
                key_name = f'{sub_key}'
                job_dict.update({key_name:job_data[key][sub_key]})
        if 'csq' in job_data.keys():
            #job_dict.update({'platform':get_client_platform(job_data['csq']['query'])})
            #job_dict.update({'pool':get_client_pool(job_data['csq']['query'])})
            job_dict.update({'csq':job_data['csq']['query']})
        session_jobs_dict.update({job_data['id']:job_dict})
    return session_jobs_dict

def get_tasks_dict_for_session_jobs_dict(gtax_instance, job_set_session_id, session_jobs_dict, offline_mode=False):
    session_task_dict = dict()
    cache_file_name = f'gtax_{job_set_session_id}_full_tmp.cache'
    if offline_mode and os.path.isfile(cache_file_name):
        with open(cache_file_name, 'rb') as cache:
            session_task_dict = pickle.load(cache)
    else:
        print(f'be patient!! - getting full info of tasks for job set session: {job_set_session_id}')
        for job in session_jobs_dict:
            print(f'getting tasks for job id:{job}')
            url = f'http://{gtax_instance}.intel.com/api/v1/jobs?include_tasks=true&include_csq=true&include_phases_task_counts=true&full_info=true&id={job}'
            data = get_gtax_data(url)
            session_task_dict.update(get_session_tasks_dict(data['data']))
        with open(cache_file_name, 'w+b') as cache:
            pickle.dump(session_task_dict, cache) 
    return session_task_dict

def get_tasks_dict_for_session_id_and_client_search(gtax_instance, job_set_session_id, client_search, offline_mode=False):
    session_task_dict = dict()
    cache_file_name = f'gtax_{job_set_session_id}_full_tmp.cache'
    if offline_mode and os.path.isfile(cache_file_name):
        with open(cache_file_name, 'rb') as cache:
            session_task_dict = pickle.load(cache)
    else:
        print(f'be patient!! - getting full info of tasks for job set session: {job_set_session_id}')
        url = f'http://{gtax_instance}.intel.com/api/v1/jobs?include_tasks=true&include_csq=true&include_phases_task_counts=true&full_info=true&client_name={client_search}&jobset_session_ids={job_set_session_id}'
        data = get_gtax_data(url)
        session_task_dict.update(get_session_tasks_dict(data['data']))
        with open(cache_file_name, 'w+b') as cache:
            pickle.dump(session_task_dict, cache) 
    return session_task_dict

def get_tasks_dict_for_session_id(gtax_instance, job_set_session_id, offline_mode=False):
    session_task_dict = dict()
    cache_file_name = f'gtax_{job_set_session_id}_full_tmp.cache'
    if offline_mode and os.path.isfile(cache_file_name):
        with open(cache_file_name, 'rb') as cache:
            session_task_dict = pickle.load(cache)
    else:
        print(f'be patient!! - getting full info of tasks for job set session: {job_set_session_id}')
        url = f'http://{gtax_instance}.intel.com/api/v1/jobs?include_tasks=true&include_csq=true&include_phases_task_counts=true&full_info=true&jobset_session_ids={job_set_session_id}'
        data = get_gtax_data(url)
        session_task_dict.update(get_session_tasks_dict(data['data']))
        with open(cache_file_name, 'w+b') as cache:
            pickle.dump(session_task_dict, cache) 
    return session_task_dict

def get_session_tasks_dict(session_jobs_data):
    session_task_dict = dict()
    for job_data in session_jobs_data:
        for task_data in job_data['tasks']:
            task_dict = dict()
            for key in ['id', 'job_id', 'task_result_id', 'task_definition_id']:
                task_dict.update({key:task_data[key]})
            if job_data['client']:
                task_dict.update({'client_name':job_data['client']['name']})
                task_dict.update({'client_id':job_data['client']['id']})
            else:
                task_dict.update({'client_name':'unassigned'})
                task_dict.update({'client_id':'unassigned'})
            if 'task_result' in task_data.keys():
                for key in ['submission_type', 'gta_result_key', 'started_date', 'completed_date', 'status', 'gta_status', 'primary_reason', 'publishing_dumps']:
                    if key == 'gta_result_key':
                        task_dict.update({'phase':get_task_phase(task_data['task_result']['gta_result_key'])})
                        task_dict.update({key:task_data['task_result'][key]})
                    else:
                        task_dict.update({key:task_data['task_result'][key]})
            if 'task_definition' in task_data.keys():
                task_dict.update({'command':task_data['task_definition']['command']})
            session_task_dict.update({task_data['id']:task_dict})
    return session_task_dict

def get_csq_dict(session_jobs_dict, gtax_instance):
    csq_dict = dict()
    client_group_set = set()
    for job in session_jobs_dict:
        if session_jobs_dict[job]['result'] != 'blocked':
            client_group_set.add(session_jobs_dict[job]['client_group'])
    for group in client_group_set:
        csq_group_set = set()
        csq_group_jobs_list = list()
        for job in filer_report_data(session_jobs_dict, [['submission_type', 'init', 'match']]):
            if session_jobs_dict[job]['client_group'] == group:
                csq_group_set.add(session_jobs_dict[job]['csq'])
        index = 0
        for csq in csq_group_set:
            csq_dict.update({f'{group}_{index}':{'client_group':group, 'csq':csq, 'csq_link':get_csq_link(gtax_instance, csq)}})
            index += 1
    csq_group_set = set()
    for job in filer_report_data(session_jobs_dict, [['result', 'blocked', 'match']]):
        csq_group_set.add(session_jobs_dict[job]['csq'])
    index = 0
    for csq in csq_group_set:
        csq_dict.update({f'blocked_{index}':{'client_group':'blocked', 'csq':csq, 'csq_link':get_csq_link(gtax_instance, csq)}})
        index += 1
    return csq_dict

def get_csq_link(gtax_instance, csq):
    csq_link = f'http://{gtax_instance}.intel.com/#/clients?csq='
    csq_html = urllib.parse.quote(csq).replace("%27","'").replace('%28','(').replace('%29',')')
    csq_link += f'{csq_html}&ensure_req_match=true'
    return csq_link

def get_client_group(client_name):
    client_regex = re.compile(r'^.+?(-\d)') 
    client_group = client_name
    if client_regex.search(client_name):
        client_group = client_regex.search(client_name).group(0)[:-2]
    return client_group

def get_client_platform(csq):
    platform_regex = re.compile(r"('platform' = '.*?')") 
    platform = 'unknown'
    if platform_regex.search(csq):
        platform = platform_regex.search(csq).group(0)[14:-1]
    return platform

def get_client_pool(csq):
    platform_regex = re.compile(r"('pool' = '.*?')") 
    platform = 'unknown'
    if platform_regex.search(csq):
        platform = platform_regex.search(csq).group(0)[10:-1]
    return platform

def get_task_phase(gta_result_key):
    phase = 'unknown'
    if gta_result_key: 
        for phase_name in ['teardown', 'setup', 'test']:
            if gta_result_key.find(phase_name) > 0:
                phase = phase_name
    else:
        phase = ''
    return phase

def get_key_stat(data_dict, key, percent=False, total=False, limit=0):
    # format condition_list : [[condition_key1, condition_key1_value, condition_key1_func], [condition_key2, condition_key2_value, condition_key2_func]]
    # func: eql, n_eql, greater, greater_eql, less, less_eql, match, contain 
    value_stat_dict = dict()
    value_list = list()
    for item in data_dict:
        value_list.append(data_dict[item][key])
    value_set = set(value_list)
    for value in value_set:
        if value_list.count(value) > limit:
            if percent:
                value_stat_dict.update({value : round(value_list.count(value)/len(value_list)*100, 2)})
            else:
                value_stat_dict.update({value : value_list.count(value)})
    if total:
        value_stat_dict.update({'total':len(value_list)})
    return value_stat_dict

def get_tasks_stat(data_dict, percent=False, total=False):
    stat_dict = {'total': 0,  'not_run': 0, 'running': 0, 'passed': 0, 'failed': 0, 'aborted': 0, 'unsupported': 0, 'timed_out': 0, 'canceled': 0, 'blocked': 0, 'ignored': 0}
    for item in data_dict:
        for key in stat_dict.keys():
            stat_dict.update({key:stat_dict[key] + data_dict[item][f'{key}']})
    if percent:
        temp_stat_dict = stat_dict
        for key in temp_stat_dict.keys():
            total = stat_dict['total']
            stat_dict.update({key:round(stat_dict[key]/total*100, 2)})
    if not total:
        stat_dict.pop('total')
    return stat_dict

def filer_report_data(report_data, condition_list, match_all=True):
    # format condition_list : [[condition_key1, condition_key1_value, condition_key1_func], [condition_key2, condition_key2_value, condition_key2_func]]
    # func: eql, n_eql, greater, greater_eql, less, less_eql, match, contain  
    report_data_filtered = dict()
    if match_all:
        for bug in report_data:
            condition_match = True
            for condition in condition_list:
                if check_condition(report_data[bug][condition[0]], condition[1], condition[2]) and condition_match == True:
                    condition_match = True
                else:
                    condition_match = False
            if condition_match == True:
                report_data_filtered.update({bug:report_data[bug]})
    else:
        for bug in report_data:
            condition_match = False
            for condition in condition_list:
                if check_condition(report_data[bug][condition[0]], condition[1], condition[2]):
                    condition_match = True
            if condition_match == True:
                report_data_filtered.update({bug:report_data[bug]})        
    return report_data_filtered

def check_condition(value, condition_value, condition_func):
    check = False
    if condition_func == 'eql':
        if float(value) == float(condition_value):
            check = True
    elif condition_func == 'n_eql':
        if float(value) != float(condition_value):
            check = True
    elif condition_func == 'greater_eql':
        if float(value) >= float(condition_value):
            check = True
    elif condition_func == 'greater':
        if float(value) > float(condition_value):
            check = True
    elif condition_func == 'less':
        if float(value) < float(condition_value):
            check = True
    elif condition_func == 'less_eql':
        if float(value) <= float(condition_value):
            check = True
    elif condition_func == 'match':
        if str(value) == str(condition_value):
            check = True
    elif condition_func == 'not_match':
        if str(value) != str(condition_value):
            check = True
    elif condition_func == 'contain':
        if str(value).find(str(condition_value)) >= 0:
            check = True
    else:
        check = False
    return check

def print_stat_table(ws, title, stat_dict, start_row, start_column, value_sort=False, no_sort=False, horizontal=False, total=-1):
    row = start_row
    column = start_column
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws.cell(row=row, column=column, value=title)
    ws.cell(row=row, column=column).font = Font(bold=True)
    if total > -1:
        stat_dict.update({'total':total})
    if horizontal:
        table_row = row + 1
        column = start_column
        for stat in sort_dict(stat_dict, value_sort):
            row = table_row
            ws.cell(row=row, column=column, value=stat[0])
            ws.cell(row=row, column=column).border = thin_border
            row += 1
            ws.cell(row=row, column=column, value=stat[1])
            ws.cell(row=row, column=column).border = thin_border
            column += 1
    else:
        for stat in sort_dict(stat_dict, value_sort):
            row += 1
            column = start_column
            ws.cell(row=row, column=column, value=stat[0])
            ws.cell(row=row, column=column).border = thin_border
            column += 1
            ws.cell(row=row, column=column, value=stat[1])
            ws.cell(row=row, column=column).border = thin_border
    return row

def sort_dict(stat_dict, value_sort=False):
    stat_sort_list = []
    for stat in stat_dict:
        stat_sort_list.append([stat, stat_dict[stat]])
    if value_sort:
        stat_sort_list.sort(key=itemgetter(1), reverse=True)
    else:
        stat_sort_list.sort(key=itemgetter(0), reverse=False)
    return stat_sort_list

def print_pie_graph(ws, title, stat_dict, start_row, start_column, graph_poz, graph_w, graph_h, value_sort=False):
    chart = PieChart()
    chart.varyColors = True
    chart.title = title
    chart.height = graph_h
    chart.width = graph_w
    row = print_stat_table(ws, title, stat_dict, start_row, start_column, value_sort)
    chart_data = Reference(ws, min_row=start_row, max_row=row, min_col=start_column+1)
    chart_titles = Reference(ws, min_row=start_row+1, max_row=row, min_col=start_column)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.add_data(chart_data, titles_from_data=True)
    chart.set_categories(chart_titles)
    ws.add_chart(chart, graph_poz)
    return row

def print_col_graph(ws, title, stat_dict, start_row, start_column, graph_poz, graph_w, graph_h, value_sort=False):
    chart = BarChart()
    chart.style = 11
    chart.shape = 4
    chart.type = "col"
    chart.legend = None
    chart.varyColors = True
    chart.title = title
    chart.height = graph_h
    chart.width = graph_w
    row = print_stat_table(ws, title, stat_dict, start_row, start_column, value_sort)
    chart_data = Reference(ws, min_row=start_row+1, max_row=row, min_col=start_column+1)
    chart_titles = Reference(ws, min_row=start_row+1, max_row=row, min_col=start_column)
    chart.add_data(chart_data, titles_from_data=False)
    chart.set_categories(chart_titles)
    ws.add_chart(chart, graph_poz)
    return row

def print_bar_graph(ws, title, stat_dict, start_row, start_column, graph_poz, graph_w, graph_h, value_sort=False):
    chart = BarChart()
    chart.style = 11
    chart.shape = 4
    chart.type = "bar"
    chart.legend = None
    chart.varyColors = True
    chart.title = title
    chart.height = graph_h
    chart.width = graph_w
    chart.x_axis.scaling.orientation = "maxMin"
    row = print_stat_table(ws, title, stat_dict, start_row, start_column, value_sort)
    chart_data = Reference(ws, min_row=start_row+1, max_row=row, min_col=start_column+1)
    chart_titles = Reference(ws, min_row=start_row+1, max_row=row, min_col=start_column)
    chart.add_data(chart_data, titles_from_data=False)
    chart.set_categories(chart_titles)
    ws.add_chart(chart, graph_poz)
    return row

def get_letter_index_list():
    letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
    letter_index = ['']
    for letter in letters:
        letter_index.append(letter)
    for letter1 in letters:
        for letter2 in letters:
            letter_index.append(f'{letter1}{letter2}')
    return letter_index

def set_width(ws, min_width=0, max_width=100):
    column_widths = []
    column_heights = []
    for row_cnt, row in enumerate(ws.iter_rows()):
        column_heights.append(0)
        for col_cnt, cell in enumerate(row):
            if len(column_widths) <= col_cnt:
                column_widths.append(min_width)
            # Ignore cells that don't wrap
            if cell.alignment.wrap_text is False:
                continue
            if is_merged_cell(ws, cell):
                continue
            length, num_lines = calculate_cell_length(cell)
            if cell.alignment.text_rotation == 0:
                column_widths[col_cnt] = max(column_widths[col_cnt], length)
                row_height = 14 * num_lines
                column_heights[row_cnt] = max(column_heights[row_cnt], row_height)
            else:
                column_heights[row_cnt] = max(column_heights[row_cnt], 5 * length)
                col_width = 14 * num_lines
                column_widths[col_cnt] = max(column_widths[col_cnt], col_width)
    max_column_width_allowed = max_width
    for cnt, column_width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(cnt)].width = min(column_width, max_column_width_allowed)
    for cnt, row_height in enumerate(column_heights, 1):
        ws.row_dimensions[cnt].height = row_height

def is_merged_cell(ws, cell):
    for cell_range in ws.merged_cells:
        if cell.column >= cell_range.min_col and cell.column <= cell_range.max_col and \
           cell.row    >= cell_range.min_row and cell.row    <= cell_range.max_row:
            return True
    return False

def calculate_cell_length(cell):
    num_lines = 1
    if isinstance(cell.value, str):
        if 'HYPERLINK' in cell.value:
            pieces = cell.value.split(',')
            val = pieces[-1].split('"')[1]
            length = len(val)
        else:
            if '\n' in cell.value:
                num_lines = len(cell.value.split('\n'))
                length = max([len(s) for s in cell.value.split('\n')])
            else:
                length = len(cell.value)
    elif isinstance(cell.value, bool):
        length = len('False')
    elif isinstance(cell.value, int):
        if cell.value == 0:
            length = 1
        else:
            if cell.value < 0:
                val = -1 * cell.value
            else:
                val = cell.value
            length = math.ceil(math.log10(abs(cell.value)))
    else:
        length = 0
    length += 2
    return length, num_lines

if __name__ == '__main__':
    usage_msg = '''get_jobs_for_jobsetsession.exe -j 7178576
       get_jobs_for_jobsetsession.exe -j 7178576 -i gtax-gcmxd-fm
       get_jobs_for_jobsetsession.exe -j 7178576 -i gtax-gcmxd-fm -t
       get_jobs_for_jobsetsession.exe -j 7178576 -i gtax-gcmxd-fm -o
       '''
    ap = argparse.ArgumentParser(usage=usage_msg)
    ap.add_argument('-i', '-instance', default='gtax-igk', help='gtax instance:[gtax-igk, gtax-gcmxd-fm]', required=False)
    ap.add_argument('-j', '-job_session', help='job session id: 7178576', required=True)
    ap.add_argument('-c', '-client',  default=None, help='client search string exp:GK_DG2%%', required=False)
    ap.add_argument('-o', '-offline', action='store_true', default=False, help='offline mode', required=False)
    ap.add_argument('-t', '-task', action='store_true', default=False, help='include tasks data', required=False)
    parsed = ap.parse_args()
    #print(parsed)
    if parsed.o:
        print('------------ offline mode! ---------------')
    if parsed.j:
        main(parsed.i, parsed.j, parsed.c, parsed.o, parsed.t)
