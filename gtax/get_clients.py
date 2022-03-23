import sys
import os.path
import argparse
from datetime import datetime
from operator import itemgetter
import time
import json
import requests
import openpyxl
import math
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList

def main(client_search_string, columns_list_string, gtax_url, query_mode, columns_mode, filename, filter_offline, count_only):
    clients_data = list()
    if query_mode == 'name':
        clients_data = get_clients_data(gtax_url, client_search_string, False, filter_offline)
    if query_mode == 'csq':
        clients_data = get_clients_data(gtax_url, client_search_string, True, filter_offline)
    if not count_only:
        clients_data_list = clients_data_2_list_dict(clients_data)
        save_report(clients_data_list, columns_list_string, gtax_url, filename, columns_mode)
    return 0

def get_clients_data(gtax_url, client_search_string, is_csq, filter_offline):
    clients_data = list()
    if client_search_string:
        print(f'getting data for clients {client_search_string} from {gtax_url}')
        if is_csq:
            client_search_string = client_search_string.replace(' ','%20')
            client_search_string = client_search_string.replace('=','%3D')
            url = f'{gtax_url}/api/v1/clients?csq={client_search_string}&include_all_properties=true&full_info=true&order_by=name'
        else:
            url = f'{gtax_url}/api/v1/clients?include_all_properties=true&full_info=true&name={client_search_string}&order_by=name'
        data = get_gtax_data(url)
        if data:
            for client in data['data']:
                if filter_offline:
                    if filter_offline == 'skip' and client['status'] != 'offline':
                            clients_data.append(client)
                    if filter_offline == 'only' and client['status'] == 'offline':
                            clients_data.append(client)
                else:
                    clients_data.append(client)
            print(f'clients found: {len(clients_data)} {offline_msg(filter_offline)}')
        else:
            print(f'no clients found!')
    return clients_data

def offline_msg(filter_offline):
    offline_text = ''
    if filter_offline:
        if filter_offline == 'skip':
            offline_text = '[offline clients skipped!]'
        if filter_offline == 'only':
            offline_text = '[offline clients only!]'
    return offline_text

def get_gtax_instance_url(gtax_instance, gtax_port):
    gtax_instance_url = f'http://{gtax_instance}.intel.com'
    if len(gtax_port) > 1:
        gtax_instance_url += f':{gtax_port}'
    return gtax_instance_url

def get_gtax_data(url):
    gtax_data = None
    headers = { 'Content-type': 'application/json' }
    response = requests.get(url, headers=headers, proxies={'http': 'http://proxy-chain.intel.com:911', 'https': 'http://proxy-chain.intel.com:912' })
    if response.status_code == 200:
        gtax_data = response.json()
    else:
        print(url)
        print(f'status code:{response.status_code}')
        print(response.text)
    return gtax_data
 
def get_client_property_value(property_name, client_data):
    client_property_value = None
    for client_property in client_data['properties']:
        if client_property['property']['name'] == property_name:
            client_property_value = client_property['property']['value']
            break
    return client_property_value

def clients_data_2_list_dict(clients_data):
    clients_data_list = list()
    if clients_data:
        for client in clients_data:
            client_property_list = list()
            for key in ['name', 'id', 'status', 'active_runner', 'is_reserved', 'is_dirty']:
                client_property_dict = dict()
                client_property_dict.update({'name':f'client_{key}'})
                client_property_dict.update({'value':client[key]})
                client_property_dict.update({'type':'system'})
                client_property_dict.update({'active': True})
                client_property_dict.update({'required':False})
                client_property_list.append(client_property_dict)
            for client_property in client['properties']:
                client_property_dict = dict()
                for key in ['name', 'value']:
                    client_property_dict.update({key:client_property['property'][key]})
                for key in ['type', 'active', 'required']:
                    client_property_dict.update({key:client_property[key]})
                client_property_list.append(client_property_dict)
            for setting in client['settings']:
                client_setting_dict = dict()
                client_setting_dict.update({'name':f"[{setting['name']}]"})
                client_setting_dict.update({'value':setting['value']})
                client_setting_dict.update({'type':'setting'})
                client_setting_dict.update({'active':'True'})
                client_setting_dict.update({'required':'False'})
                client_property_list.append(client_setting_dict)
            clients_data_list.append(client_property_list)
    return clients_data_list

def get_all_clients_data_keys(clients_data_list, select=all):
    keys_set = set()
    all_keys = list()
    for client_data in clients_data_list:
        for i in range(len(client_data)):
            if select == 'all_auto' and client_data[i]['type'] == 'auto_detected':
                keys_set.add(client_data[i]['name'])
            elif select == 'all_user' and client_data[i]['type'] == 'user_defined':
                keys_set.add(client_data[i]['name'])
            elif select == 'all_system' and client_data[i]['type'] == 'system':
                keys_set.add(client_data[i]['name'])
            elif select == 'all_required' and client_data[i]['required'] == True:
                keys_set.add(client_data[i]['name'])
            elif select == 'all' and client_data[i]['type'] != 'setting':
                keys_set.add(client_data[i]['name']) 
            elif select == 'all_settings' and client_data[i]['type'] == 'setting':
                keys_set.add(client_data[i]['name']) 
    for key in ['client_name', 'client_id', 'client_status', 'client_active_runner', 'client_is_reserved', 'client_is_dirty']:
        if key in keys_set:
            keys_set.remove(key)
            all_keys.append(key)
    for key in sorted(keys_set):
        all_keys.append(key)
    return all_keys

def get_client_property(client_data, property_name, property_key):
    user_present = False
    property_value = ''
    #check user first
    for i in range(len(client_data)):
        if property_name == client_data[i]['name'] and client_data[i]['type'] == 'user_defined':
            property_value = client_data[i][property_key]
            user_present = True
            break
    #then check auto if not user
    if not user_present:
        for i in range(len(client_data)):
            if property_name == client_data[i]['name']:
                property_value = client_data[i][property_key]
                break
    return property_value

def print_clients_data_table(ws, clients_data_list, columns_list_string, gtax_url, start_row, start_column):
    columns_list = columns_list_string.split(',')
    if columns_list[0] in ['all', 'all_auto', 'all_user', 'all_system', 'all_required', 'all_settings']:
        data_keys = get_all_clients_data_keys(clients_data_list, select=columns_list[0])
    else:
        data_keys = columns_list
    if 'client_name' not in data_keys:
        data_keys.insert(0, 'client_name') 
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    row = start_row
    column = start_column
    for key in data_keys:
        ws.cell(row=row, column=column, value=key)
        ws.cell(row=row, column=column).font = Font(bold = True)
        column += 1
    for client_data in clients_data_list:
        client_id = get_client_property(client_data, 'client_id', 'value')
        row += 1
        column = start_column
        for key in data_keys:
            if key in ['client_name', 'client_id']:
                link_text = get_client_property(client_data,  key, 'value')
                client_id = get_client_property(client_data,  'client_id', 'value')
                ws.cell(row=row, column=column, value=f'=HYPERLINK("{gtax_url}/#/clients/{client_id}?tab=properties","{link_text}")')
                ws.cell(row=row, column=column).font = Font(color='0000FF')
            else:
                if get_client_property(client_data,  key, 'required') == True:
                    ws.cell(row=row, column=column).font = Font(bold=True)
                    ws.cell(row=row, column=column).fill = PatternFill("solid", fgColor="FFFF00")
                if get_client_property(client_data,  key, 'active') == False:
                    ws.cell(row=row, column=column).font = Font(color='FFFFFF')
                    ws.cell(row=row, column=column).fill = PatternFill("solid", fgColor="AAAAAA")
                if get_client_property(client_data,  key, 'type') == 'auto_detected':
                    ws.cell(row=row, column=column).font = Font(italic=True)
                    ws.cell(row=row, column=column).fill = PatternFill("solid", fgColor="EEEEEE")
                ws.cell(row=row, column=column, value=get_client_property(client_data,  key, 'value'))
            ws.cell(row=row, column=column).border = thin_border
            column += 1
    return row

def save_report(clients_data_list, columns_list, gtax_url, filename, columns_mode):
    if clients_data_list:
        if columns_mode == 'columns':
            output_wb = openpyxl.Workbook()
            ws = output_wb.active
            ws.title = f'gtax_clients-{ts}'
            print_clients_data_table(ws, clients_data_list, columns_list, gtax_url, 1, 1)
            ws.auto_filter.ref = ws.dimensions
            set_width(ws)
            output_wb.save(filename)
        else:
            # tabs mode
            output_wb = openpyxl.Workbook()
            tabs_count = 0
            tabs = get_tabs_from_file(columns_list)
            for tab in tabs:
                if tabs_count == 0:
                    ws = output_wb.active
                    ws.title =  tab['name']
                else:
                    ws = output_wb.create_sheet(tab['name'])
                print_clients_data_table(ws, clients_data_list, tab['columns'], gtax_url, 1, 1)
                ws.auto_filter.ref = ws.dimensions
                set_width(ws)
                tabs_count += 1
            output_wb.save(filename)
        print(f'output saved to {filename}')

def set_width(ws, min_width=0):
    column_widths = []
    column_heights = []
    for row_cnt, row in enumerate(ws.iter_rows()):
        column_heights.append(0)
        for col_cnt, cell in enumerate(row):
            if len(column_widths) <= col_cnt:
                column_widths.append(min_width)
            # Ignore cells that don't wrap
            if cell.alignment.wrap_text is False:
                continue
            if is_merged_cell(ws, cell):
                continue
            length, num_lines = calculate_cell_length(cell)
            if cell.alignment.text_rotation == 0:
                column_widths[col_cnt] = max(column_widths[col_cnt], length)
                row_height = 14 * num_lines
                column_heights[row_cnt] = max(column_heights[row_cnt], row_height)
            else:
                column_heights[row_cnt] = max(column_heights[row_cnt], 5 * length)
                col_width = 14 * num_lines
                column_widths[col_cnt] = max(column_widths[col_cnt], col_width)
    max_column_width_allowed = 60
    for cnt, column_width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(cnt)].width = min(column_width, max_column_width_allowed)
    for cnt, row_height in enumerate(column_heights, 1):
        ws.row_dimensions[cnt].height = row_height

def is_merged_cell(ws, cell):
    for cell_range in ws.merged_cells:
        if cell.column >= cell_range.min_col and cell.column <= cell_range.max_col and \
           cell.row    >= cell_range.min_row and cell.row    <= cell_range.max_row:
            return True
    return False

def calculate_cell_length(cell):
    num_lines = 1
    if isinstance(cell.value, str):
        if 'HYPERLINK' in cell.value:
            pieces = cell.value.split(',')
            val = pieces[-1].split('"')[1]
            length = len(val)
        else:
            if '\n' in cell.value:
                num_lines = len(cell.value.split('\n'))
                length = max([len(s) for s in cell.value.split('\n')])
            else:
                length = len(cell.value)
    elif isinstance(cell.value, bool):
        length = len('False')
    elif isinstance(cell.value, int):
        if cell.value == 0:
            length = 1
        else:
            if cell.value < 0:
                val = -1 * cell.value
            else:
                val = cell.value
            length = math.ceil(math.log10(abs(cell.value)))
    else:
        length = 0
    length += 2
    return length, num_lines

def get_tabs_from_file(file_name):
    with open(file_name) as json_file:
        data = json.load(json_file)['tabs']
    return data

def get_filename_timestamp():
    ts = str(datetime.now())
    ts = ts.replace(':', '_')
    return ts[5:-7]

def print_tabs_config_template():
    template_file = 'tabs_template.json'
    tabs_template = {'tabs':[{'name': 'clients', 'columns': 'client_name,client_id,label_id,client_status,pool,board_name,platform'}, {'name': 'displays', 'columns': 'display_external,display_hdmi,display_dp,display_edp,display_raritan,display_id,display_dp_id,display_hdmi_id,display_raritan_id,display_raritan_connected_port,kvm'}, {'name': 'user', 'columns': 'all_user'}, {'name': 'required', 'columns': 'all_required'}, {'name': 'settings', 'columns': 'all_settings'}]}
    with open(template_file, 'w') as outfile:
        json.dump(tabs_template, outfile, indent=4)
    print(json.dumps(tabs_template, indent=4))
    print(f'template saved:{template_file}')

def log_script_call():
    app_name = os.path.basename(sys.argv[0])
    app_path = os.path.dirname(os.path.abspath(sys.argv[0]))
    log_path = os.path.join(app_path, 'logs')
    params_list = sys.argv[1:]
    call_log_file = os.path.join(log_path, f"{app_name.split('.')[0]}_calls.log")
    command = f'{app_name}'
    for param in params_list:
        if param.find(chr(39)) > 0 or param.find(chr(32)) > 0:
            command += f' "{param}"'
        else:
            command += f' {param}'
    if not os.path.exists(log_path):
        os.makedirs(log_path)
    with open(call_log_file, 'a') as log_file:
        log_file.write(f'[{str(datetime.now())}] {command}\n')

if __name__ == '__main__':
    usage_msg = '''get_clients.exe -n GK-RKL%% -c client_name,label_id,client_status,pool,platform,board_name,cpu_sku,cpu_stepping,cpu_qdf,display_hdmi,display_hdmi_id,sbios,bios_me_firmware_version,owner,netbox_id
       get_clients.exe -n FM-RKL%% -c all
       get_clients.exe -n FM-RKL%% -c all_auto
       get_clients.exe -n FM-RKL%% -c all_user
       get_clients.exe -n FM-RKL%% -c all_required
       get_clients.exe -n FM-RKL%% -i gtax-gcmxd-fm
       get_clients.exe -n FM-RKL%% -c all_user -i gtax-gcmxd-fm'''
    ts = get_filename_timestamp()
    filename = f'gtax_clients_{ts}.xlsx'
    ap = argparse.ArgumentParser(usage=usage_msg)
    ap.add_argument('-i', '-instance', default='gtax-igk', help='gtax instance:[gtax-igk, gtax-gcmxd-fm, gtax-ril-fm]', required=False)
    ap.add_argument('-p', '-port', default='', help='gtax instance port', required=False)
    ap.add_argument('-n', '-name', help="search clients by name: GK-RKL%%", required=False)
    ap.add_argument('-csq', help="csq query: ('pool' = 'CI_RKL')", required=False) 
    ap.add_argument('-c', '-columns', default='all', help='columns list: [all/all_auto/all_user/all_settings/client_name,label_id,client_status,pool]', required=False)
    ap.add_argument('-t', '-tabs', help='tabs_config_file.json', required=False)
    ap.add_argument('-o', '-output', default=filename, help='output file name', required=False)
    ap.add_argument('-offline', default=False, const='all', nargs='?', choices=['skip', 'only'], help='offline clients skip or only', required=False)
    ap.add_argument('-tt', '-tabs_template', action='store_true', default=False, help='show tabs config template file', required=False)
    ap.add_argument('-count', action='store_true', default=False, help='print count only', required=False)

    parsed = ap.parse_args()
    gtax_url = get_gtax_instance_url(parsed.i, parsed.p)
    log_script_call()
    if not (parsed.n or parsed.csq or parsed.tt):
        ap.error('\nNo params! Use -n or -csq or -tt')
    else:
        search_string = ''
        query_mode = 'name'
        cloumns_string = ''
        if parsed.n: 
            query_mode = 'name'
            search_string = parsed.n
        if parsed.csq: 
            query_mode = 'csq'
            search_string = parsed.csq
        if parsed.tt:
            print_tabs_config_template()
        else:
            if parsed.t:
                main(search_string, parsed.t, gtax_url, query_mode, 'tabs', parsed.o, parsed.offline, False)
            else:
                main(search_string, parsed.c, gtax_url, query_mode, 'columns', parsed.o, parsed.offline, parsed.count)